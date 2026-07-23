"""
routes/dept_admin.py — Department Admin (HOD) blueprint.
Manages everything within the HOD's assigned department only.
"""

from flask import (Blueprint, render_template, request,
                   redirect, url_for, flash, abort, make_response, jsonify)
from auth_utils import (dept_admin_required, write_audit_log,
                        current_user, dept_isolation_check)
from db import get_service_client
from notifications import get_user_notifications, notify_dept_notice, create_notification, delete_notifications_for_notice
from report_utils import (excel_letterhead, style_header_row,
                          excel_signature_block, pdf_header_style_cmds,
                          pdf_signature_block)
from datetime import datetime, date, timedelta
import secrets, string, json

dept_admin_bp = Blueprint("dept_admin", __name__)


def _dept_id():
    user = current_user()
    dept = user.get("department_id")
    if not dept:
        abort(403)
    return dept


def _dept_student_ids(db, dept_id):
    enrolled = (db.table("enrollments")
                  .select("student_id, classes!inner(department_id)")
                  .eq("classes.department_id", dept_id)
                  .execute().data or [])
    return list({row["student_id"] for row in enrolled if row.get("student_id")})


def _student_class_name(db, student_id):
    """Return enrolled class name for a student."""
    rows = (db.table("enrollments")
              .select("classes(name)")
              .eq("student_id", student_id)
              .limit(1).execute().data or [])
    if not rows:
        return "—"
    cls = rows[0].get("classes")
    if isinstance(cls, list):
        cls = cls[0] if cls else {}
    return (cls or {}).get("name") or "—"


def _student_in_dept(db, student_id, dept_id):
    """True if student is enrolled in a class belonging to dept_id."""
    rows = (db.table("enrollments")
              .select("student_id, classes!inner(department_id)")
              .eq("student_id", student_id)
              .eq("classes.department_id", dept_id)
              .limit(1).execute().data or [])
    return bool(rows)


def _attach_student_classes(db, students):
    """Attach class_obj and admission_number on each student dict."""
    if not students:
        return students
    ids = [s["id"] for s in students]
    enr_rows = (db.table("enrollments")
                  .select("student_id, classes(id, name)")
                  .in_("student_id", ids)
                  .execute().data or [])
    class_by_sid = {}
    for e in enr_rows:
        sid = e["student_id"]
        if sid in class_by_sid:
            continue
        cls = e.get("classes")
        if isinstance(cls, list):
            cls = cls[0] if cls else None
        if cls:
            class_by_sid[sid] = cls
    for s in students:
        s["class_obj"] = class_by_sid.get(s["id"])
        s["admission_number"] = s.get("admission_no", "")
    return students


def _gen_password(length=10):
    chars = string.ascii_letters + string.digits + "!@#$"
    while True:
        pwd = "".join(secrets.choice(chars) for _ in range(length))
        if any(c.isdigit() for c in pwd) and any(c in "!@#$" for c in pwd):
            return pwd


# ── Dashboard ─────────────────────────────────────────────────────────────────

def _dept_dashboard_payload(db, dept_id):
    """Compute all dept-admin dashboard KPIs and chart datasets.

    Returns a dict of native Python values (no json.dumps) so it can be
    reused by both the HTML view and the realtime JSON endpoint.
    """
    stats = {}
    payload = {
        "stats": stats,
        "att_unit_labels": [], "att_unit_present": [], "att_unit_absent": [],
        "trend_labels": [], "trend_present": [], "trend_absent": [],
        "app_status": {"pending": 0, "approved": 0, "rejected": 0},
        "clearance_stats": {"pending": 0, "approved": 0, "rejected": 0},
        "attachment_stats": {"pending": 0, "active": 0, "approved": 0, "completed": 0, "rejected": 0},
        "class_labels": [], "class_counts": [],
        "atype_labels": [], "atype_counts": [],
    }
    try:
        from stats_utils import (exact_count, clearance_kpi, attachment_status_counts,
                                  count_table, count_status_map)

        stats["classes"]  = db.table("classes").select("id", count="exact").eq("department_id", dept_id).execute().count or 0
        stats["trainers"] = db.table("user_profiles").select("id", count="exact").eq("role", "trainer").eq("department_id", dept_id).execute().count or 0
        stats["students"] = db.table("user_profiles").select("id", count="exact").eq("role", "student").eq("department_id", dept_id).execute().count or 0
        stats["units"]    = db.table("units").select("id", count="exact").eq("department_id", dept_id).execute().count or 0

        stats["assessments"] = exact_count(
            db.table("assessments").select("id, units!inner(department_id)", count="exact")
            .eq("units.department_id", dept_id))
        stats["pending"] = exact_count(
            db.table("assessments").select("id, units!inner(department_id)", count="exact")
            .eq("units.department_id", dept_id).eq("status", "pending"))
        stats["approved"] = exact_count(
            db.table("assessments").select("id, units!inner(department_id)", count="exact")
            .eq("units.department_id", dept_id).eq("status", "approved"))
        stats["rejected"] = exact_count(
            db.table("assessments").select("id, units!inner(department_id)", count="exact")
            .eq("units.department_id", dept_id).eq("status", "rejected"))

        try:
            stats["applications"] = db.table("course_applications").select("id", count="exact").eq("department_id", dept_id).execute().count or 0
            stats["pending_applications"] = db.table("course_applications").select("id", count="exact").eq("department_id", dept_id).eq("status", "pending").execute().count or 0
        except Exception:
            stats["applications"] = 0
            stats["pending_applications"] = 0

        # Attendance per unit
        all_att = (db.table("attendance")
                   .select("status, units!inner(name, department_id)")
                   .eq("units.department_id", dept_id).execute().data or [])
        att_per_unit = {}
        for row in all_att:
            uname = (row.get("units") or {}).get("name", "Unknown")
            bucket = att_per_unit.setdefault(uname, {"present": 0, "absent": 0})
            if row.get("status") == "present":
                bucket["present"] += 1
            else:
                bucket["absent"] += 1
        payload["att_unit_labels"] = list(att_per_unit.keys())
        payload["att_unit_present"] = [att_per_unit[u]["present"] for u in payload["att_unit_labels"]]
        payload["att_unit_absent"] = [att_per_unit[u]["absent"] for u in payload["att_unit_labels"]]

        # 7-day attendance trend
        for i in range(6, -1, -1):
            day = (date.today() - timedelta(days=i)).isoformat()
            payload["trend_labels"].append(day[5:])
            day_rows = (db.table("attendance")
                        .select("status, units!inner(department_id)")
                        .eq("units.department_id", dept_id)
                        .eq("attendance_date", day).execute().data or [])
            payload["trend_present"].append(sum(1 for r in day_rows if r.get("status") == "present"))
            payload["trend_absent"].append(sum(1 for r in day_rows if r.get("status") != "present"))

        # Application status breakdown
        app_status = count_status_map(db, "course_applications",
                                      ("pending", "approved", "rejected"), department_id=dept_id)
        for k in ("pending", "approved", "rejected"):
            app_status.setdefault(k, 0)
        payload["app_status"] = app_status

        # Clearance requests
        try:
            cl = clearance_kpi(db, department_id=dept_id)
            payload["clearance_stats"] = {"pending": cl["pending"], "approved": cl["completed"], "rejected": cl["rejected"]}
        except Exception:
            pass

        # Industrial attachments
        try:
            att_student_ids = [u["id"] for u in (db.table("user_profiles").select("id")
                               .eq("role", "student").eq("department_id", dept_id).execute().data or [])]
            payload["attachment_stats"] = attachment_status_counts(db, att_student_ids)
        except Exception:
            pass

        try:
            stats["trips_total"] = exact_count(db.table("academic_trips").select("id", count="exact").eq("department_id", dept_id))
            stats["trips_pending"] = exact_count(db.table("academic_trips").select("id", count="exact").eq("department_id", dept_id).eq("status", "submitted"))
        except Exception:
            stats["trips_total"] = stats["trips_pending"] = 0
        try:
            stats["summative_nyc"] = exact_count(db.table("summative_competences").select("id", count="exact").eq("department_id", dept_id).eq("competence", "not_yet_competent"))
        except Exception:
            stats["summative_nyc"] = 0

        # Classes with student counts
        classes_data = (db.table("classes").select("id, name, courses(name)")
                        .eq("department_id", dept_id).order("name").execute().data or [])
        for cls in classes_data[:10]:
            payload["class_labels"].append(cls["name"])
            payload["class_counts"].append(count_table(db, "enrollments", class_id=cls["id"]))

        # Assessment types breakdown
        typed_assessments = (db.table("assessments")
                             .select("assessment_type, units!inner(department_id)")
                             .eq("units.department_id", dept_id).execute().data or [])
        atype_map = {}
        for a in typed_assessments:
            t = a.get("assessment_type") or "Other"
            atype_map[t] = atype_map.get(t, 0) + 1
        payload["atype_labels"] = list(atype_map.keys())
        payload["atype_counts"] = [atype_map[t] for t in payload["atype_labels"]]
    except Exception as e:
        print(f"[dept_admin] dashboard payload error: {e}")
    return payload


@dept_admin_bp.route("/dashboard/live")
@dept_admin_required
def dashboard_live():
    """Realtime JSON feed for the dept-admin dashboard (polled by the client)."""
    from flask import jsonify
    db = get_service_client()
    payload = _dept_dashboard_payload(db, _dept_id())
    payload["server_time"] = datetime.now().strftime("%H:%M:%S")
    return jsonify(payload)


@dept_admin_bp.route("/")
@dept_admin_bp.route("/dashboard")
@dept_admin_required
def dashboard():
    db = get_service_client()
    dept_id = _dept_id()
    dept = db.table("departments").select("*").eq("id", dept_id).single().execute().data or {}
    stats = {}
    unread_notifications = []
    recent_assessments = []
    recent_attendance = []
    units_list = []
    try:
        stats["classes"]     = db.table("classes").select("id", count="exact").eq("department_id", dept_id).execute().count or 0
        stats["trainers"]    = db.table("user_profiles").select("id", count="exact").eq("role", "trainer").eq("department_id", dept_id).execute().count or 0
        stats["students"]    = db.table("user_profiles").select("id", count="exact").eq("role", "student").eq("department_id", dept_id).execute().count or 0
        stats["units"]       = db.table("units").select("id", count="exact").eq("department_id", dept_id).execute().count or 0
        
        # Fetch unread notifications
        unread_notifications = get_user_notifications(current_user()["id"], unread_only=True, limit=3)

        # Assessments stats — exact counts via unit join filters
        from stats_utils import exact_count, clearance_kpi, attachment_status_counts, count_table

        stats["assessments"] = exact_count(
            db.table("assessments").select("id, units!inner(department_id)", count="exact")
            .eq("units.department_id", dept_id)
        )
        stats["pending"] = exact_count(
            db.table("assessments").select("id, units!inner(department_id)", count="exact")
            .eq("units.department_id", dept_id).eq("status", "pending")
        )
        stats["approved"] = exact_count(
            db.table("assessments").select("id, units!inner(department_id)", count="exact")
            .eq("units.department_id", dept_id).eq("status", "approved")
        )
        stats["rejected"] = exact_count(
            db.table("assessments").select("id, units!inner(department_id)", count="exact")
            .eq("units.department_id", dept_id).eq("status", "rejected")
        )

        # Course applications
        try:
            stats["applications"] = db.table("course_applications").select("id", count="exact").eq("department_id", dept_id).execute().count or 0
            stats["pending_applications"] = db.table("course_applications").select("id", count="exact").eq("department_id", dept_id).eq("status", "pending").execute().count or 0
        except Exception:
            stats["applications"] = 0
            stats["pending_applications"] = 0

        # Recent assessments specifically for this department
        recent_assessments = (db.table("assessments")
            .select("*, user_profiles!assessments_student_id_fkey(full_name, admission_no), units!inner(name, department_id), classes(name)")
            .eq("units.department_id", dept_id)
            .order("uploaded_at", desc=True).limit(8).execute().data or [])

        # Recent attendance specifically for this department
        recent_attendance = (db.table("attendance")
            .select("*, user_profiles!attendance_student_id_fkey(full_name, admission_no, enrollments(classes(name))), units!inner(name, code, department_id)")
            .eq("units.department_id", dept_id)
            .order("attendance_date", desc=True).limit(10).execute().data or [])
            
        # Flatten recent_attendance class names
        for att in recent_attendance:
            student = att.get("user_profiles") or {}
            enrolls = student.get("enrollments") or []
            first_enroll = enrolls[0] if enrolls else {}
            cls = first_enroll.get("classes") or {}
            att["classes"] = cls
            
        units_list = db.table("units").select("id, name, code").eq("department_id", dept_id).order("name").execute().data or []

        # ── Analytics: Attendance per unit ───────────────────────
        all_att = (db.table("attendance")
                   .select("status, units!inner(name, department_id)")
                   .eq("units.department_id", dept_id)
                   .execute().data or [])
        att_per_unit = {}
        for row in all_att:
            uname = (row.get("units") or {}).get("name", "Unknown")
            bucket = att_per_unit.setdefault(uname, {"present": 0, "absent": 0})
            if row.get("status") == "present":
                bucket["present"] += 1
            else:
                bucket["absent"] += 1
        att_unit_labels = list(att_per_unit.keys())
        att_unit_present = [att_per_unit[u]["present"] for u in att_unit_labels]
        att_unit_absent  = [att_per_unit[u]["absent"]  for u in att_unit_labels]

        # ── Analytics: 7-day attendance trend ────────────────────
        trend_labels, trend_present, trend_absent = [], [], []
        for i in range(6, -1, -1):
            day = (date.today() - timedelta(days=i)).isoformat()
            trend_labels.append(day[5:])  # MM-DD
            day_rows = (db.table("attendance")
                        .select("status, units!inner(department_id)")
                        .eq("units.department_id", dept_id)
                        .eq("attendance_date", day)
                        .execute().data or [])
            trend_present.append(sum(1 for r in day_rows if r.get("status") == "present"))
            trend_absent.append(sum(1 for r in day_rows if r.get("status") != "present"))

        # ── Analytics: Application status breakdown ───────────────
        from stats_utils import count_status_map
        app_status = count_status_map(
            db, "course_applications", ("pending", "approved", "rejected"),
            department_id=dept_id,
        )
        for k in ("pending", "approved", "rejected"):
            app_status.setdefault(k, 0)

        # ── Analytics: Clearance requests (map DB statuses → UI labels) ───
        clearance_stats = {"pending": 0, "approved": 0, "rejected": 0}
        try:
            cl = clearance_kpi(db, department_id=dept_id)
            clearance_stats = {
                "pending": cl["pending"],
                "approved": cl["completed"],
                "rejected": cl["rejected"],
            }
        except Exception:
            pass

        # ── Analytics: Industrial attachments (all dept students, no 200 cap) ─
        attachment_stats = {"pending": 0, "active": 0, "approved": 0, "completed": 0, "rejected": 0}
        try:
            att_student_ids = [
                u["id"]
                for u in (
                    db.table("user_profiles")
                    .select("id")
                    .eq("role", "student")
                    .eq("department_id", dept_id)
                    .execute()
                    .data
                    or []
                )
            ]
            attachment_stats = attachment_status_counts(db, att_student_ids)
        except Exception:
            pass

        # Trips + summative for this department
        try:
            stats["trips_total"] = exact_count(
                db.table("academic_trips").select("id", count="exact").eq("department_id", dept_id)
            )
            stats["trips_pending"] = exact_count(
                db.table("academic_trips").select("id", count="exact")
                .eq("department_id", dept_id).eq("status", "submitted")
            )
        except Exception:
            stats["trips_total"] = stats["trips_pending"] = 0
        try:
            stats["summative_nyc"] = exact_count(
                db.table("summative_competences").select("id", count="exact")
                .eq("department_id", dept_id).eq("competence", "not_yet_competent")
            )
        except Exception:
            stats["summative_nyc"] = 0

        # ── Analytics: Classes with student counts ────────────────
        classes_data = (db.table("classes")
                        .select("id, name, courses(name)")
                        .eq("department_id", dept_id)
                        .order("name")
                        .execute().data or [])
        class_labels, class_counts = [], []
        for cls in classes_data[:10]:
            enroll_count = count_table(db, "enrollments", class_id=cls["id"])
            class_labels.append(cls["name"])
            class_counts.append(enroll_count)

        # ── Analytics: Assessment types breakdown ─────────────────
        typed_assessments = (db.table("assessments")
                             .select("assessment_type, units!inner(department_id)")
                             .eq("units.department_id", dept_id)
                             .execute().data or [])
        atype_map = {}
        for a in typed_assessments:
            t = a.get("assessment_type") or "Other"
            atype_map[t] = atype_map.get(t, 0) + 1
        atype_labels = list(atype_map.keys())
        atype_counts = [atype_map[t] for t in atype_labels]

    except Exception as e:
        flash(f"Error loading dashboard: {e}", "danger")
        att_unit_labels = att_unit_present = att_unit_absent = []
        trend_labels = trend_present = trend_absent = []
        app_status = {"pending": 0, "approved": 0, "rejected": 0}
        clearance_stats = {"pending": 0, "approved": 0, "rejected": 0}
        attachment_stats = {"pending": 0, "active": 0, "approved": 0, "completed": 0, "rejected": 0}
        class_labels = class_counts = []
        atype_labels = atype_counts = []

    return render_template("dept_admin/dashboard_enhanced.html",
                           dept=dept, stats=stats,
                           recent_assessments=recent_assessments,
                           recent_attendance=recent_attendance,
                           units_list=units_list,
                           unread_notifications=unread_notifications,
                           # analytics
                           att_unit_labels=json.dumps(att_unit_labels),
                           att_unit_present=json.dumps(att_unit_present),
                           att_unit_absent=json.dumps(att_unit_absent),
                           trend_labels=json.dumps(trend_labels),
                           trend_present=json.dumps(trend_present),
                           trend_absent=json.dumps(trend_absent),
                           app_status=app_status,
                           clearance_stats=clearance_stats,
                           attachment_stats=attachment_stats,
                           class_labels=json.dumps(class_labels),
                           class_counts=json.dumps(class_counts),
                           atype_labels=json.dumps(atype_labels),
                           atype_counts=json.dumps(atype_counts),
                           department_name=dept.get("name",""))


# ── Welcome alias ─────────────────────────────────────────────────────────────

@dept_admin_bp.route("/welcome")
@dept_admin_required
def welcome():
    return redirect(url_for("dept_admin.dashboard"))


# ── Courses ───────────────────────────────────────────────────────────────────

@dept_admin_bp.route("/courses", methods=["GET", "POST"])
@dept_admin_required
def courses():
    db = get_service_client()
    dept_id = _dept_id()
    error = None

    # Delete
    if request.args.get("delete"):
        try:
            course_id = request.args["delete"]
            # Verify course belongs to this department
            row = db.table("courses").select("department_id").eq("id", course_id).single().execute().data
            if not row or row.get("department_id") != dept_id:
                flash("Course not found or not in your department.", "danger")
            else:
                db.table("courses").delete().eq("id", course_id).execute()
                write_audit_log("delete_course", target=course_id)
                flash("Course deleted.", "success")
        except Exception as exc:
            error = f"Error deleting course: {exc}"
        return redirect(url_for("dept_admin.courses"))

    if request.method == "POST":
        action = request.form.get("action", "add")

        if action == "add":
            name = request.form.get("name", "").strip()
            code = request.form.get("code", "").strip().upper()
            if not name or not code:
                error = "Course name and code are required."
            else:
                try:
                    existing = db.table("courses").select("id").eq("code", code).eq("department_id", dept_id).execute()
                    if existing.data:
                        error = f"A course with code '{code}' already exists in this department."
                    else:
                        db.table("courses").insert({
                            "name": name,
                            "code": code,
                            "department_id": dept_id
                        }).execute()
                        write_audit_log("create_course", target=code)
                        flash(f"Course '{name}' added successfully.", "success")
                        return redirect(url_for("dept_admin.courses"))
                except Exception as exc:
                    error = f"Error adding course: {exc}"

        elif action == "edit":
            course_id = request.form.get("course_id", "").strip()
            name      = request.form.get("name", "").strip()
            code      = request.form.get("code", "").strip().upper()
            if not course_id or not name or not code:
                error = "All fields are required to update a course."
            else:
                try:
                    row = db.table("courses").select("department_id").eq("id", course_id).single().execute().data
                    if not row or row.get("department_id") != dept_id:
                        error = "Course not found in your department."
                    else:
                        db.table("courses").update({"name": name, "code": code}).eq("id", course_id).execute()
                        write_audit_log("edit_course", target=course_id)
                        flash(f"Course updated to '{name}'.", "success")
                        return redirect(url_for("dept_admin.courses"))
                except Exception as exc:
                    error = f"Error updating course: {exc}"

    courses_list = (db.table("courses")
                    .select("*, classes(id)")
                    .eq("department_id", dept_id)
                    .order("name")
                    .execute().data or [])
    # Annotate with class count
    for c in courses_list:
        c["_class_count"] = len(c.get("classes") or [])

    return render_template("dept_admin/courses.html",
                           courses=courses_list,
                           error=error)


# ── Classes ───────────────────────────────────────────────────────────────────

@dept_admin_bp.route("/classes", methods=["GET", "POST"])
@dept_admin_required
def classes():
    db = get_service_client()
    dept_id = _dept_id()
    error = None
    if request.method == "POST":
        action = request.form.get("action", "create")
        if action == "create":
            name         = request.form.get("name", "").strip().upper()
            course_id    = request.form.get("course_id", "").strip()
            intake_year  = request.form.get("intake_year") or None
            intake_month = request.form.get("intake_month") or None
            level        = request.form.get("level") or None
            cycle        = request.form.get("cycle") or None
            if not name or not course_id:
                error = "Class name and course are required."
            else:
                try:
                    db.table("classes").insert({
                        "name": name, "course_id": course_id,
                        "department_id": dept_id, "intake_year": intake_year,
                        "intake_month": intake_month, "level": level, "cycle": cycle
                    }).execute()
                    write_audit_log("create_class", target=name)
                    flash("Class added successfully.", "success")
                    return redirect(url_for("dept_admin.classes"))
                except Exception as exc:
                    error = f"Error creating class: {exc}"
        elif action == "edit":
            class_id     = request.form.get("class_id", "").strip()
            name         = request.form.get("name", "").strip().upper()
            course_id    = request.form.get("course_id", "").strip()
            intake_year  = request.form.get("intake_year") or None
            intake_month = request.form.get("intake_month") or None
            level        = request.form.get("level") or None
            cycle        = request.form.get("cycle") or None
            if not class_id or not name or not course_id:
                error = "Class name and course are required."
            else:
                row = db.table("classes").select("department_id").eq("id", class_id).single().execute().data
                if not row or row.get("department_id") != dept_id:
                    abort(403)
                try:
                    db.table("classes").update({
                        "name": name, "course_id": course_id,
                        "intake_year": intake_year, "intake_month": intake_month,
                        "level": level, "cycle": cycle
                    }).eq("id", class_id).execute()
                    write_audit_log("edit_class", target=name)
                    flash("Class updated successfully.", "success")
                    return redirect(url_for("dept_admin.classes"))
                except Exception as exc:
                    error = f"Error updating class: {exc}"
        elif action == "delete":
            class_id = request.form.get("class_id")
            row = db.table("classes").select("department_id").eq("id", class_id).single().execute().data
            if not row or row.get("department_id") != dept_id:
                abort(403)
            try:
                db.table("classes").delete().eq("id", class_id).execute()
                write_audit_log("delete_class", target=str(class_id))
                flash("Class deleted.", "success")
            except Exception as exc:
                flash(f"Cannot delete class (it may have students enrolled): {exc}", "danger")
            return redirect(url_for("dept_admin.classes"))
    classes_list = db.table("classes").select("*, courses(name)").eq("department_id", dept_id).order("name").execute().data or []
    courses = db.table("courses").select("*").eq("department_id", dept_id).order("name").execute().data or []
    return render_template("dept_admin/classes.html", classes=classes_list, courses=courses, error=error)


# ── Units ─────────────────────────────────────────────────────────────────────

@dept_admin_bp.route("/units", methods=["GET", "POST"])
@dept_admin_required
def units():
    db = get_service_client()
    dept_id = _dept_id()
    error = None
    if request.method == "POST":
        action    = request.form.get("action", "create")
        if action == "create":
            code      = request.form.get("code", "").strip().upper()
            name      = request.form.get("name", "").strip()
            course_id = request.form.get("course_id", "").strip()
            if not all([code, name, course_id]):
                error = "Code, name, and course are required."
            else:
                try:
                    db.table("units").insert({"code": code, "name": name,
                        "department_id": dept_id, "course_id": course_id}).execute()
                    write_audit_log("create_unit", target=code)
                    flash("Unit added successfully.", "success")
                    return redirect(url_for("dept_admin.units"))
                except Exception as exc:
                    error = f"Error creating unit: {exc}"
        elif action == "edit":
            unit_id   = request.form.get("unit_id", "").strip()
            code      = request.form.get("code", "").strip().upper()
            name      = request.form.get("name", "").strip()
            course_id = request.form.get("course_id", "").strip()
            if not all([unit_id, code, name, course_id]):
                error = "All fields are required."
            else:
                row = db.table("units").select("department_id").eq("id", unit_id).single().execute().data
                if not row or row.get("department_id") != dept_id:
                    abort(403)
                try:
                    db.table("units").update({
                        "code": code, "name": name, "course_id": course_id
                    }).eq("id", unit_id).execute()
                    write_audit_log("edit_unit", target=code)
                    flash("Unit updated successfully.", "success")
                    return redirect(url_for("dept_admin.units"))
                except Exception as exc:
                    error = f"Error updating unit: {exc}"
        elif action == "delete":
            unit_id = request.form.get("unit_id", "").strip()
            row = db.table("units").select("department_id").eq("id", unit_id).single().execute().data
            if not row or row.get("department_id") != dept_id:
                abort(403)
            try:
                db.table("units").delete().eq("id", unit_id).execute()
                write_audit_log("delete_unit", target=str(unit_id))
                flash("Unit deleted.", "success")
            except Exception as exc:
                flash(f"Cannot delete unit (it may have assessments or assignments linked to it): {exc}", "danger")
            return redirect(url_for("dept_admin.units"))
    units_list = db.table("units").select("*, courses(name)").eq("department_id", dept_id).order("code").execute().data or []
    courses    = db.table("courses").select("*").eq("department_id", dept_id).order("name").execute().data or []
    return render_template("dept_admin/units.html", units=units_list, courses=courses, error=error)


# ── Trainers ──────────────────────────────────────────────────────────────────

@dept_admin_bp.route("/trainers", methods=["GET", "POST"])
@dept_admin_required
def trainers():
    db = get_service_client()
    dept_id = _dept_id()
    error = None
    new_creds = None
    if request.method == "POST":
        email     = request.form.get("email", "").strip().lower()
        full_name = request.form.get("full_name", "").strip()
        staff_no  = request.form.get("staff_no", "").strip()
        password  = _gen_password()
        if not all([email, full_name]):
            error = "Email and full name are required."
        else:
            existing = db.table("user_profiles").select("id").eq("email", email).execute()
            if existing.data:
                error = "Email already exists."
            else:
                try:
                    from auth_utils import create_staff_auth_user
                    user_id = create_staff_auth_user(email=email, password=password,
                        full_name=full_name, role="trainer", department_id=dept_id)
                    if staff_no:
                        db.table("user_profiles").update({"staff_no": staff_no}).eq("id", user_id).execute()
                    write_audit_log("create_trainer", target=f"user:{user_id}")
                    new_creds = {"full_name": full_name, "email": email, "password": password}
                    flash(f"Trainer {full_name} added.", "success")
                except Exception as exc:
                    error = f"Error: {exc}"
    trainers_list = db.table("user_profiles").select("*, departments(name)").eq("role", "trainer").eq("department_id", dept_id).order("full_name").execute().data or []
    return render_template("dept_admin/trainers.html", trainers=trainers_list, error=error, new_creds=new_creds)


# ── Students ──────────────────────────────────────────────────────────────────

@dept_admin_bp.route("/students", methods=["GET", "POST"])
@dept_admin_required
def students():
    db = get_service_client()
    dept_id = _dept_id()
    error = None
    new_creds = None
    if request.method == "POST":
        admission_no = request.form.get("admission_no", "").strip()
        email        = request.form.get("email", "").strip().lower()
        full_name    = request.form.get("full_name", "").strip()
        class_id     = request.form.get("class_id", "").strip()
        password     = _gen_password()
        if not all([admission_no, email, full_name, class_id]):
            error = "All fields are required."
        else:
            dup = db.table("user_profiles").select("id").eq("admission_no", admission_no).execute()
            if dup.data:
                error = "Admission number already exists."
            else:
                try:
                    from auth_utils import create_student_auth_user
                    user_id = create_student_auth_user(admission_no=admission_no, password=password,
                        email=email, full_name=full_name, department_id=dept_id, class_id=class_id)
                    db.table("enrollments").insert({"student_id": user_id, "class_id": class_id}).execute()
                    write_audit_log("create_student", target=f"user:{user_id}")
                    new_creds = {"full_name": full_name, "admission_no": admission_no,
                                 "email": email, "password": password}
                    flash(f"Student {full_name} added.", "success")
                except Exception as exc:
                    error = f"Error: {exc}"
    students_list = (db.table("enrollments")
        .select("*, user_profiles(id, full_name, admission_no, email, mobile_number, is_active), classes(name)")
        .eq("classes.department_id", dept_id)
        .execute().data or [])
    # Fallback: query user_profiles directly
    if not students_list:
        students_list = db.table("user_profiles").select("*").eq("role", "student").eq("department_id", dept_id).order("full_name").execute().data or []
    classes = db.table("classes").select("*").eq("department_id", dept_id).order("name").execute().data or []
    search = request.args.get("q", "")
    return render_template("dept_admin/students.html", students=students_list, classes=classes,
                           error=error, new_creds=new_creds, search=search)


# ── Assign Units to Trainers ──────────────────────────────────────────────────

@dept_admin_bp.route("/trainer-units", methods=["GET", "POST"])
@dept_admin_bp.route("/assign-units", methods=["GET", "POST"])
@dept_admin_required
def trainer_units():
    db = get_service_client()
    dept_id = _dept_id()
    error = None
    classes_list = db.table("classes").select("id, name").eq("department_id", dept_id).order("name").execute().data or []
    active_year = datetime.now().year

    if request.method == "POST":
        action     = request.form.get("action", "assign")
        trainer_id = request.form.get("trainer_id", "").strip()
        unit_id    = request.form.get("unit_id", "").strip()
        class_id   = request.form.get("class_id", "").strip()
        year       = request.form.get("year", str(active_year)).strip()
        term       = request.form.get("term", "1").strip()
        if action == "assign":
            if not all([trainer_id, unit_id]):
                error = "Trainer and unit are required."
            else:
                # Verify trainer belongs to this department
                t_row = db.table("user_profiles").select("department_id").eq("id", trainer_id).single().execute().data
                if not t_row or t_row.get("department_id") != dept_id:
                    abort(403)
                # Verify unit belongs to this department
                u_row = db.table("units").select("department_id").eq("id", unit_id).single().execute().data
                if not u_row or u_row.get("department_id") != dept_id:
                    abort(403)
                # Train global assignment (trainer ↔ unit) — ignore if already exists
                try:
                    db.table("trainer_units").insert({"trainer_id": trainer_id, "unit_id": unit_id}).execute()
                    write_audit_log("assign_unit", target=f"trainer:{trainer_id}")
                except Exception:
                    pass  # global assignment may already exist
                # Class-level linking (class ↔ unit ↔ trainer) for attendance
                if class_id:
                    try:
                        db.table("class_units").insert({
                            "class_id": class_id,
                            "unit_id": unit_id,
                            "trainer_id": trainer_id,
                            "year": int(year),
                            "term": int(term)
                        }).execute()
                        flash("Unit linked to class for attendance.", "success")
                    except Exception as cx:
                        err_str2 = str(cx)
                        if "duplicate" in err_str2.lower() or "unique" in err_str2.lower():
                            # Record exists — update trainer_id instead
                            try:
                                db.table("class_units").update({
                                    "trainer_id": trainer_id
                                }).eq("class_id", class_id).eq("unit_id", unit_id).eq("year", int(year)).eq("term", int(term)).execute()
                                flash("Unit trainer updated for this class.", "success")
                            except Exception:
                                flash("Unit already linked to this class for that year/term.", "warning")
                        else:
                            flash(f"Class linking failed: {cx}", "warning")
                else:
                    flash("Unit assigned globally (class not selected, so it won't appear in attendance dropdown).", "info")
                return redirect(url_for("dept_admin.trainer_units"))
        elif action == "unassign":
            assign_id = request.form.get("assign_id", "").strip()
            if assign_id:
                try:
                    db.table("trainer_units").delete().eq("id", assign_id).execute()
                    write_audit_log("unassign_unit", target=f"assignment:{assign_id}")
                    flash("Assignment removed successfully.", "success")
                except Exception as exc:
                    flash(f"Error removing assignment: {exc}", "danger")
                return redirect(url_for("dept_admin.trainer_units"))

    # Use explicit FK hint for the user_profiles join to avoid ambiguity
    try:
        assignments = (db.table("trainer_units")
            .select("id, trainer_id, unit_id, assigned_at, user_profiles!trainer_units_trainer_id_fkey(full_name, staff_no), units(name, code)")
            .execute().data or [])
        # Filter to only this department's trainers
        dept_trainer_ids = {t["id"] for t in
            db.table("user_profiles").select("id").eq("role", "trainer").eq("department_id", dept_id).execute().data or []}
        assignments = [a for a in assignments if a.get("trainer_id") in dept_trainer_ids]
    except Exception:
        # Fallback: fetch without join, manually attach profile data
        assignments = (db.table("trainer_units")
            .select("id, trainer_id, unit_id, assigned_at")
            .execute().data or [])
        dept_trainer_ids = {t["id"] for t in
            db.table("user_profiles").select("id").eq("role", "trainer").eq("department_id", dept_id).execute().data or []}
        assignments = [a for a in assignments if a.get("trainer_id") in dept_trainer_ids]
        # Attach profile and unit info manually
        profiles = {p["id"]: p for p in
            db.table("user_profiles").select("id, full_name, staff_no").in_("id", list(dept_trainer_ids)).execute().data or []}
        dept_units = {u["id"]: u for u in
            db.table("units").select("id, name, code").eq("department_id", dept_id).execute().data or []}
        for a in assignments:
            a["user_profiles"] = profiles.get(a["trainer_id"], {})
            a["units"] = dept_units.get(a["unit_id"], {})

    # Fetch class-level assignments (class_units)
    try:
        class_assignments = (db.table("class_units")
            .select("*, classes(name), units(name, code), user_profiles!class_units_trainer_id_fkey(full_name)")
            .execute().data or [])
        # Filter to only this department's classes
        dept_class_ids = {c["id"] for c in classes_list}
        class_assignments = [ca for ca in class_assignments if ca.get("class_id") in dept_class_ids]
    except Exception:
        class_assignments = (db.table("class_units")
            .select("*")
            .execute().data or [])
        dept_class_ids = {c["id"] for c in classes_list}
        class_assignments = [ca for ca in class_assignments if ca.get("class_id") in dept_class_ids]
        class_map = {c["id"]: c for c in classes_list}
        unit_map  = {u["id"]: u for u in (db.table("units").select("id, name, code").eq("department_id", dept_id).execute().data or [])}
        for ca in class_assignments:
            ca["classes"] = class_map.get(ca["class_id"], {})
            ca["units"]   = unit_map.get(ca["unit_id"], {})
            ca["user_profiles"] = {}

    trainers = db.table("user_profiles").select("id, full_name, staff_no").eq("role", "trainer").eq("department_id", dept_id).order("full_name").execute().data or []
    units    = db.table("units").select("id, name, code").eq("department_id", dept_id).order("name").execute().data or []
    return render_template("dept_admin/trainer_units.html",
                           assignments=assignments, class_assignments=class_assignments,
                           trainers=trainers, units=units,
                           classes=classes_list, active_year=active_year, error=error)


# ── Attendance Overview ───────────────────────────────────────────────────────

@dept_admin_bp.route("/attendance")
@dept_admin_required
def attendance():
    from datetime import date as _date, datetime as _dt
    db       = get_service_client()
    dept_id  = _dept_id()

    class_filter  = request.args.get("class_id", "").strip()
    unit_filter   = request.args.get("unit_id",  "").strip()
    term_filter   = request.args.get("term",     "").strip()
    year_filter   = request.args.get("year",     str(_date.today().year)).strip()
    week_filter   = request.args.get("week",     "").strip()
    lesson_filter = request.args.get("lesson",   "").strip()

    classes  = (db.table("classes").select("id, name")
                .eq("department_id", dept_id).order("name").execute().data or [])
    units    = (db.table("units").select("id, name, code")
                .eq("department_id", dept_id).order("name").execute().data or [])
    dept_obj = (db.table("departments").select("id, name")
                .eq("id", dept_id).single().execute().data or {})

    def _norm_lesson(l):
        s = str(l).strip()
        return f"L{s}" if s in ("1", "2", "3", "4") else s

    ALL_LESSONS = ["L1", "L2", "L3", "L4"]
    ALL_WEEKS   = list(range(1, 16))
    LESSONS     = ALL_LESSONS[:]
    WEEKS       = ALL_WEEKS[:]

    term_int = int(term_filter) if term_filter else None
    year_int = int(year_filter) if year_filter else None

    # Apply optional week / lesson filters
    if week_filter:
        try:
            w = int(week_filter)
            if 1 <= w <= 15:
                WEEKS = [w]
            else:
                week_filter = ""
        except ValueError:
            week_filter = ""

    if lesson_filter:
        norm = _norm_lesson(lesson_filter)
        if norm in ALL_LESSONS:
            LESSONS = [norm]
        else:
            lesson_filter = ""

    TERM_LABELS = {1: "Term 1 (Jan–Apr)", 2: "Term 2 (May–Aug)", 3: "Term 3 (Sep–Dec)"}
    term_label  = TERM_LABELS.get(term_int, f"Term {term_int}") if term_int else ""

    matrix       = []
    unit_obj     = cls_obj = None
    trainer_name = ""

    if class_filter and unit_filter:
        for u in units:
            if u["id"] == unit_filter:
                unit_obj = u; break
        for c in classes:
            if c["id"] == class_filter:
                cls_obj = c; break

        # Trainer for this class/unit
        try:
            cu = (db.table("class_units")
                  .select("user_profiles!class_units_trainer_id_fkey(full_name)")
                  .eq("class_id", class_filter).eq("unit_id", unit_filter)
                  .limit(1).execute().data or [])
            if cu:
                trainer_name = (cu[0].get("user_profiles") or {}).get("full_name", "")
        except Exception:
            pass

        enr_rows = (db.table("enrollments")
                    .select("student_id, "
                            "user_profiles!enrollments_student_id_fkey"
                            "(id, full_name, admission_no)")
                    .eq("class_id", class_filter).execute().data or [])

        students_ordered = []
        for e in enr_rows:
            up  = e.get("user_profiles") or {}
            sid = up.get("id") or e.get("student_id")
            if sid and not any(s["id"] == sid for s in students_ordered):
                students_ordered.append({
                    "id":           sid,
                    "full_name":    up.get("full_name", "—"),
                    "admission_no": up.get("admission_no", "—"),
                })
        students_ordered.sort(key=lambda s: s["full_name"])

        student_ids = [s["id"] for s in students_ordered]
        if student_ids:
            q = (db.table("attendance")
                 .select("student_id, week, lesson, status")
                 .eq("unit_id", unit_filter)
                 .in_("student_id", student_ids))
            if term_int:    q = q.eq("term", term_int)
            if year_int:    q = q.eq("year", year_int)
            if week_filter: q = q.eq("week", int(week_filter))
            att_rows = q.execute().data or []
        else:
            att_rows = []

        pivot = {}
        for r in att_rows:
            key = (r["week"], _norm_lesson(r["lesson"]))
            pivot.setdefault(r["student_id"], {})[key] = r["status"]

        for s in students_ordered:
            cells = {}
            present = absent = 0
            for w in WEEKS:
                for l in LESSONS:
                    st = pivot.get(s["id"], {}).get((w, l))
                    cells[(w, l)] = st
                    if st == "present":  present += 1
                    elif st == "absent": absent  += 1
            total = present + absent
            pct   = round(present / total * 100, 1) if total else 0
            matrix.append({
                "id":           s["id"],
                "full_name":    s["full_name"],
                "admission_no": s["admission_no"],
                "cells":        cells,
                "present":      present,
                "absent":       absent,
                "total":        total,
                "pct":          pct,
            })

    generated_at = _dt.now().strftime("%d %b %Y %H:%M")

    return render_template(
        "dept_admin/attendance.html",
        classes=classes,
        units=units,
        class_filter=class_filter,
        unit_filter=unit_filter,
        term_filter=term_filter,
        year_filter=year_filter,
        week_filter=week_filter,
        lesson_filter=lesson_filter,
        term_int=term_int,
        year_int=year_int,
        term_label=term_label,
        matrix=matrix,
        unit_obj=unit_obj,
        cls_obj=cls_obj,
        dept_obj=dept_obj,
        trainer_name=trainer_name,
        generated_at=generated_at,
        WEEKS=WEEKS,
        ALL_WEEKS=ALL_WEEKS,
        LESSONS=LESSONS,
        ALL_LESSONS=ALL_LESSONS,
    )


# ── Attendance Matrix PDF Download ───────────────────────────────────────────

@dept_admin_bp.route("/unit-attendance-pdf")
@dept_admin_required
def unit_attendance_pdf():
    """Landscape unit attendance register (all taught weeks) for HOD."""
    from unit_attendance_register import build_unit_attendance_register

    db = get_service_client()
    dept_id = _dept_id()

    class_id = request.args.get("class_id", "").strip()
    unit_id  = request.args.get("unit_id", "").strip()
    year     = request.args.get("year", datetime.now().year, type=int)
    term     = request.args.get("term", 1, type=int)

    if not (class_id and unit_id):
        flash("Select a class and unit first.", "warning")
        return redirect(url_for("dept_admin.attendance"))

    # Scope: class must belong to this department
    cls = (db.table("classes").select("id, department_id")
             .eq("id", class_id).single().execute().data or {})
    if not cls or cls.get("department_id") != dept_id:
        flash("Class not found in your department.", "error")
        return redirect(url_for("dept_admin.attendance"))

    data = build_unit_attendance_register(
        db, class_id=class_id, unit_id=unit_id, year=year, term=term, trainer_id=None,
    )
    if not data:
        flash("No attendance records found for this unit in the selected period.", "warning")
        return redirect(url_for("dept_admin.attendance",
                                class_id=class_id, unit_id=unit_id, year=year, term=term))

    return render_template("shared/unit_attendance_pdf.html", **data)


@dept_admin_bp.route("/attendance/pdf")
@dept_admin_required
def attendance_matrix_pdf():
    """Generate the attendance matrix as a PDF with official header."""
    import io, os
    db      = get_service_client()
    dept_id = _dept_id()

    class_filter  = request.args.get("class_id", "")
    unit_filter   = request.args.get("unit_id",  "")
    term_filter   = request.args.get("term",     "")
    year_filter   = request.args.get("year",     "")
    week_filter   = request.args.get("week",     "")
    lesson_filter = request.args.get("lesson",   "")

    if not (class_filter and unit_filter):
        flash("Select a class and unit first.", "warning")
        return redirect(url_for("dept_admin.attendance"))

    def _norm_lesson(l):
        s = str(l).strip()
        return f"L{s}" if s in ("1", "2", "3", "4") else s

    ALL_LESSONS = ["L1", "L2", "L3", "L4"]
    ALL_WEEKS   = list(range(1, 16))
    LESSONS     = ALL_LESSONS[:]
    WEEKS       = ALL_WEEKS[:]

    term_int = int(term_filter) if term_filter else None
    year_int = int(year_filter) if year_filter else None

    if week_filter:
        try:
            w = int(week_filter)
            if 1 <= w <= 15:
                WEEKS = [w]
            else:
                week_filter = ""
        except ValueError:
            week_filter = ""

    if lesson_filter:
        norm = _norm_lesson(lesson_filter)
        if norm in ALL_LESSONS:
            LESSONS = [norm]
        else:
            lesson_filter = ""

    TERM_LABELS = {1: "Term 1 (Jan–Apr)", 2: "Term 2 (May–Aug)", 3: "Term 3 (Sep–Dec)"}
    term_label  = TERM_LABELS.get(term_int, f"Term {term_int}") if term_int else ""

    unit_obj  = (db.table("units").select("id, name, code").eq("id", unit_filter).single().execute().data or {})
    cls_obj   = (db.table("classes").select("id, name").eq("id", class_filter).single().execute().data or {})
    dept_obj  = (db.table("departments").select("name").eq("id", dept_id).single().execute().data or {})

    trainer_name = ""
    try:
        cu = (db.table("class_units")
              .select("user_profiles!class_units_trainer_id_fkey(full_name)")
              .eq("class_id", class_filter).eq("unit_id", unit_filter)
              .limit(1).execute().data or [])
        if cu:
            trainer_name = (cu[0].get("user_profiles") or {}).get("full_name", "")
    except Exception:
        pass

    enr_rows = (db.table("enrollments")
                .select("student_id, user_profiles!enrollments_student_id_fkey(id, full_name, admission_no)")
                .eq("class_id", class_filter).execute().data or [])

    students_ordered = []
    for e in enr_rows:
        up = e.get("user_profiles") or {}
        sid = up.get("id") or e.get("student_id")
        if sid and not any(s["id"] == sid for s in students_ordered):
            students_ordered.append({"id": sid, "full_name": up.get("full_name", "—"),
                                     "admission_no": up.get("admission_no", "—")})
    students_ordered.sort(key=lambda s: s["full_name"])

    student_ids = [s["id"] for s in students_ordered]
    att_rows = []
    if student_ids:
        q = (db.table("attendance").select("student_id, week, lesson, status")
             .eq("unit_id", unit_filter).in_("student_id", student_ids))
        if term_int:    q = q.eq("term", term_int)
        if year_int:    q = q.eq("year", year_int)
        if week_filter: q = q.eq("week", int(week_filter))
        att_rows = q.execute().data or []

    pivot = {}
    for r in att_rows:
        pivot.setdefault(r["student_id"], {})[(r["week"], _norm_lesson(r["lesson"]))] = r["status"]

    matrix = []
    for s in students_ordered:
        cells = {}
        present = absent = 0
        for w in WEEKS:
            for l in LESSONS:
                st = pivot.get(s["id"], {}).get((w, l))
                cells[(w, l)] = st
                if st == "present":  present += 1
                elif st == "absent": absent  += 1
        total = present + absent
        pct   = round(present / total * 100, 1) if total else 0
        matrix.append({"full_name": s["full_name"], "admission_no": s["admission_no"],
                       "cells": cells, "present": present, "absent": absent,
                       "total": total, "pct": pct})

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, Image, HRFlowable)
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        _CELL_FONT = "Helvetica"
        try:
            import matplotlib.font_manager as _fm
            _fp = _fm.findfont(_fm.FontProperties(family="DejaVu Sans"))
            pdfmetrics.registerFont(TTFont("DejaVu", _fp))
            _CELL_FONT = "DejaVu"
        except Exception:
            pass

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=10*mm, rightMargin=10*mm,
                                topMargin=8*mm, bottomMargin=10*mm)

        styles = getSampleStyleSheet()

        def _ps(name, font="Helvetica", size=9, align=TA_CENTER, color=None, **kw):
            args = dict(parent=styles["Normal"], fontName=font,
                        fontSize=size, alignment=align, **kw)
            if color:
                args["textColor"] = color
            return ParagraphStyle(name, **args)

        CENTER  = _ps("C")
        BOLD_C  = _ps("BC",  font="Helvetica-Bold")
        LEFT    = _ps("L",   align=TA_LEFT)
        BOLD_L  = _ps("BL",  font="Helvetica-Bold", align=TA_LEFT)
        INST    = _ps("INST",font="Helvetica-Bold", size=14, align=TA_CENTER)
        DEPT    = _ps("DEPT",font="Helvetica",      size=10, align=TA_CENTER,
                      color=colors.HexColor("#374151"))
        DOCTIT  = _ps("DT",  font="Helvetica-Bold", size=12, align=TA_CENTER,
                      color=colors.HexColor("#0F2C54"))
        META_LB = _ps("MLB", font="Helvetica-Bold", size=8,  align=TA_LEFT,
                      color=colors.HexColor("#6B7280"))
        META_VL = _ps("MVL", font="Helvetica-Bold", size=9,  align=TA_LEFT,
                      color=colors.HexColor("#111827"))

        coat_path = os.path.join(os.path.dirname(__file__), "..", "static", "assets", "KENYACOATOFARMS.png")
        logo_path = os.path.join(os.path.dirname(__file__), "..", "static", "assets", "THIKATTILOGO.jpg")

        LOGO_H = 20*mm
        def _img(path, h=LOGO_H):
            try:
                img = Image(path)
                img.drawHeight = h
                img.drawWidth  = h * (img.imageWidth / img.imageHeight)
                return img
            except Exception:
                return Paragraph("", CENTER)

        # ── Top banner: Coat of Arms | Title block | TTTI Logo ────────────
        now_str = datetime.now().strftime("%d %b %Y %H:%M")
        title_block = [
            Paragraph("THIKA TECHNICAL TRAINING INSTITUTE", INST),
            Spacer(1, 1*mm),
            Paragraph(f"Department of {dept_obj.get('name', '')}", DEPT),
            Spacer(1, 1.5*mm),
            Paragraph("UNIT ATTENDANCE REGISTER", DOCTIT),
        ]
        banner_data = [[_img(coat_path), title_block, _img(logo_path)]]
        banner_tbl  = Table(banner_data, colWidths=[24*mm, doc.width - 48*mm, 24*mm])
        banner_tbl.setStyle(TableStyle([
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN",        (0, 0), (0,  0),  "LEFT"),
            ("ALIGN",        (2, 0), (2,  0),  "RIGHT"),
            ("ALIGN",        (1, 0), (1,  0),  "CENTER"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING",   (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
        ]))

        # ── Metadata grid ─────────────────────────────────────────────────
        def _mf(label, value):
            return [Paragraph(label, META_LB), Paragraph(str(value) if value else "—", META_VL)]

        unit_label = f"{unit_obj.get('code','')} — {unit_obj.get('name','')}"
        week_label = f"Week {week_filter}" if week_filter else "All weeks (1–15)"
        les_label  = lesson_filter if lesson_filter else "All lessons"

        meta_rows = [
            [*_mf("CLASS",           cls_obj.get("name",  "")),
             *_mf("YEAR",            year_int or ""),
             *_mf("TERM",            term_label or "All"),
             *_mf("WEEK",            week_label)],
            [*_mf("UNIT",            unit_label),
             *_mf("LESSON",          les_label),
             *_mf("TRAINER",         trainer_name or "—"),
             *_mf("DATE GENERATED",  now_str)],
            [*_mf("TOTAL STUDENTS",  len(matrix)),
             Paragraph("", CENTER), Paragraph("", CENTER),
             Paragraph("", CENTER), Paragraph("", CENTER),
             Paragraph("", CENTER), Paragraph("", CENTER),
             Paragraph("", CENTER)],
        ]
        col_w = doc.width / 8
        meta_tbl = Table(meta_rows, colWidths=[col_w] * 8)
        LBLUE = colors.HexColor("#EFF6FF")
        meta_tbl.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, -1), LBLUE),
            ("ROWBACKGROUNDS",(0,0), (-1, -1), [LBLUE, colors.HexColor("#F0FDF4")]),
            ("BOX",          (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
            ("INNERGRID",    (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            ("LEFTPADDING",  (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ]))

        # ── Build matrix table ────────────────────────────────────────────
        NAVY  = colors.HexColor("#0F2C54")
        DBLUE = colors.HexColor("#1A3D6E")
        WHITE = colors.white
        GREEN = colors.HexColor("#D1FAE5")
        RED   = colors.HexColor("#FEE2E2")
        GREY  = colors.HexColor("#F3F4F6")

        # Row 1: week span headers
        week_row = ["#", "Trainee", "Adm No"]
        for w in WEEKS:
            week_row += [f"Week {w}"] + [""] * (len(LESSONS) - 1)
        week_row += ["Present", "Absent", "Total", "%"]

        # Row 2: lesson sub-headers
        lesson_row = ["", "", ""]
        for _ in WEEKS:
            for l in LESSONS:
                lesson_row.append(l)
        lesson_row += ["", "", "", ""]

        # Data rows
        data_rows = []
        for i, row in enumerate(matrix, 1):
            r = [str(i), row["full_name"], row["admission_no"]]
            for w in WEEKS:
                for l in LESSONS:
                    st = row["cells"].get((w, l))
                    r.append("✓" if st == "present" else ("✗" if st == "absent" else ""))
            r += [str(row["present"]), str(row["absent"]), str(row["total"]),
                  f"{row['pct']}%"]
            data_rows.append(r)

        if not data_rows:
            data_rows = [["—"] * len(week_row)]

        tbl_data   = [week_row, lesson_row] + data_rows
        n_cols     = len(week_row)
        fixed_cols = 3
        summary_cols = 4
        lesson_cols  = n_cols - fixed_cols - summary_cols

        # Column widths
        name_w    = 38*mm
        adm_w     = 18*mm
        num_w     = 6*mm
        idx_w     = 6*mm
        summ_w    = 9*mm
        lesson_cw = max(5*mm, (doc.width - name_w - adm_w - idx_w - summary_cols*summ_w) / lesson_cols)
        col_widths = [idx_w, name_w, adm_w] + [lesson_cw] * lesson_cols + [summ_w] * summary_cols

        tbl = Table(tbl_data, colWidths=col_widths, repeatRows=2)

        # Span week header cells
        span_cmds = []
        for wi, w in enumerate(WEEKS):
            cs = fixed_cols + wi * len(LESSONS)
            ce = cs + len(LESSONS) - 1
            span_cmds.append(("SPAN", (cs, 0), (ce, 0)))

        cell_cmds = []
        for ri, row in enumerate(data_rows, 2):
            for ci, val in enumerate(row[fixed_cols:fixed_cols + lesson_cols], fixed_cols):
                if val == "✓":
                    cell_cmds.append(("BACKGROUND", (ci, ri), (ci, ri), GREEN))
                    cell_cmds.append(("TEXTCOLOR",  (ci, ri), (ci, ri), colors.HexColor("#15803d")))
                elif val == "✗":
                    cell_cmds.append(("BACKGROUND", (ci, ri), (ci, ri), RED))
                    cell_cmds.append(("TEXTCOLOR",  (ci, ri), (ci, ri), colors.HexColor("#dc2626")))

        HDR_FILL = colors.HexColor("#DCE6F4")
        HDR_TEXT = colors.HexColor("#0F2744")
        tbl.setStyle(TableStyle([
            # Header rows — light institutional style
            ("BACKGROUND",  (0, 0), (-1, 0), HDR_FILL),
            ("TEXTCOLOR",   (0, 0), (-1, 0), HDR_TEXT),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0), 6),
            ("BACKGROUND",  (0, 1), (-1, 1), HDR_FILL),
            ("TEXTCOLOR",   (0, 1), (-1, 1), HDR_TEXT),
            ("FONTNAME",    (0, 1), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 1), (-1, 1), 5),
            # Data rows — use Unicode-capable font for ✓/✗
            ("FONTNAME",    (0, 2), (-1, -1), _CELL_FONT),
            ("FONTSIZE",    (0, 2), (-1, -1), 6),
            ("ROWBACKGROUNDS", (0, 2), (-1, -1), [WHITE, GREY]),
            # Summary columns bold
            ("FONTNAME",    (-4, 2), (-1, -1), "Helvetica-Bold"),
            # Alignment
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("ALIGN",       (1, 2), (1, -1),  "LEFT"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            # Grid
            ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
            ("LINEBELOW",   (0, 1), (-1, 1),  0.8, HDR_TEXT),
            # Padding
            ("TOPPADDING",  (0, 0), (-1, -1), 1.5),
            ("BOTTOMPADDING",(0, 0),(-1, -1), 1.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 1.5),
            ("RIGHTPADDING",(0, 0), (-1, -1), 1.5),
        ] + span_cmds + cell_cmds))

        term_lbl = f"Term{term_int}_" if term_int else ""
        year_lbl = f"{year_int}_" if year_int else ""
        fname    = (f"Unit_Attendance_Register_{unit_obj.get('code','')}_"
                    f"{cls_obj.get('name','').replace(' ','_')}_"
                    f"{term_lbl}{year_lbl}{datetime.now().strftime('%Y%m%d')}.pdf")

        # ── Signing block (shared official sign-off) ───────────────────────
        sign_flowables = pdf_signature_block(
            doc.width,
            officer_title="Departmental Monitoring Officer",
            extra_officers=("Head of Department",),
        )

        elements = [
            banner_tbl,
            Spacer(1, 3*mm),
            meta_tbl,
            HRFlowable(width="100%", thickness=1.5, color=NAVY, spaceAfter=4*mm),
            tbl,
            Spacer(1, 6*mm),
        ] + sign_flowables
        doc.build(elements)

        buf.seek(0)
        resp = make_response(buf.read())
        resp.headers["Content-Type"]        = "application/pdf"
        resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp

    except ImportError:
        flash("PDF generation requires reportlab. Run: pip install reportlab pillow", "warning")
        return redirect(url_for("dept_admin.attendance",
                                class_id=class_filter, unit_id=unit_filter,
                                term=term_filter, year=year_filter,
                                week=week_filter, lesson=lesson_filter))


# ── Assessments Overview ──────────────────────────────────────────────────────

@dept_admin_bp.route("/assessments")
@dept_admin_required
def assessments():
    db = get_service_client()
    dept_id = _dept_id()
    status_filter = request.args.get("status", "")
    query = (db.table("assessments")
        .select("*, user_profiles!assessments_student_id_fkey(full_name, admission_no), units(name, code, department_id), classes(name)")
        .order("uploaded_at", desc=True).limit(200))
    if status_filter:
        query = query.eq("status", status_filter)
    records = query.execute().data or []
    records = [r for r in records if r.get("units", {}).get("department_id") == dept_id]
    return render_template("dept_admin/assessments.html",
                           assessments=records, status_filter=status_filter)


# ── Download Unit Report CSV ──────────────────────────────────────────────────

@dept_admin_bp.route("/download-unit-report/<unit_id>")
@dept_admin_required
def download_unit_report(unit_id):
    db = get_service_client()
    dept_id = _dept_id()
    unit = db.table("units").select("*").eq("id", unit_id).single().execute().data
    if not unit or unit.get("department_id") != dept_id:
        abort(403)
    assessments = (db.table("assessments")
        .select("*, user_profiles!assessments_student_id_fkey(full_name, admission_no), classes(name)")
        .eq("unit_id", unit_id).execute().data or [])
    import csv, io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Admission No", "Full Name", "Class", "Type", "No.", "Term", "Cycle", "Year", "Status"])
    for a in assessments:
        writer.writerow([
            a.get("user_profiles", {}).get("admission_no", ""),
            a.get("user_profiles", {}).get("full_name", ""),
            a.get("classes", {}).get("name", ""),
            a.get("assessment_type", ""), a.get("assessment_no", ""),
            a.get("term", ""), a.get("cycle", ""), a.get("year", ""), a.get("status", "")
        ])
    output.seek(0)
    resp = make_response(output.getvalue())
    resp.headers["Content-Disposition"] = f"attachment; filename=unit_{unit.get('code','report')}.csv"
    resp.headers["Content-type"] = "text/csv"
    return resp


# ── Exam Bookings ─────────────────────────────────────────────────────────────

@dept_admin_bp.route("/exam-bookings")
@dept_admin_required
def exam_bookings():
    from datetime import date as _date
    db = get_service_client()
    dept_id       = _dept_id()
    status_filter = request.args.get("status", "all")
    q             = request.args.get("q", "").strip()
    year_filter   = request.args.get("year",  "").strip()
    term_filter   = request.args.get("term",  "").strip()

    # term → month range
    # (start_month, end_month, last_day_of_end_month)
    TERM_MONTHS = {"1": ("01","04","30"), "2": ("05","08","31"), "3": ("09","12","31")}

    enr = (db.table("enrollments")
           .select("student_id, classes!inner(department_id, name)")
           .eq("classes.department_id", dept_id)
           .execute().data or [])

    student_class = {}
    for e in enr:
        sid = e.get("student_id")
        cls = (e.get("classes") or {}).get("name", "")
        if sid and sid not in student_class:
            student_class[sid] = cls

    student_ids = list(student_class.keys())

    bookings = []
    if student_ids:
        query = (db.table("exam_bookings")
            .select("*, "
                    "units(name, code), "
                    "student:user_profiles!exam_bookings_student_id_fkey"
                    "(full_name, admission_no, mobile_number), "
                    "reviewer:user_profiles!exam_bookings_approved_by_fkey(full_name)")
            .in_("student_id", student_ids)
            .order("created_at", desc=True))

        if status_filter and status_filter != "all":
            query = query.eq("status", status_filter)

        yr = year_filter or str(_date.today().year)
        if year_filter:
            query = query.gte("exam_date", f"{yr}-01-01").lte("exam_date", f"{yr}-12-31")
        if term_filter and term_filter in TERM_MONTHS:
            m0, m1, last = TERM_MONTHS[term_filter]
            query = query.gte("exam_date", f"{yr}-{m0}-01").lte("exam_date", f"{yr}-{m1}-{last}")

        bookings = query.execute().data or []

    for b in bookings:
        b["student_user"]     = b.get("student")  or {}
        b["approved_by_user"] = b.get("reviewer") or {}
        b["class_name"]       = student_class.get(b.get("student_id"), "—")

    if q:
        ql = q.lower()
        bookings = [b for b in bookings
                    if ql in b["student_user"].get("full_name",  "").lower()
                    or ql in b["student_user"].get("admission_no", "").lower()]

    all_student_bookings = []
    if student_ids:
        try:
            all_student_bookings = (db.table("exam_bookings")
                .select("status").in_("student_id", student_ids).execute().data or [])
        except Exception:
            pass
    counts = {
        "all":      len(all_student_bookings),
        "pending":  sum(1 for b in all_student_bookings if b["status"] == "pending"),
        "approved": sum(1 for b in all_student_bookings if b["status"] == "approved"),
        "rejected": sum(1 for b in all_student_bookings if b["status"] == "rejected"),
        "completed": sum(1 for b in all_student_bookings if b["status"] == "completed"),
    }

    return render_template("dept_admin/exam_bookings.html",
                           bookings=bookings,
                           status_filter=status_filter,
                           q=q,
                           year_filter=year_filter,
                           term_filter=term_filter,
                           counts=counts)


@dept_admin_bp.route("/exam-bookings/export")
@dept_admin_required
def export_exam_bookings():
    """Export exam bookings per class as a styled Excel workbook."""
    import io
    from datetime import date as _date
    db      = get_service_client()
    dept_id = _dept_id()

    status_filter = request.args.get("status", "")
    class_filter  = request.args.get("class_id", "")
    year_filter   = request.args.get("year",  "").strip()
    term_filter   = request.args.get("term",  "").strip()
    TERM_MONTHS   = {"1": ("01","04","30"), "2": ("05","08","31"), "3": ("09","12","31")}

    # Get all enrollments for this department
    enr = (db.table("enrollments")
             .select("student_id, classes!inner(id, department_id, name)")
             .eq("classes.department_id", dept_id)
             .execute().data or [])

    student_class = {}
    for e in enr:
        sid = e.get("student_id")
        cls = e.get("classes") or {}
        if sid and sid not in student_class:
            student_class[sid] = {"class_id": cls.get("id"), "class_name": cls.get("name", "Unknown")}

    student_ids = list(student_class.keys())
    if not student_ids:
        flash("No students found in your department.", "warning")
        return redirect(url_for("dept_admin.exam_bookings"))

    query = (db.table("exam_bookings")
               .select("id, student_id, unit_id, exam_session, serial_number, status, "
                       "attempt_type, previous_grade, created_at, "
                       "units(name, code), "
                       "student:user_profiles!exam_bookings_student_id_fkey(full_name, admission_no, mobile_number)")
               .in_("student_id", student_ids)
               .order("student_id")
               .order("created_at"))

    if status_filter and status_filter != "all":
        query = query.eq("status", status_filter)

    yr = year_filter or str(_date.today().year)
    if year_filter:
        query = query.gte("exam_date", f"{yr}-01-01").lte("exam_date", f"{yr}-12-31")
    if term_filter and term_filter in TERM_MONTHS:
        m0, m1, last = TERM_MONTHS[term_filter]
        query = query.gte("exam_date", f"{yr}-{m0}-01").lte("exam_date", f"{yr}-{m1}-{last}")

    bookings = query.execute().data or []

    if not bookings:
        flash("No bookings found for the selected filters.", "info")
        return redirect(url_for("dept_admin.exam_bookings"))

    # Attach class info
    for b in bookings:
        b["_cls"] = student_class.get(b.get("student_id"), {}).get("class_name", "Unknown")

    # Filter by class if requested
    if class_filter:
        bookings = [b for b in bookings if student_class.get(b.get("student_id"), {}).get("class_id") == class_filter]

    # Group by class name
    from collections import OrderedDict
    by_class = OrderedDict()
    for b in bookings:
        cn = b["_cls"]
        by_class.setdefault(cn, []).append(b)

    # ── Build Excel ──────────────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    dept_obj = (db.table("departments").select("name")
                .eq("id", dept_id).single().execute().data or {})
    dept_name = dept_obj.get("name", "")
    hod_name = (current_user() or {}).get("full_name", "")

    wb = Workbook()
    wb.remove(wb.active)

    alt     = PatternFill("solid", fgColor="EFF6FF")
    pending = PatternFill("solid", fgColor="FEF3C7")
    approved= PatternFill("solid", fgColor="DCFCE7")
    rejected= PatternFill("solid", fgColor="FEE2E2")
    thin    = Side(style="thin", color="B0C4D8")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    HEADERS = ["S/N", "Admission No.", "Full Name", "Phone",
               "Unit Code", "Unit Name", "Attempt Type", "Prev. Grade",
               "Serial No.", "Session / Term", "Submitted", "Status"]
    WIDTHS  = [6, 16, 28, 16, 12, 34, 20, 12, 24, 22, 14, 12]

    AT_LABELS = {
        "first_attempt": "First Attempt",
        "retake":        "Retake (NYC/Fail)",
        "missing_unit":  "Missed Unit",
    }

    for cls_name, rows in by_class.items():
        ws = wb.create_sheet(title=cls_name[:31])

        # ── Letterhead ───────────────────────────────────────────────────────────
        next_row = excel_letterhead(
            ws, "Exam Bookings Report", len(HEADERS),
            dept_name=dept_name,
            meta_lines=[
                f"Class: {cls_name}",
                f"Generated: {datetime.now().strftime('%d %B %Y')}    |    "
                f"Total Entries: {len(rows)}    |    "
                f"Status Filter: {status_filter.upper() if status_filter else 'ALL'}",
            ],
        )

        # ── Headers ──────────────────────────────────────────────────────────────
        hdr_row = next_row
        for ci, h in enumerate(HEADERS, 1):
            ws.cell(row=hdr_row, column=ci, value=h)
        style_header_row(ws, hdr_row, len(HEADERS))
        ws.row_dimensions[hdr_row].height = 26

        # ── Group by student ─────────────────────────────────────────────────────
        by_student = OrderedDict()
        for b in rows:
            st  = b.get("student") or {}
            adm = st.get("admission_no", b["student_id"])
            by_student.setdefault(adm, {"student": st, "rows": []})["rows"].append(b)

        sn = 1
        for adm, data in by_student.items():
            st = data["student"]
            for b in data["rows"]:
                unit = b.get("units") or {}
                at   = b.get("attempt_type") or "first_attempt"
                status = (b.get("status") or "pending").lower()

                row_vals = [
                    sn,
                    st.get("admission_no", ""),
                    st.get("full_name", ""),
                    st.get("mobile_number", ""),
                    unit.get("code", ""),
                    unit.get("name", ""),
                    AT_LABELS.get(at, at),
                    b.get("previous_grade", ""),
                    b.get("serial_number", ""),
                    b.get("exam_session", ""),
                    (b.get("created_at") or "")[:10],
                    status.upper(),
                ]
                ws.append(row_vals)
                dr = ws.max_row

                # Status colour on last cell
                stat_fill = {"approved": approved, "rejected": rejected, "pending": pending}.get(status)
                row_fill  = alt if sn % 2 == 0 else None

                for ci in range(1, len(HEADERS) + 1):
                    c = ws.cell(row=dr, column=ci)
                    if ci == len(HEADERS) and stat_fill:
                        c.fill = stat_fill
                    elif row_fill:
                        c.fill = row_fill
                    c.border = border
                    c.alignment = Alignment(vertical="center", wrap_text=True)

                sn += 1

        # ── Column widths + freeze ────────────────────────────────────────────────
        from openpyxl.utils import get_column_letter
        for ci, w in enumerate(WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

        # ── Official sign-off ─────────────────────────────────────────────────────
        excel_signature_block(ws, ws.max_row + 2,
                              officer_title="Head of Department",
                              officer_name=hod_name)

    if not wb.sheetnames:
        flash("No data to export.", "info")
        return redirect(url_for("dept_admin.exam_bookings"))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"ExamBookings_{datetime.now().strftime('%Y%m%d')}.xlsx"
    from flask import Response as _Resp
    return _Resp(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@dept_admin_bp.route("/exam-bookings/trainee/<student_id>/approved-pdf")
@dept_admin_required
def trainee_approved_bookings_pdf(student_id):
    """Printable PDF of all approved exam bookings for one trainee."""
    db      = get_service_client()
    dept    = db.table("departments").select("name").eq("id", _dept_id()).single().execute().data or {}
    student = (db.table("user_profiles")
               .select("full_name, admission_no, mobile_number")
               .eq("id", student_id).single().execute().data or {})

    enr = (db.table("enrollments")
           .select("classes(name)").eq("student_id", student_id)
           .limit(1).execute().data or [])
    class_name = ((enr[0].get("classes") or {}).get("name", "")) if enr else ""

    bookings = (db.table("exam_bookings")
                .select("*, units(name, code)")
                .eq("student_id", student_id)
                .eq("status", "approved")
                .order("exam_date")
                .execute().data or [])

    return render_template(
        "dept_admin/trainee_approved_bookings_pdf.html",
        student=student,
        class_name=class_name,
        dept_name=dept.get("name", ""),
        bookings=bookings,
        date_gen=datetime.now().strftime("%d %B %Y"),
    )


@dept_admin_bp.route("/exam-bookings/<booking_id>/approve", methods=["POST"])
@dept_admin_required
def approve_exam_booking(booking_id):
    db = get_service_client()
    dept_id = _dept_id()
    user = current_user()
    booking = db.table("exam_bookings").select("*").eq("id", booking_id).single().execute().data
    if not booking:
        abort(404)
    unit = db.table("units").select("department_id").eq("id", booking["unit_id"]).single().execute().data
    if not unit or unit["department_id"] != dept_id:
        abort(403)
    try:
        db.table("exam_bookings").update({"status": "approved", "approved_by": user["id"],
            "approved_at": datetime.now().isoformat()}).eq("id", booking_id).execute()
        write_audit_log("approve_exam_booking", target=f"booking:{booking_id}")
        from notifications import create_notification
        create_notification(user_id=booking["student_id"], title="Exam Booking Approved",
            message=f"Your exam booking for {booking.get('exam_date')} has been approved.",
            notification_type="success", action_url="/student/exam-bookings")
        flash("Exam booking approved.", "success")
    except Exception as e:
        flash(f"Error: {e}", "danger")
    return redirect(url_for("dept_admin.exam_bookings"))


@dept_admin_bp.route("/exam-bookings/batch-approve", methods=["POST"])
@dept_admin_required
def batch_approve_exam_bookings():
    """Approve all pending bookings sharing a Form 1A serial number."""
    db = get_service_client()
    dept_id = _dept_id()
    user = current_user()
    serial = (request.form.get("serial_number") or "").strip()
    if not serial:
        flash("Serial number required.", "error")
        return redirect(url_for("dept_admin.exam_bookings"))

    rows = (db.table("exam_bookings")
              .select("id, student_id, unit_id, status")
              .eq("serial_number", serial)
              .eq("status", "pending")
              .execute().data or [])
    approved = 0
    student_id = None
    for booking in rows:
        unit = (db.table("units").select("department_id")
                  .eq("id", booking["unit_id"]).limit(1).execute().data or [None])[0]
        if not unit or unit.get("department_id") != dept_id:
            continue
        db.table("exam_bookings").update({
            "status": "approved",
            "approved_by": user["id"],
            "approved_at": datetime.now().isoformat(),
        }).eq("id", booking["id"]).execute()
        approved += 1
        student_id = booking["student_id"]
    if approved:
        write_audit_log("batch_approve_exam_bookings", target=f"serial:{serial}", detail={"count": approved})
        if student_id:
            try:
                from notifications import create_notification
                create_notification(
                    user_id=student_id,
                    title="Exam Booking Approved",
                    message=f"All units on Form 1A ({serial}) have been approved by your HOD.",
                    notification_type="success",
                    action_url="/student/exam-bookings",
                )
            except Exception:
                pass
        flash(f"Approved {approved} unit(s) for serial {serial}.", "success")
    else:
        flash("No pending bookings found for that serial in your department.", "warning")
    return redirect(url_for("dept_admin.exam_bookings", status="pending"))


@dept_admin_bp.route("/exam-bookings/<booking_id>/reject", methods=["POST"])
@dept_admin_required
def reject_exam_booking(booking_id):
    db = get_service_client()
    dept_id = _dept_id()
    user = current_user()
    booking = db.table("exam_bookings").select("*").eq("id", booking_id).single().execute().data
    if not booking:
        abort(404)
    unit = db.table("units").select("department_id").eq("id", booking["unit_id"]).single().execute().data
    if not unit or unit["department_id"] != dept_id:
        abort(403)
    reason = request.form.get("rejection_reason", "")
    try:
        db.table("exam_bookings").update({"status": "rejected", "approved_by": user["id"],
            "approved_at": datetime.now().isoformat(), "rejection_reason": reason}).eq("id", booking_id).execute()
        write_audit_log("reject_exam_booking", target=f"booking:{booking_id}")
        from notifications import create_notification
        create_notification(user_id=booking["student_id"], title="Exam Booking Rejected",
            message=f"Your exam booking for {booking.get('exam_date')} was rejected. {reason}",
            notification_type="warning", action_url="/student/exam-bookings")
        flash("Exam booking rejected.", "success")
    except Exception as e:
        flash(f"Error: {e}", "danger")
    return redirect(url_for("dept_admin.exam_bookings"))


# ── Marks ─────────────────────────────────────────────────────────────────────

def _compute_grade(obtained, max_m):
    """Return (percentage, CDACC code) using the TVET CDACC competency scale.
    M 80-100 · P 65-79 · C 50-64 · NYC 0-49."""
    from grading_utils import compute_grade
    return compute_grade(obtained, max_m)


def _fetch_marks(db, dept_id, year, term, class_id, unit_id, trainer_id):
    """
    Fetch formative marks for all classes/units in the department.
    Returns a flat list of dicts ready for the template.
    """
    # Unit IDs belonging to this department
    dept_units = db.table("units").select("id").eq("department_id", dept_id).execute().data or []
    unit_ids = [u["id"] for u in dept_units]
    if not unit_ids:
        return []

    # Assessment definitions
    fa_q = (db.table("formative_assessments")
            .select("id, unit_id, class_id, trainer_id, assessment_type, "
                    "assessment_name, max_marks, year, term, created_at, "
                    "units(name, code), classes(name), "
                    "trainer:user_profiles!formative_assessments_trainer_id_fkey(full_name)")
            .in_("unit_id", unit_ids)
            .eq("year", int(year)))
    if term:       fa_q = fa_q.eq("term",       int(term))
    if class_id:   fa_q = fa_q.eq("class_id",   class_id)
    if unit_id:    fa_q = fa_q.eq("unit_id",    unit_id)
    if trainer_id: fa_q = fa_q.eq("trainer_id", trainer_id)

    try:
        formative_assessments = fa_q.order("created_at", desc=True).execute().data or []
    except Exception as e:
        print(f"[marks] formative_assessments error: {e}")
        return []

    if not formative_assessments:
        return []

    fa_map = {a["id"]: a for a in formative_assessments}
    a_ids  = list(fa_map.keys())

    # Student marks for those assessments
    try:
        fm_rows = (db.table("formative_marks")
                   .select("assessment_id, student_id, marks_obtained, "
                           "student:user_profiles!formative_marks_student_id_fkey"
                           "(full_name, admission_no)")
                   .in_("assessment_id", a_ids)
                   .execute().data or [])
    except Exception as e:
        print(f"[marks] formative_marks error: {e}")
        return []

    rows = []
    for m in fm_rows:
        fa  = fa_map.get(m["assessment_id"], {})
        pct, grade = _compute_grade(m.get("marks_obtained"), fa.get("max_marks", 100))
        rows.append({
            "student":         m.get("student") or {},
            "unit":            fa.get("units")   or {},
            "class_":          fa.get("classes") or {},
            "trainer":         fa.get("trainer") or {},
            "assessment_name": fa.get("assessment_name", ""),
            "assessment_type": fa.get("assessment_type", ""),
            "max_marks":       fa.get("max_marks", 100),
            "marks_obtained":  m.get("marks_obtained"),
            "percentage":      pct,
            "grade":           grade,
            "year":            fa.get("year"),
            "term":            fa.get("term"),
        })

    rows.sort(key=lambda r: (
        r["class_"].get("name", ""),
        r["student"].get("full_name", ""),
        r["unit"].get("name", ""),
        r["assessment_name"],
    ))
    return rows


@dept_admin_bp.route("/marks")
@dept_admin_required
def marks():
    db = get_service_client()
    dept_id    = _dept_id()
    year       = request.args.get("year",       str(datetime.now().year))
    term       = request.args.get("term",       "")
    class_id   = request.args.get("class_id",   "")
    unit_id    = request.args.get("unit_id",    "")
    trainer_id = request.args.get("trainer_id", "")

    classes  = db.table("classes").select("id, name").eq("department_id", dept_id).order("name").execute().data or []
    units    = db.table("units").select("id, name, code").eq("department_id", dept_id).order("name").execute().data or []
    trainers = (db.table("user_profiles").select("id, full_name")
                .eq("role", "trainer").eq("department_id", dept_id)
                .order("full_name").execute().data or [])

    marks_list = _fetch_marks(db, dept_id, year, term, class_id, unit_id, trainer_id)

    # Summary stats
    distinct_students = len({r["student"].get("admission_no") for r in marks_list if r["student"].get("admission_no")})
    pass_count  = sum(1 for r in marks_list if r["grade"] in ("M", "P", "C"))  # Competent & above
    pass_rate   = round(pass_count / len(marks_list) * 100) if marks_list else 0

    return render_template(
        "dept_admin/marks.html",
        marks=marks_list,
        classes=classes, units=units, trainers=trainers,
        year=year, term=term, class_id=class_id, unit_id=unit_id, trainer_id=trainer_id,
        distinct_students=distinct_students,
        pass_rate=pass_rate,
    )


@dept_admin_bp.route("/marks/download-pdf")
@dept_admin_required
def download_marks_pdf():
    db         = get_service_client()
    dept_id    = _dept_id()
    user       = current_user()
    year       = request.args.get("year",       str(datetime.now().year))
    term       = request.args.get("term",       "")
    class_id   = request.args.get("class_id",   "")
    unit_id    = request.args.get("unit_id",    "")
    trainer_id = request.args.get("trainer_id", "")

    marks_list = _fetch_marks(db, dept_id, year, term, class_id, unit_id, trainer_id)

    # Fetch department info
    dept_info = db.table("departments").select("name").eq("id", dept_id).single().execute().data or {}
    dept_name = dept_info.get("name", "")

    # HOD full name
    hod_name  = user.get("full_name", "Head of Department")

    # Resolve class/unit names for report subtitle
    class_name = ""
    unit_name  = ""
    if class_id:
        c = db.table("classes").select("name").eq("id", class_id).limit(1).execute().data or []
        class_name = c[0]["name"] if c else ""
    if unit_id:
        u = db.table("units").select("name", "code").eq("id", unit_id).limit(1).execute().data or []
        if u:
            unit_name = f"{u[0].get('code','')} — {u[0].get('name','')}"

    generated_at = datetime.now().strftime("%d %B %Y at %H:%M")

    # Stats
    total        = len(marks_list)
    pass_count   = sum(1 for m in marks_list if m.get("grade") in ("M", "P", "C"))
    pass_rate    = round(pass_count / total * 100) if total else 0
    avg_pct      = round(sum(m.get("percentage",0) for m in marks_list) / total, 1) if total else 0

    return render_template(
        "dept_admin/marks_pdf.html",
        marks=marks_list,
        year=year, term=term,
        class_name=class_name, unit_name=unit_name,
        dept_name=dept_name,
        hod_name=hod_name,
        generated_at=generated_at,
        total=total, pass_count=pass_count,
        pass_rate=pass_rate, avg_pct=avg_pct,
    )


# ── Trainer Documents ─────────────────────────────────────────────────────────

@dept_admin_bp.route("/trainer-documents")
@dept_admin_required
def trainer_documents():
    db = get_service_client()
    dept_id = _dept_id()
    doc_type   = request.args.get("document_type", "")
    year       = request.args.get("year", str(datetime.now().year))
    term       = request.args.get("term", "")
    trainer_id = request.args.get("trainer_id", "")
    query = (db.table("trainer_documents")
        .select("*, units(name, code, department_id), classes(name), user_profiles(full_name, staff_no, department_id)")
        .eq("academic_year", int(year)))
    if term:       query = query.eq("term", term)
    if doc_type:   query = query.eq("document_type", doc_type)
    if trainer_id: query = query.eq("trainer_id", trainer_id)
    docs = query.order("created_at", desc=True).execute().data or []
    docs = [d for d in docs if (d.get("units") and d.get("units", {}).get("department_id") == dept_id) or (d.get("user_profiles") and d.get("user_profiles", {}).get("department_id") == dept_id)]
    trainers = (db.table("user_profiles").select("id, full_name, staff_no")
        .eq("role", "trainer").eq("department_id", dept_id).order("full_name").execute().data or [])
    return render_template("dept_admin/trainer_documents.html",
                           documents=docs, trainers=trainers,
                           document_type=doc_type, year=year, term=term, trainer_id=trainer_id)


@dept_admin_bp.route("/trainer-document-view/<document_id>")
@dept_admin_required
def view_trainer_document(document_id):
    db = get_service_client()
    dept_id = _dept_id()
    result = (db.table("trainer_documents")
              .select("*, trainer:user_profiles!trainer_documents_trainer_id_fkey(department_id)")
              .eq("id", document_id).limit(1).execute().data or [])
    doc = result[0] if result else None
    if not doc:
        abort(404)
    allowed = (doc.get("trainer") or {}).get("department_id") == dept_id
    if not allowed and doc.get("unit_id"):
        unit = (db.table("units").select("department_id")
                  .eq("id", doc["unit_id"]).limit(1).execute().data or [])
        allowed = unit and unit[0].get("department_id") == dept_id
    if not allowed:
        abort(403)
    file_url = doc.get("file_url", "")
    bucket = "assessment-scripts" if "/assessment-scripts/" in file_url else "documents"
    split_key = f"/{bucket}/"
    storage_path = file_url.split(split_key)[-1] if split_key in file_url else None
    if not storage_path:
        return redirect(file_url)
    try:
        raw = bytes(db.storage.from_(bucket).download(storage_path))
    except Exception:
        abort(404)
    ct = doc.get("file_type") or "application/octet-stream"
    fn = doc.get("file_name") or "document"
    resp = make_response(raw)
    resp.headers["Content-Type"] = ct
    resp.headers["Content-Disposition"] = f"inline; filename=\"{fn}\""
    return resp


# ── Trainee POE ───────────────────────────────────────────────────────────────

@dept_admin_bp.route("/trainee-poe")
@dept_admin_required
def trainee_poe():
    import os
    db = get_service_client()
    dept_id      = _dept_id()
    supabase_url = os.environ.get("SUPABASE_URL", "").strip()

    def _fmt_size(b):
        if not b: return "0 B"
        for u in ["B", "KB", "MB", "GB"]:
            if b < 1024: return f"{b:.1f} {u}"
            b /= 1024
        return f"{b:.1f} GB"

    try:
        rows = (db.table("assessments")
            .select("id, status, script_file_path, script_file_name, script_file_size, "
                    "uploaded_at, assessment_type, assessment_no, term, year, "
                    "student:user_profiles!assessments_student_id_fkey(full_name, admission_no), "
                    "units!inner(name, code, department_id), "
                    "classes(name)")
            .eq("units.department_id", dept_id)
            .order("uploaded_at", desc=True)
            .execute().data or [])
    except Exception as e:
        print(f"[trainee_poe] {e}")
        rows = []

    # Build class → unit → files structure
    folder_map = {}
    for a in rows:
        cls_name  = (a.get("classes") or {}).get("name") or "Uncategorised"
        unit_obj  = a.get("units") or {}
        unit_name = f"{unit_obj.get('code','?')} — {unit_obj.get('name','?')}" if unit_obj.get("name") else "Unknown Unit"
        student   = a.get("student") or {}
        fp        = a.get("script_file_path") or ""
        status    = (a.get("status") or "pending").title()

        file_obj = {
            "id":             str(a.get("id", "")),
            "name":           a.get("script_file_name") or f"{a.get('assessment_type','?')} #{a.get('assessment_no','?')}",
            "url":            f"{supabase_url}/storage/v1/object/public/assessment-scripts/{fp}" if fp else "",
            "status":         status,
            "admissionNumber": student.get("admission_no") or "N/A",
            "studentName":    student.get("full_name") or "Unknown",
            "formattedSize":  _fmt_size(a.get("script_file_size") or 0),
            "size":           a.get("script_file_size") or 0,
            "assessmentType": a.get("assessment_type") or "",
            "assessmentNo":   str(a.get("assessment_no") or ""),
            "term":           str(a.get("term") or ""),
            "year":           str(a.get("year") or ""),
            "uploadedAt":     (a.get("uploaded_at") or "")[:10],
            "className":      cls_name,
            "unitName":       unit_name,
        }
        folder_map.setdefault(cls_name, {}).setdefault(unit_name, []).append(file_obj)

    classes_data = []
    for cls_name in sorted(folder_map):
        units_list = []
        for unit_name in sorted(folder_map[cls_name]):
            files = sorted(folder_map[cls_name][unit_name],
                           key=lambda f: (f["admissionNumber"] == "N/A", f["admissionNumber"]))
            units_list.append({"name": unit_name, "files": files})
        classes_data.append({"name": cls_name, "units": units_list})

    total_size = _fmt_size(sum(a.get("script_file_size") or 0 for a in rows))

    return render_template(
        "dept_admin/trainee_poe.html",
        classes_data=classes_data,
        total_classes=len(folder_map),
        total_units=sum(len(v) for v in folder_map.values()),
        total_files=len(rows),
        total_size=total_size,
    )


# ── Trainees Documents ────────────────────────────────────────────────────────

TRAINEE_DOC_TYPES = [
    ("passport_photo",          "Passport Photo",                             True),
    ("admission_letter",        "Admission Letter",                           True),
    ("medical_form",            "Medical Examination Form",                   True),
    ("personal_data_form",      "Personal Data Form",                         True),
    ("declaration_form",        "Declaration Form",                           True),
    ("kcse_result_slip",        "KCSE Result Slip",                           True),
    ("kcse_certificate",        "KCSE Certificate",                           True),
    ("kcpe_result_slip",        "KCPE Result Slip",                           True),
    ("birth_certificate",       "Birth Certificate",                          True),
    ("national_id",             "National ID",                                True),
    ("guardian_id",             "Guardian ID Copies",                         False),
    ("consent_form",            "Consent Form",                               True),
    ("most_recent_result_slip", "Previous Module Result Slip (Continuing)",   False),
]

# Table that trainee uploads go into (matches routes/student.py)
_TRAINEE_DOCS_TABLE = "student_personal_documents"


def _resolve_doc_url(doc):
    return doc.get("file_url") or ""


def _parse_hod_verification(docs):
    """
    Read overall verification status from student_personal_documents.
    Returns (status_str, comment_str) using the first non-null status found.
    """
    for d in docs.values():
        st = d.get("status") or ""
        if st in ("approved", "rejected", "pending"):
            return st, d.get("rejection_reason") or ""
    return "pending", ""


@dept_admin_bp.route("/trainees-documents")
@dept_admin_required
def trainees_documents():
    db = get_service_client()
    dept_id = _dept_id()
    q = request.args.get("q", "").strip()

    # All students enrolled in any class of this department
    enr = (db.table("enrollments")
           .select("student_id, classes!inner(department_id, name), "
                   "user_profiles!enrollments_student_id_fkey(id, full_name, admission_no, email)")
           .eq("classes.department_id", dept_id)
           .execute().data or [])

    seen = {}
    for e in enr:
        up  = e.get("user_profiles") or {}
        sid = up.get("id") or e.get("student_id")
        if sid and sid not in seen:
            seen[sid] = {
                "id":           sid,
                "full_name":    up.get("full_name", "—"),
                "admission_no": up.get("admission_no", "—"),
                "email":        up.get("email", ""),
                "class_name":   (e.get("classes") or {}).get("name", ""),
            }
    students = list(seen.values())

    if q:
        ql = q.lower()
        students = [s for s in students
                    if ql in s["full_name"].lower() or ql in s["admission_no"].lower()]

    student_ids = [s["id"] for s in students]
    doc_map = {}
    if student_ids:
        try:
            td_rows = (db.table(_TRAINEE_DOCS_TABLE)
                       .select("*")
                       .in_("student_id", student_ids)
                       .execute().data or [])
            for d in td_rows:
                sid = d["student_id"]
                doc_map.setdefault(sid, {})[d["document_type"]] = d
                d["_url"] = _resolve_doc_url(d)
        except Exception as e:
            print(f"[trainees_documents] fetch error: {e}")

    total_required = sum(1 for _, _, req in TRAINEE_DOC_TYPES if req)
    for s in students:
        docs = doc_map.get(s["id"], {})
        s["docs"]        = docs
        s["uploaded"]    = len(docs)
        s["required_ok"] = sum(1 for dt, _, req in TRAINEE_DOC_TYPES if req and dt in docs)
        s["total_req"]   = total_required
        overall_status, _ = _parse_hod_verification(docs)
        s["overall_status"] = overall_status

    students.sort(key=lambda s: s["full_name"])

    return render_template(
        "dept_admin/trainees_documents.html",
        students=students,
        doc_types=TRAINEE_DOC_TYPES,
        q=q,
    )


@dept_admin_bp.route("/trainees-documents/<student_id>")
@dept_admin_required
def trainee_document_detail(student_id):
    db = get_service_client()
    dept_id = _dept_id()

    student = db.table("user_profiles").select(
        "id, full_name, admission_no, email, mobile_number, department_id"
    ).eq("id", student_id).single().execute().data
    if not student:
        flash("Student not found.", "error")
        return redirect(url_for("dept_admin.trainees_documents"))

    if not _student_in_dept(db, student_id, dept_id):
        flash("Student not found in your department.", "error")
        return redirect(url_for("dept_admin.trainees_documents"))

    enr = (db.table("enrollments")
           .select("class_id, classes!inner(name, department_id)")
           .eq("student_id", student_id)
           .eq("classes.department_id", dept_id)
           .execute().data or [])
    class_name = (enr[0].get("classes") or {}).get("name", "") if enr else ""

    td_rows = []
    try:
        td_rows = (db.table(_TRAINEE_DOCS_TABLE)
                   .select("*")
                   .eq("student_id", student_id)
                   .execute().data or [])
    except Exception as e:
        print(f"[trainee_document_detail] fetch error: {e}")

    docs = {}
    for d in td_rows:
        d["_url"] = _resolve_doc_url(d)
        docs[d["document_type"]] = d

    overall_status, hod_comment = _parse_hod_verification(docs)

    return render_template(
        "dept_admin/trainee_document_detail.html",
        student=student,
        class_name=class_name,
        docs=docs,
        doc_types=TRAINEE_DOC_TYPES,
        overall_status=overall_status,
        hod_comment=hod_comment,
    )


@dept_admin_bp.route("/trainees-documents/<student_id>/verify", methods=["POST"])
@dept_admin_required
def verify_trainee_documents(student_id):
    db = get_service_client()
    dept_id = _dept_id()
    if not _student_in_dept(db, student_id, dept_id):
        flash("Student not found in your department.", "error")
        return redirect(url_for("dept_admin.trainees_documents"))

    status  = request.form.get("status", "pending")
    comment = request.form.get("comment", "").strip()

    if status not in ("pending", "approved", "rejected"):
        status = "pending"

    from datetime import datetime as _dt
    verified_at = _dt.utcnow().isoformat()
    admin_id    = current_user().get("id")

    existing = (db.table(_TRAINEE_DOCS_TABLE)
                .select("id")
                .eq("student_id", student_id)
                .execute().data or [])

    errors = []
    for doc in existing:
        try:
            db.table(_TRAINEE_DOCS_TABLE).update({
                "status":           status,
                "rejection_reason": comment or None,
                "verified_by":      admin_id,
                "verified_at":      verified_at,
            }).eq("id", doc["id"]).execute()
        except Exception as e:
            errors.append(str(e))
            print(f"[verify_trainee_documents] update error: {e}")

    if errors:
        flash(f"Some documents could not be updated: {errors[0]}", "warning")

    write_audit_log("verify_trainee_documents",
                    target=f"student:{student_id}",
                    detail={"status": status})
    flash(f"Verification saved — documents marked as {status}.", "success")
    return redirect(url_for("dept_admin.trainee_document_detail", student_id=student_id))


# ── Class List ────────────────────────────────────────────────────────────────

@dept_admin_bp.route("/class-list")
@dept_admin_required
def class_list():
    db = get_service_client()
    dept_id = _dept_id()
    dept = db.table("departments").select("*").eq("id", dept_id).single().execute().data or {}
    class_id = request.args.get("class_id", "")
    classes = db.table("classes").select("*, courses(name)").eq("department_id", dept_id).order("name").execute().data or []
    cls = None
    students = []
    if class_id:
        cls = next((c for c in classes if str(c["id"]) == class_id), None)
        enrollments = (db.table("enrollments")
            .select("*, user_profiles(id, full_name, admission_no, email, mobile_number)")
            .eq("class_id", class_id).execute().data or [])
        students = [e.get("user_profiles", {}) for e in enrollments if e.get("user_profiles")]
        for s in students:
            s["admission_number"] = s.get("admission_no", "")
    return render_template("dept_admin/class_list.html",
                           dept=dept, classes=classes, cls=cls,
                           students=students, class_id=class_id)


@dept_admin_bp.route("/class-list/pdf")
@dept_admin_required
def class_list_pdf():
    db = get_service_client()
    dept_id = _dept_id()
    dept = db.table("departments").select("*").eq("id", dept_id).single().execute().data or {}
    class_id = request.args.get("class_id", "")
    cls = None
    students = []
    if class_id:
        cls = db.table("classes").select("*, courses(name)").eq("id", class_id).single().execute().data
        enrollments = (db.table("enrollments")
            .select("*, user_profiles(id, full_name, admission_no, email, mobile_number)")
            .eq("class_id", class_id).execute().data or [])
        students = [e.get("user_profiles", {}) for e in enrollments if e.get("user_profiles")]
        for s in students:
            s["admission_number"] = s.get("admission_no", "")
    return render_template("dept_admin/class_list_pdf.html",
                           cls=cls, dept_name=dept.get("name",""),
                           students=students, date_gen=datetime.now().strftime("%d %b %Y"))


# ── Trainee Attendance Search ─────────────────────────────────────────────────

@dept_admin_bp.route("/trainee-search")
@dept_admin_required
def trainee_search():
    db = get_service_client()
    dept_id = _dept_id()
    q = request.args.get("q", "").strip()
    student_id = request.args.get("student_id", "")
    unit_id = request.args.get("unit_id", "")
    students = []
    student = None
    records = []
    units_list = []
    summary = {}

    if q:
        students = (db.table("user_profiles").select("*")
            .eq("role", "student").eq("department_id", dept_id)
            .or_(f"admission_no.ilike.%{q}%,full_name.ilike.%{q}%")
            .limit(20).execute().data or [])
        for s in students:
            s["class_name"] = _student_class_name(db, s["id"])

    if student_id:
        allowed = _dept_student_ids(db, dept_id)
        if student_id not in allowed:
            flash("Trainee not found in your department.", "error")
            return redirect(url_for("dept_admin.trainee_search", q=q))
        student = db.table("user_profiles").select("*").eq("id", student_id).single().execute().data
        if student:
            student["class_name"] = _student_class_name(db, student_id)
            units_list = (db.table("attendance").select("unit_id, units!inner(id, name, code)")
                .eq("student_id", student_id)
                .execute().data or [])
            seen = set()
            deduped = []
            for u in units_list:
                uid = u["unit_id"]
                if uid not in seen:
                    seen.add(uid)
                    deduped.append(u.get("units", u))
            units_list = deduped

    if student_id and unit_id:
        records = (db.table("attendance")
            .select("*, units(name, code), trainers:user_profiles!attendance_trainer_id_fkey(full_name)")
            .eq("student_id", student_id).eq("unit_id", unit_id)
            .order("attendance_date", desc=True).execute().data or [])

        total = len(records)
        present = sum(1 for a in records if a["status"] == "present")
        unit_info = db.table("units").select("name, code").eq("id", unit_id).single().execute().data or {}
        summary = {"total": total, "present": present, "absent": total - present,
                   "pct": round(present / total * 100, 1) if total else 0,
                   "unit_code": unit_info.get("code", ""),
                   "unit_name": unit_info.get("name", "")}

    return render_template("dept_admin/trainee_search.html",
                           query=q, students=students, student=student,
                           student_id=student_id, unit_id=unit_id,
                           units_list=units_list, records=records, summary=summary)


@dept_admin_bp.route("/trainee-report-pdf")
@dept_admin_required
def trainee_report_pdf():
    db = get_service_client()
    dept_id = _dept_id()
    dept = db.table("departments").select("*").eq("id", dept_id).single().execute().data or {}
    student_id = request.args.get("student_id", "")
    unit_id = request.args.get("unit_id", "")
    student = None
    records = []
    summary = {}
    if student_id and unit_id:
        allowed = _dept_student_ids(db, dept_id)
        if student_id not in allowed:
            flash("Trainee not found in your department.", "error")
            return redirect(url_for("dept_admin.trainee_search"))
        student = db.table("user_profiles").select("*").eq("id", student_id).single().execute().data
        if student:
            student["class_name"] = _student_class_name(db, student_id)
            if not student.get("admission_number"):
                student["admission_number"] = student.get("admission_no", "")
            records = (db.table("attendance")
                .select("*, units(name, code), trainers:user_profiles!attendance_trainer_id_fkey(full_name)")
                .eq("student_id", student_id).eq("unit_id", unit_id)
                .order("attendance_date", desc=True).execute().data or [])
            unit_info = db.table("units").select("name, code").eq("id", unit_id).single().execute().data or {}
            total = len(records)
            present = sum(1 for a in records if a["status"] == "present")
            summary = {"total": total, "present": present, "absent": total - present,
                       "pct": round(present / total * 100, 1) if total else 0,
                       "unit_code": unit_info.get("code", ""),
                       "unit_name": unit_info.get("name", "")}
    return render_template("dept_admin/trainee_report_pdf.html",
                           dept_name=dept.get("name",""), student=student,
                           records=records, summary=summary,
                           generated=datetime.now().strftime("%d %b %Y"))


# ── Assessment Sheet ──────────────────────────────────────────────────────────

@dept_admin_bp.route("/assessment-sheet")
@dept_admin_required
def assessment_sheet():
    db = get_service_client()
    dept_id = _dept_id()
    class_id = request.args.get("class_id", "")
    unit_id  = request.args.get("unit_id", "")
    year     = request.args.get("year", "")
    term     = request.args.get("term", "")
    min_pct  = int(request.args.get("min_pct", 80))

    classes = db.table("classes").select("id, name").eq("department_id", dept_id).order("name").execute().data or []
    units   = db.table("units").select("id, name, code").eq("department_id", dept_id).order("name").execute().data or []
    cls = None
    unit = None
    eligible = []
    term_label = f"Term {term}" if term else "All Terms"

    if class_id and unit_id:
        cls = db.table("classes").select("id, name").eq("id", class_id).single().execute().data
        unit = db.table("units").select("id, name, code").eq("id", unit_id).single().execute().data
        enrollments = (db.table("enrollments")
            .select("student_id, user_profiles!inner(id, full_name, admission_no)")
            .eq("class_id", class_id).execute().data or [])
        for e in enrollments:
            student = e.get("user_profiles") or {}
            sid = student.get("id")
            if not sid:
                continue
            att_query = db.table("attendance").select("status").eq("student_id", sid).eq("unit_id", unit_id)
            if year:
                att_query = att_query.eq("year", int(year))
            if term:
                att_query = att_query.eq("term", int(term))
            att_records = att_query.execute().data or []
            total = len(att_records)
            if total == 0:
                continue
            present = sum(1 for a in att_records if a["status"] == "present")
            pct = round(present / total * 100, 1)
            if pct >= min_pct:
                eligible.append({
                    "admission_number": student.get("admission_no", ""),
                    "full_name": student.get("full_name", ""),
                    "present": present,
                    "total": total,
                    "pct": pct
                })
        eligible.sort(key=lambda x: x["full_name"])

    return render_template("dept_admin/assessment_sheet.html",
                           classes=classes, units=units,
                           class_id=class_id, unit_id=unit_id,
                           year=year, term=term, min_pct=min_pct,
                           cls=cls, unit=unit, term_label=term_label,
                           eligible=eligible)


@dept_admin_bp.route("/assessment-sheet/pdf")
@dept_admin_required
def assessment_sheet_pdf():
    db = get_service_client()
    dept_id = _dept_id()
    dept = db.table("departments").select("*").eq("id", dept_id).single().execute().data or {}
    class_id = request.args.get("class_id", "")
    unit_id  = request.args.get("unit_id", "")
    year     = request.args.get("year", "")
    term     = request.args.get("term", "")
    min_pct  = int(request.args.get("min_pct", 80))

    cls = None
    unit = None
    eligible = []
    term_label = f"Term {term}" if term else "All Terms"

    if class_id and unit_id:
        cls = db.table("classes").select("id, name").eq("id", class_id).single().execute().data
        unit = db.table("units").select("id, name, code").eq("id", unit_id).single().execute().data
        enrollments = (db.table("enrollments")
            .select("student_id, user_profiles!inner(id, full_name, admission_no)")
            .eq("class_id", class_id).execute().data or [])
        for e in enrollments:
            student = e.get("user_profiles") or {}
            sid = student.get("id")
            if not sid:
                continue
            att_query = db.table("attendance").select("status").eq("student_id", sid).eq("unit_id", unit_id)
            if year:
                att_query = att_query.eq("year", int(year))
            if term:
                att_query = att_query.eq("term", int(term))
            att_records = att_query.execute().data or []
            total = len(att_records)
            if total == 0:
                continue
            present = sum(1 for a in att_records if a["status"] == "present")
            pct = round(present / total * 100, 1)
            if pct >= min_pct:
                eligible.append({
                    "admission_number": student.get("admission_no", ""),
                    "full_name": student.get("full_name", ""),
                    "present": present,
                    "total": total,
                    "pct": pct
                })
        eligible.sort(key=lambda x: x["full_name"])

    return render_template("dept_admin/assessment_sheet_pdf.html",
                           cls=cls, unit=unit, dept_name=dept.get("name",""),
                           term_label=term_label, year=year, min_pct=min_pct,
                           eligible=eligible, date_gen=datetime.now().strftime("%d %b %Y"))


# ── Credentials ───────────────────────────────────────────────────────────────

@dept_admin_bp.route("/credentials", methods=["GET", "POST"])
@dept_admin_required
def credentials():
    from auth_utils import reset_user_password, generate_temp_password
    db = get_service_client()
    dept_id = _dept_id()
    tab = request.args.get("tab", "trainers")
    search_t = request.args.get("search_t", "")
    search_s = request.args.get("search_s", "")
    filter_class = request.args.get("filter_class", "")

    if request.method == "POST":
        action = request.form.get("action")
        uid = request.form.get("user_id")
        # Verify the target belongs to this department (dept isolation)
        target = None
        if uid:
            res = (db.table("user_profiles")
                   .select("id, full_name, role, department_id, admission_no, staff_no, email")
                   .eq("id", uid).limit(1).execute().data or [])
            target = res[0] if res else None

        if not target or target.get("department_id") != dept_id:
            flash("User not found in your department.", "error")
        elif action == "set_password":
            new_pw = request.form.get("password", "").strip()
            if len(new_pw) < 6:
                flash("Password must be at least 6 characters.", "error")
            else:
                ok, m = reset_user_password(uid, new_pw)
                if ok:
                    flash(f"Password for {target['full_name']} set to: {new_pw}", "success")
                else:
                    flash(m, "error")
        elif action == "reset_password":
            new_pw = generate_temp_password()
            ok, m = reset_user_password(uid, new_pw)
            if ok:
                flash(f"New temporary password for {target['full_name']}: {new_pw}", "success")
            else:
                flash(m, "error")
        return redirect(url_for("dept_admin.credentials", tab=tab, search_t=search_t,
                                search_s=search_s, filter_class=filter_class))

    tq = db.table("user_profiles").select("id, full_name, email, staff_no, is_active, must_change_password, departments(name)")
    tq = tq.eq("role", "trainer").eq("department_id", dept_id)
    if search_t:
        tq = tq.or_(f"full_name.ilike.%{search_t}%,staff_no.ilike.%{search_t}%,email.ilike.%{search_t}%")
    trainers_list = tq.order("full_name").execute().data or []

    sq = db.table("user_profiles").select(
        "id, full_name, admission_no, email, is_active, must_change_password")
    sq = sq.eq("role", "student").eq("department_id", dept_id)
    if search_s:
        sq = sq.or_(f"full_name.ilike.%{search_s}%,admission_no.ilike.%{search_s}%")
    students_list = sq.order("full_name").execute().data or []
    _attach_student_classes(db, students_list)

    if filter_class:
        students_list = [
            s for s in students_list
            if s.get("class_obj") and str(s["class_obj"].get("id")) == str(filter_class)
        ]

    classes_list = db.table("classes").select("id, name").eq("department_id", dept_id).order("name").execute().data or []
    return render_template("dept_admin/credentials.html",
                           tab=tab, search_t=search_t, search_s=search_s,
                           filter_class=filter_class,
                           trainers_list=trainers_list, students_list=students_list,
                           classes_list=classes_list)


# ── Import Data ───────────────────────────────────────────────────────────────

@dept_admin_bp.route("/import", methods=["GET", "POST"])
@dept_admin_required
def import_data():
    db = get_service_client()
    dept_id = _dept_id()
    results = []
    error   = None
    classes = db.table("classes").select("id, name").eq("department_id", dept_id).order("name").execute().data or []
    if request.method == "POST":
        import_type = request.form.get("import_type", "students")
        class_id    = request.form.get("class_id", "")
        if "file" not in request.files or request.files["file"].filename == "":
            error = "Please select an Excel (.xlsx) file."
        else:
            file = request.files["file"]
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
                required = ["admission_no", "full_name", "email"]
                missing  = [h for h in required if h not in headers]
                if missing:
                    error = f"Missing columns: {', '.join(missing)}"
                else:
                    adm_i   = headers.index("admission_no")
                    name_i  = headers.index("full_name")
                    email_i = headers.index("email")
                    phone_i = headers.index("mobile_number") if "mobile_number" in headers else None
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or row[adm_i] is None:
                            continue
                        adm   = str(row[adm_i]).strip()
                        name  = str(row[name_i]).strip()
                        email = str(row[email_i]).strip().lower()
                        phone = str(row[phone_i]).strip() if phone_i is not None and row[phone_i] else ""
                        dup = db.table("user_profiles").select("id").eq("admission_no", adm).execute().data
                        if dup:
                            results.append({"row": adm, "status": "skipped", "msg": "Already exists"})
                            continue
                        try:
                            from auth_utils import create_student_auth_user
                            pwd = _gen_password()
                            uid = create_student_auth_user(admission_no=adm, password=pwd,
                                email=email, full_name=name, department_id=dept_id, class_id=None)
                            if class_id:
                                db.table("enrollments").insert({"student_id": uid, "class_id": class_id}).execute()
                            if phone:
                                db.table("user_profiles").update({"mobile_number": phone}).eq("id", uid).execute()
                            results.append({"row": adm, "status": "created", "msg": f"Password: {pwd}"})
                        except Exception as exc:
                            results.append({"row": adm, "status": "error", "msg": str(exc)[:80]})
                wb.close()
                if not error:
                    created = sum(1 for r in results if r["status"] == "created")
                    flash(f"Import complete. {created} students created.", "success")
            except Exception as exc:
                error = f"Error reading file: {exc}"
    result_summary = None
    if results:
        result_summary = {
            "success": sum(1 for r in results if r["status"] == "created"),
            "errors": [r["msg"] for r in results if r["status"] == "error"]
        }
    return render_template("dept_admin/import.html", classes=classes, results=results, result=result_summary, error=error)


# ── Companies ─────────────────────────────────────────────────────────────────

INDUSTRIES = ['Electrical Engineering','Mechanical Engineering','Information Technology',
    'Civil Engineering','Automotive Engineering','Hospitality','Business Management',
    'Health Sciences','Agriculture','Construction','Manufacturing','Other']


@dept_admin_bp.route("/companies")
@dept_admin_required
def companies():
    db = get_service_client()
    dept_id = _dept_id()
    industry = request.args.get("industry", "")
    # Department-scoped list — edit/delete only works for own department's companies
    query = db.table("companies").select("*, departments(name)").eq("department_id", dept_id)
    if industry:
        query = query.eq("industry_classification", industry)
    companies_list = query.order("name").execute().data or []
    departments = db.table("departments").select("id, name").order("name").execute().data or []
    return render_template("dept_admin/companies.html",
                           companies=companies_list, industries=INDUSTRIES,
                           industry=industry, departments=departments)


@dept_admin_bp.route("/companies/<company_id>")
@dept_admin_required
def get_company(company_id):
    db = get_service_client()
    dept_id = _dept_id()
    company = db.table("companies").select("*").eq("id", company_id).single().execute().data
    if not company or company.get("department_id") != dept_id:
        return {"error": "Not found"}, 404
    return {k: company.get(k) for k in ("id","name","industry_classification","address","city",
        "phone_number","email","website","latitude","longitude","geofence_radius_meters",
        "available_slots","contact_person","contact_phone","contact_email","description","is_active")}


@dept_admin_bp.route("/companies/add", methods=["POST"])
@dept_admin_required
def add_company():
    db = get_service_client()
    dept_id = _dept_id()
    user = current_user()
    name = request.form.get("name", "").strip()
    industry_classification = request.form.get("industry_classification", "")
    if not name or not industry_classification:
        flash("Name and industry are required.", "error")
        return redirect(url_for("dept_admin.companies"))
    try:
        lat = request.form.get("latitude")
        lng = request.form.get("longitude")
        db.table("companies").insert({
            "name": name, "industry_classification": industry_classification,
            "address": request.form.get("address") or None,
            "city": request.form.get("city") or None,
            "phone_number": request.form.get("phone_number") or None,
            "email": request.form.get("email") or None,
            "website": request.form.get("website") or None,
            "latitude": float(lat) if lat else None,
            "longitude": float(lng) if lng else None,
            "geofence_radius_meters": int(request.form.get("geofence_radius_meters") or 300),
            "available_slots": int(request.form.get("available_slots") or 0),
            "department_id": dept_id,
            "contact_person": request.form.get("contact_person") or None,
            "contact_phone": request.form.get("contact_phone") or None,
            "contact_email": request.form.get("contact_email") or None,
            "description": request.form.get("description") or None,
            "created_by": user["id"]
        }).execute()
        write_audit_log("add_company", target=name)
        flash("Company added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("dept_admin.companies"))


@dept_admin_bp.route("/companies/<company_id>/edit", methods=["POST"])
@dept_admin_required
def edit_company(company_id):
    db = get_service_client()
    dept_id = _dept_id()
    company = db.table("companies").select("department_id").eq("id", company_id).single().execute().data
    if not company or company.get("department_id") != dept_id:
        abort(403)
    try:
        lat = request.form.get("latitude")
        lng = request.form.get("longitude")
        update = {
            "name": request.form.get("name"),
            "industry_classification": request.form.get("industry_classification"),
            "address": request.form.get("address") or None,
            "city": request.form.get("city") or None,
            "phone_number": request.form.get("phone_number") or None,
            "email": request.form.get("email") or None,
            "website": request.form.get("website") or None,
            "geofence_radius_meters": int(request.form.get("geofence_radius_meters") or 300),
            "available_slots": int(request.form.get("available_slots") or 0),
            "contact_person": request.form.get("contact_person") or None,
            "contact_phone": request.form.get("contact_phone") or None,
            "contact_email": request.form.get("contact_email") or None,
            "description": request.form.get("description") or None,
            "is_active": request.form.get("is_active") == "on"
        }
        if lat: update["latitude"] = float(lat)
        if lng: update["longitude"] = float(lng)
        db.table("companies").update(update).eq("id", company_id).execute()
        write_audit_log("edit_company", target=f"company:{company_id}")
        flash("Company updated.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("dept_admin.companies"))


@dept_admin_bp.route("/companies/<company_id>/delete", methods=["POST"])
@dept_admin_required
def delete_company(company_id):
    db = get_service_client()
    dept_id = _dept_id()
    company = db.table("companies").select("department_id").eq("id", company_id).single().execute().data
    if not company or company.get("department_id") != dept_id:
        abort(403)
    try:
        db.table("companies").delete().eq("id", company_id).execute()
        write_audit_log("delete_company", target=f"company:{company_id}")
        flash("Company deleted.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("dept_admin.companies"))


# ── Course Applications (Public Pre-Registration) ──────────────────────────────

@dept_admin_bp.route("/applications")
@dept_admin_required
def applications():
    db = get_service_client()
    dept_id = _dept_id()
    try:
        applications = (db.table("course_applications")
            .select("*")
            .eq("department_id", dept_id)
            .order("created_at", desc=True)
            .execute().data or [])
    except Exception:
        applications = []
    return render_template("dept_admin/applications.html", applications=applications)


@dept_admin_bp.route("/applications/<app_id>/review", methods=["POST"])
@dept_admin_required
def review_application(app_id):
    db = get_service_client()
    dept_id = _dept_id()
    action = request.form.get("action", "")
    notes = request.form.get("notes", "").strip()

    app = db.table("course_applications").select("*").eq("id", app_id).single().execute().data
    if not app or app.get("department_id") != dept_id:
        abort(403)

    status_map = {"approve": "approved", "reject": "rejected"}
    new_status = status_map.get(action)
    if not new_status:
        flash("Invalid action.", "error")
        return redirect(url_for("dept_admin.applications"))

    try:
        db.table("course_applications").update({
            "status": new_status,
            "reviewed_at": datetime.utcnow().isoformat(),
            "reviewed_by": current_user()["id"],
            "review_notes": notes,
        }).eq("id", app_id).execute()
        write_audit_log("review_course_application", target=f"app:{app_id},status:{new_status}")
        flash(f"Application {new_status}.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("dept_admin.applications"))


# ── Digital Logbook Review ───────────────────────────────────────────────────

@dept_admin_bp.route("/logbooks")
@dept_admin_required
def logbooks():
    """View all digital logbook entries for students in this department."""
    import os
    from datetime import date as _date
    db           = get_service_client()
    dept_id      = _dept_id()
    supabase_url = os.environ.get("SUPABASE_URL", "").strip()

    status_filter = request.args.get("status", "")
    adm_filter    = request.args.get("admission_no", "").strip().upper()
    year_filter   = request.args.get("year",  str(_date.today().year)).strip()
    period_filter = request.args.get("period", "").strip()   # 1=Jan-Apr 2=May-Aug 3=Sep-Dec

    PERIOD_RANGES = {
        "1": (f"{year_filter}-01-01", f"{year_filter}-04-30"),
        "2": (f"{year_filter}-05-01", f"{year_filter}-08-31"),
        "3": (f"{year_filter}-09-01", f"{year_filter}-12-31"),
    }
    PERIOD_LABELS = {
        "1": "January – April",
        "2": "May – August",
        "3": "September – December",
    }

    # ── Students via enrollments (reliable across all setups) ─────────────────
    enr = (db.table("enrollments")
           .select("student_id, classes!inner(department_id)")
           .eq("classes.department_id", dept_id)
           .execute().data or [])
    student_ids = list({e["student_id"] for e in enr})

    records = []
    if student_ids:
        # ── Step 1: fetch logbook entries (no risky FK-name joins) ─────────────
        query = (db.table("digital_logbook")
                   .select("id, student_id, attachment_id, log_date, entry_time, "
                           "tasks_performed, skills_applied, hours_worked, "
                           "challenges_encountered, achievements, "
                           "mentor_approval_status, mentor_comments, "
                           "trainer_comments, evidence_urls, created_at, "
                           "student:user_profiles!digital_logbook_student_id_fkey"
                           "(full_name, admission_no)")
                   .in_("student_id", student_ids)
                   .order("log_date", desc=True)
                   .limit(1000))

        if status_filter:
            query = query.eq("mentor_approval_status", status_filter)

        # Date range on log_date — only apply when explicitly set
        if period_filter and period_filter in PERIOD_RANGES:
            d0, d1 = PERIOD_RANGES[period_filter]
            query = query.gte("log_date", d0).lte("log_date", d1)
        elif year_filter and year_filter != "all":
            query = query.gte("log_date", f"{year_filter}-01-01").lte("log_date", f"{year_filter}-12-31")

        records = query.execute().data or []

        if adm_filter:
            records = [r for r in records
                       if adm_filter in (r.get("student") or {}).get("admission_no", "").upper()]

        # ── Step 2: enrich with company name via attachment_id (best-effort) ──
        att_ids = list({r["attachment_id"] for r in records if r.get("attachment_id")})
        att_map = {}
        if att_ids:
            try:
                att_rows = (db.table("industrial_attachments")
                            .select("id, start_date, end_date, companies(name)")
                            .in_("id", att_ids)
                            .execute().data or [])
                for a in att_rows:
                    att_map[a["id"]] = a
            except Exception:
                pass

        for entry in records:
            entry["_attachment"] = att_map.get(entry.get("attachment_id"), {})
            ev_paths = entry.get("evidence_urls") or []
            entry["_evidence"] = [
                {"url":  f"{supabase_url}/storage/v1/object/public/assessment-evidence/{p}",
                 "ext":  p.rsplit(".", 1)[-1].lower() if "." in p else "bin",
                 "name": p.rsplit("/", 1)[-1]}
                for p in ev_paths if p
            ]

    return render_template("dept_admin/logbooks.html",
                           logbooks=records,
                           status_filter=status_filter,
                           adm_filter=adm_filter,
                           year_filter=year_filter,
                           period_filter=period_filter,
                           PERIOD_LABELS=PERIOD_LABELS)


@dept_admin_bp.route("/logbooks/<log_id>/review", methods=["POST"])
@dept_admin_required
def review_logbook(log_id):
    db      = get_service_client()
    user    = current_user()
    dept_id = _dept_id()
    action  = request.form.get("action", "")
    comment = (request.form.get("comment") or "").strip()

    if action not in ("approve", "reject"):
        flash("Invalid action.", "error")
        return redirect(url_for("dept_admin.logbooks"))

    # Enforce dept isolation
    log_row = (db.table("digital_logbook")
                 .select("id, student_id")
                 .eq("id", log_id)
                 .limit(1)
                 .execute().data or [])
    if not log_row:
        flash("Log entry not found.", "error")
        return redirect(url_for("dept_admin.logbooks"))

    student_ok = (db.table("user_profiles")
                    .select("id")
                    .eq("id", log_row[0]["student_id"])
                    .eq("department_id", dept_id)
                    .limit(1)
                    .execute().data or [])
    if not student_ok:
        abort(403)

    new_status = "approved" if action == "approve" else "rejected"
    payload = {
        "mentor_approval_status": new_status,
        "mentor_approved_by":     user["id"],
        "mentor_approved_at":     datetime.utcnow().isoformat(),
    }
    if comment:
        payload["mentor_comments"] = comment

    try:
        db.table("digital_logbook").update(payload).eq("id", log_id).execute()
        write_audit_log("review_logbook", target=f"log:{log_id}",
                        detail={"action": action})
        flash(f"Log entry {new_status}.", "success")
    except Exception as exc:
        flash(f"Error: {exc}", "error")

    return redirect(url_for("dept_admin.logbooks",
                            status=request.form.get("status_filter", "")))


# ── GIS Placement Tracking & Digital Logbook ─────────────────────────────────

@dept_admin_bp.route("/attachments")
@dept_admin_required
def attachments():
    db = get_service_client()
    dept_id = _dept_id()
    status_filter = (request.args.get("status") or "pending").strip()

    student_ids = _dept_student_ids(db, dept_id)
    attachments = []
    stats = {"total": 0, "pending": 0, "approved": 0, "rejected": 0, "active": 0}

    if student_ids:
        attachments = (db.table("industrial_attachments")
            .select("*, "
                    "companies(name, address, contact_person, contact_phone), "
                    "units(name, code), "
                    "student:user_profiles!industrial_attachments_student_id_fkey"
                    "(full_name, admission_no, mobile_number)")
            .in_("student_id", student_ids)
            .order("created_at", desc=True)
            .limit(300)
            .execute().data or [])

        for item in attachments:
            letter_status = item.get("acceptance_letter_status") or "pending"
            item["_letter_status"] = letter_status
            stats["total"] += 1
            if letter_status in stats:
                stats[letter_status] += 1
            if (item.get("status") or "").lower() == "active":
                stats["active"] += 1

        if status_filter and status_filter != "all":
            attachments = [
                item for item in attachments
                if item.get("_letter_status") == status_filter
            ]

    return render_template(
        "dept_admin/attachments.html",
        attachments=attachments,
        status_filter=status_filter,
        stats=stats,
    )


@dept_admin_bp.route("/attachments/<att_id>/review", methods=["POST"])
@dept_admin_required
def review_attachment(att_id):
    db = get_service_client()
    dept_id = _dept_id()
    user = current_user()
    decision = (request.form.get("decision") or "").strip().lower()
    comment = (request.form.get("comment") or "").strip()

    if decision not in {"approve", "reject"}:
        flash("Invalid attachment review decision.", "error")
        return redirect(url_for("dept_admin.attachments"))

    student_ids = _dept_student_ids(db, dept_id)
    if not student_ids:
        flash("No trainees are linked to your department.", "warning")
        return redirect(url_for("dept_admin.attachments"))

    rows = (db.table("industrial_attachments")
              .select("id, student_id, status, acceptance_letter_url, companies(name)")
              .eq("id", att_id)
              .in_("student_id", student_ids)
              .limit(1)
              .execute().data or [])
    if not rows:
        flash("Attachment record not found in your department.", "error")
        return redirect(url_for("dept_admin.attachments"))

    attachment = rows[0]
    if not attachment.get("acceptance_letter_url"):
        flash("The trainee has not uploaded an official acceptance letter yet.", "warning")
        return redirect(url_for("dept_admin.attachments"))

    approved = decision == "approve"
    payload = {
        "acceptance_letter_status": "approved" if approved else "rejected",
        "dept_review_comments": comment or None,
        "dept_reviewed_by": user["id"],
        "dept_reviewed_at": datetime.utcnow().isoformat(),
        "status": "approved" if approved else "rejected",
        "approved_by": user["id"] if approved else None,
        "approved_at": datetime.utcnow().isoformat() if approved else None,
    }

    try:
        db.table("industrial_attachments").update(payload).eq("id", att_id).execute()
        create_notification(
            user_id=attachment["student_id"],
            title="Attachment Letter Approved" if approved else "Attachment Letter Needs Attention",
            message=(
                f"Your industrial attachment acceptance letter for {(attachment.get('companies') or {}).get('name', 'the selected company')} "
                f"has been approved by your department."
                if approved else
                f"Your industrial attachment acceptance letter for {(attachment.get('companies') or {}).get('name', 'the selected company')} was not approved. Please review the department comments and resubmit if needed."
            ),
            notification_type="success" if approved else "warning",
            action_url="/student/industrial-attachment",
        )
        write_audit_log(
            "review_attachment_acceptance_letter",
            target=f"attachment:{att_id}",
            detail={"decision": decision, "comment": comment},
        )
        flash(
            "Acceptance letter approved and the attachment has moved to department-approved status."
            if approved else
            "Acceptance letter review recorded. The trainee has been notified.",
            "success" if approved else "warning",
        )
    except Exception as exc:
        flash(f"Could not save attachment review: {exc}", "error")

    return redirect(url_for("dept_admin.attachments", status=request.form.get("status_filter", "pending")))


@dept_admin_bp.route("/gis-tracking")
@dept_admin_required
def gis_tracking():
    from datetime import date as _date, datetime as _dt
    import os
    db      = get_service_client()
    dept_id = _dept_id()

    # ── Filters ────────────────────────────────────────────────────────────────
    period_filter = request.args.get("period", "").strip()   # 1=Jan-Apr, 2=May-Aug, 3=Sep-Dec
    year_filter   = request.args.get("year",   str(_date.today().year)).strip()
    PERIOD_RANGES = {
        "1": (f"{year_filter}-01-01", f"{year_filter}-04-30"),
        "2": (f"{year_filter}-05-01", f"{year_filter}-08-31"),
        "3": (f"{year_filter}-09-01", f"{year_filter}-12-31"),
    }
    PERIOD_LABELS = {"1": "January – April", "2": "May – August", "3": "September – December"}

    # ── Dept students ──────────────────────────────────────────────────────────
    student_ids = _dept_student_ids(db, dept_id)

    # ── All placements (no status filter — show everything submitted) ──────────
    placements = []
    try:
        if student_ids:
            q = (db.table("industrial_attachments")
                 .select("*, "
                         "companies(name, address, latitude, longitude, "
                         "  contact_person, contact_phone, city, "
                         "  industry_classification, geofence_radius_meters), "
                         "units(name, code), "
                         "student:user_profiles!industrial_attachments_student_id_fkey"
                         "(full_name, admission_no, mobile_number)")
                 .in_("student_id", student_ids)
                 .order("start_date", desc=True))

            if period_filter and period_filter in PERIOD_RANGES:
                d0, d1 = PERIOD_RANGES[period_filter]
                q = q.gte("start_date", d0).lte("start_date", d1)
            elif year_filter:
                q = q.gte("start_date", f"{year_filter}-01-01").lte("start_date", f"{year_filter}-12-31")

            placements = q.execute().data or []
    except Exception as e:
        flash(f"Error loading placements: {e}", "warning")

    # ── Course of study per student (enrollment → class/dept) ─────────────────
    course_map = {}
    try:
        if student_ids:
            enr_rows = (db.table("enrollments")
                        .select("student_id, classes(name, departments(name))")
                        .in_("student_id", student_ids)
                        .execute().data or [])
            for e in enr_rows:
                sid = e.get("student_id")
                if not sid or sid in course_map:
                    continue
                cls  = e.get("classes") or {}
                dept = cls.get("departments") or {}
                dept_name  = dept.get("name", "")
                class_name = cls.get("name", "")
                if dept_name and class_name:
                    course_map[sid] = f"{dept_name} – {class_name}"
                else:
                    course_map[sid] = dept_name or class_name or "—"
    except Exception:
        pass
    for p in placements:
        p["_course"] = course_map.get(p.get("student_id"), "—")

    # ── Compute days spent for each placement ──────────────────────────────────
    today = _date.today()
    for p in placements:
        try:
            sd = _date.fromisoformat(p["start_date"]) if p.get("start_date") else None
            ed_raw = p.get("end_date")
            ed = _date.fromisoformat(ed_raw) if ed_raw else None
            if sd:
                end = min(ed, today) if ed else today
                p["_days_spent"]  = max(0, (end - sd).days)
                p["_total_days"]  = (ed - sd).days if ed else None
                p["_is_ongoing"]  = not ed or ed >= today
            else:
                p["_days_spent"] = p["_total_days"] = 0
                p["_is_ongoing"] = False
        except Exception:
            p["_days_spent"] = p["_total_days"] = 0
            p["_is_ongoing"] = False

    # ── Live locations ─────────────────────────────────────────────────────────
    live_locations = []
    try:
        if student_ids:
            loc_rows = (db.table("location_logs")
                .select("student_id, latitude, longitude, check_in_time, "
                        "check_out_time, accuracy_meters, is_within_geofence")
                .in_("student_id", student_ids)
                .not_.is_("latitude", "null")
                .order("check_in_time", desc=True)
                .execute().data or [])
            seen = set()
            for loc in loc_rows:
                sid = loc["student_id"]
                if sid not in seen:
                    seen.add(sid)
                    live_locations.append(loc)
    except Exception as e:
        print(f"[gis_tracking] location_logs: {e}")

    student_map = {p.get("student_id"): p.get("student") or {} for p in placements}
    for loc in live_locations:
        loc["student"] = student_map.get(loc["student_id"], {})

    # ── Digital Logbook ────────────────────────────────────────────────────────
    supabase_url  = os.environ.get("SUPABASE_URL", "").strip()
    logbook_entries = []
    logbook_error   = None
    try:
        if student_ids:
            logbook_entries = (db.table("digital_logbook")
                .select("id, student_id, log_date, entry_time, tasks_performed, "
                        "skills_applied, hours_worked, challenges_encountered, "
                        "achievements, mentor_approval_status, mentor_comments, "
                        "trainer_comments, evidence_urls, created_at, "
                        "student:user_profiles!digital_logbook_student_id_fkey"
                        "(full_name, admission_no), "
                        "attachment:industrial_attachments!digital_logbook_attachment_id_fkey"
                        "(companies(name))")
                .in_("student_id", student_ids)
                .order("log_date", desc=True).limit(500)
                .execute().data or [])
        for entry in logbook_entries:
            ev_paths = entry.get("evidence_urls") or []
            entry["_evidence"] = [
                {"url": f"{supabase_url}/storage/v1/object/public/assessment-evidence/{p}",
                 "ext": p.rsplit(".", 1)[-1].lower() if "." in p else "bin",
                 "name": p.rsplit("/", 1)[-1]}
                for p in ev_paths if p
            ]
    except Exception as e:
        logbook_error = str(e)

    # ── Summary stats ──────────────────────────────────────────────────────────
    active_count   = sum(1 for p in placements if p.get("_is_ongoing"))
    companies_set  = {(p.get("companies") or {}).get("name") for p in placements if (p.get("companies") or {}).get("name")}
    avg_days       = round(sum(p["_days_spent"] for p in placements) / len(placements), 1) if placements else 0
    max_days       = max((p["_days_spent"] for p in placements), default=0)

    return render_template(
        "dept_admin/gis_tracking.html",
        placements=placements,
        live_locations=live_locations,
        logbook_entries=logbook_entries,
        logbook_error=logbook_error,
        total_students=len(student_ids),
        active_count=active_count,
        companies_count=len(companies_set),
        avg_days=avg_days,
        max_days=max_days,
        period_filter=period_filter,
        year_filter=year_filter,
        PERIOD_LABELS=PERIOD_LABELS,
    )


# ── GIS / Attachment Export ────────────────────────────────────────────────────

def _get_period_range(year: int, period: str):
    ranges = {
        "1": (f"{year}-01-01", f"{year}-04-30"),
        "2": (f"{year}-05-01", f"{year}-08-31"),
        "3": (f"{year}-09-01", f"{year}-12-31"),
    }
    return ranges.get(period, (f"{year}-01-01", f"{year}-12-31"))


def _period_label(period: str) -> str:
    return {"1": "January–April", "2": "May–August", "3": "September–December"}.get(period, "Full Year")


@dept_admin_bp.route("/gis-tracking/export")
@dept_admin_required
def gis_tracking_export():
    from datetime import date
    import io, os
    from flask import current_app, Response
    db      = get_service_client()
    dept_id = _dept_id()
    fmt    = request.args.get("format", "excel")
    period = request.args.get("period", "")
    year   = int(request.args.get("year", date.today().year))

    enr = (db.table("enrollments")
             .select("student_id, classes!inner(department_id)")
             .eq("classes.department_id", dept_id)
             .execute().data or [])
    student_ids = list({e["student_id"] for e in enr})
    if not student_ids:
        flash("No students found in your department.", "warning")
        return redirect(url_for("dept_admin.gis_tracking"))

    start_date, end_date = _get_period_range(year, period)
    rows_raw = (db.table("industrial_attachments")
                  .select("id, student_id, start_date, end_date, status, "
                          "student:user_profiles!industrial_attachments_student_id_fkey"
                          "(full_name, admission_no, mobile_number), "
                          "companies(name, city, address, contact_person, contact_phone)")
                  .in_("student_id", student_ids)
                  .gte("start_date", start_date)
                  .lte("start_date", end_date)
                  .order("student_id")
                  .execute().data or [])

    today = date.today()
    rows  = []
    for r in rows_raw:
        st = r.get("student") or {}
        co = r.get("companies") or {}
        try:
            sd = date.fromisoformat(r["start_date"]) if r.get("start_date") else None
            ed = date.fromisoformat(r["end_date"])   if r.get("end_date")   else None
            days_spent = max(0, (min(ed, today) if ed else today) - sd).days if sd else 0
        except Exception:
            days_spent = 0
        rows.append({
            "Admission No": st.get("admission_no", ""),
            "Full Name":    st.get("full_name", ""),
            "Phone":        st.get("mobile_number", ""),
            "Company":      co.get("name", ""),
            "City":         co.get("city", "") or co.get("address", ""),
            "Mentor":       co.get("contact_person", ""),
            "Start Date":   r.get("start_date", ""),
            "End Date":     r.get("end_date", "") or "Ongoing",
            "Days Spent":   str(days_spent),
            "Status":       (r.get("status") or "").title(),
        })

    HEADERS    = ["Admission No", "Full Name", "Phone", "Company", "City",
                  "Mentor", "Start Date", "End Date", "Days Spent", "Status"]
    label      = _period_label(period)
    dept_info  = db.table("departments").select("name").eq("id", dept_id).execute().data or []
    dept_name  = dept_info[0]["name"] if dept_info else "Department"
    report_title = f"Industrial Attachment Report — {dept_name}"
    period_str   = f"{label} {year}"
    logo_path    = os.path.join(current_app.static_folder, "assets", "THIKATTILOGO.jpg")

    # ── PDF ───────────────────────────────────────────────────────────────────
    if fmt == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, Image, HRFlowable)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=14*mm, rightMargin=14*mm,
                                topMargin=12*mm, bottomMargin=16*mm)
        styles = getSampleStyleSheet()
        navy   = colors.HexColor("#0F2C54")
        blue   = colors.HexColor("#1565C0")
        gray   = colors.HexColor("#6B7280")
        lgray  = colors.HexColor("#94A3B8")

        def ps(name, **kw):
            return ParagraphStyle(name, parent=styles["Normal"], **kw)

        inst_s  = ps("inst",  fontName="Helvetica-Bold", fontSize=15, textColor=navy, alignment=1, spaceAfter=2)
        dept_s  = ps("dept",  fontName="Helvetica-Bold", fontSize=11, textColor=blue, alignment=1, spaceAfter=2)
        addr_s  = ps("addr",  fontSize=8,  textColor=gray, alignment=1, spaceAfter=0)
        title_s = ps("title", fontName="Helvetica-Bold", fontSize=12, textColor=navy, alignment=1, spaceBefore=6, spaceAfter=2)
        sub_s   = ps("sub",   fontSize=9,  textColor=gray, alignment=1, spaceAfter=0)
        sign_s  = ps("sign",  fontSize=9,  textColor=colors.black, spaceAfter=1)
        sigh_s  = ps("sigh",  fontName="Helvetica-Bold", fontSize=9, textColor=navy, spaceAfter=2)
        sec_s   = ps("sec",   fontName="Helvetica-Bold", fontSize=10, textColor=navy, spaceBefore=8, spaceAfter=4)

        story = []

        # ── Institute header ──────────────────────────────────────────────────
        logo_cell = Image(logo_path, width=22*mm, height=22*mm) if os.path.exists(logo_path) else Spacer(22*mm, 22*mm)
        hdr_tbl = Table([[
            logo_cell,
            [
                Paragraph("THIKA TECHNICAL TRAINING INSTITUTE", inst_s),
                Paragraph("Industrial Training Department", dept_s),
                Paragraph("P.O. Box 1120-01000, Thika | Tel: +254 67 22284 | www.thika-tti.ac.ke", addr_s),
            ]
        ]], colWidths=[28*mm, None])
        hdr_tbl.setStyle(TableStyle([
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",   (0, 0), (0, 0), 4),
            ("RIGHTPADDING",  (0, 0), (0, 0), 10),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("BOX",           (0, 0), (-1, -1), 1.5, navy),
            ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor("#EFF6FF")),
        ]))
        story.append(hdr_tbl)
        story.append(HRFlowable(width="100%", thickness=2, color=navy, spaceAfter=4, spaceBefore=4))
        story.append(Paragraph(report_title, title_s))
        story.append(Paragraph(f"Period: {period_str}  &nbsp;·&nbsp;  Total Records: {len(rows)}", sub_s))
        story.append(Spacer(1, 5*mm))

        # ── Data table ────────────────────────────────────────────────────────
        col_hdrs = ["Adm No", "Full Name", "Phone", "Company", "City", "Mentor", "Start", "End", "Days", "Status"]
        data     = [col_hdrs] + [[r.get(k, "") for k in HEADERS] for r in rows]
        col_w    = [w*mm for w in [20, 44, 24, 44, 30, 36, 19, 19, 14, 18]]
        data_tbl = Table(data, colWidths=col_w, repeatRows=1)
        data_tbl.setStyle(TableStyle(pdf_header_style_cmds(0) + [
            ("FONTSIZE",      (0, 0), (-1, 0), 8),
            ("FONTSIZE",      (0, 1), (-1, -1), 7.5),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF6FF")]),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
            ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ]))
        story.append(data_tbl)
        story.append(Spacer(1, 8*mm))

        # ── Signing area (shared official sign-off) ───────────────────────────
        story.append(Paragraph("OFFICIAL AUTHORISATION", sec_s))
        story.extend(pdf_signature_block(
            doc.width, officer_title="Industrial Liaison Officer"))

        doc.build(story)
        buf.seek(0)
        fname = f"attachment_report_{label.replace('–', '-')}_{year}.pdf"
        return Response(buf.read(), mimetype="application/pdf",
                        headers={"Content-Disposition": f"attachment; filename={fname}"})

    # ── Excel ─────────────────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Attachments"

    NAVY   = "0F2C54"
    BLUE   = "1565C0"
    LGRAY  = "94A3B8"
    STRIPE = "EFF6FF"
    thin   = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    NUM_COLS = len(HEADERS)

    def last_col():
        return get_column_letter(NUM_COLS)

    # ── Logo (rows 1-3, col A) ────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    if os.path.exists(logo_path):
        try:
            from openpyxl.drawing.image import Image as XLImg
            xl_logo = XLImg(logo_path)
            xl_logo.width  = 66
            xl_logo.height = 66
            xl_logo.anchor = "A1"
            ws.add_image(xl_logo)
        except Exception:
            pass

    # Institute header text (cols B onwards, rows 1-3)
    def hcell(ref, val, bold=False, size=11, color=NAVY, center=True):
        c = ws[ref]
        c.value     = val
        c.font      = Font(bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal="center" if center else "left",
                                vertical="center", wrap_text=False)

    ws.merge_cells(f"B1:{last_col()}1")
    hcell("B1", "THIKA TECHNICAL TRAINING INSTITUTE", bold=True, size=14, color=NAVY)
    ws.merge_cells(f"B2:{last_col()}2")
    hcell("B2", "Industrial Training Department", bold=True, size=11, color=BLUE)
    ws.merge_cells(f"B3:{last_col()}3")
    hcell("B3", "P.O. Box 1120-01000, Thika  |  Tel: +254 67 22284  |  www.thika-tti.ac.ke",
          size=9, color="6B7280")

    # ── Report title ──────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 22
    ws.merge_cells(f"A4:{last_col()}4")
    hcell("A4", report_title, bold=True, size=13, color=NAVY)

    ws.row_dimensions[5].height = 16
    ws.merge_cells(f"A5:{last_col()}5")
    hcell("A5", f"Period: {period_str}     |     Total Records: {len(rows)}", size=10, color="6B7280")
    ws["A5"].font = Font(italic=True, size=10, color="6B7280")

    # Separator row
    ws.row_dimensions[6].height = 6

    # ── Column headers (light institutional style) ────────────────────────────
    ws.append([None])   # row 6 spacer
    ws.append(HEADERS)  # row 7
    hdr_row  = ws.max_row
    ws.row_dimensions[hdr_row].height = 18
    style_header_row(ws, hdr_row, NUM_COLS)

    # ── Data rows ─────────────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor=STRIPE)
    for i, row in enumerate(rows):
        ws.append([row.get(h, "") for h in HEADERS])
        dr = ws.max_row
        ws.row_dimensions[dr].height = 15
        for ci in range(1, NUM_COLS + 1):
            c = ws.cell(row=dr, column=ci)
            c.border    = border
            c.alignment = Alignment(vertical="center", wrap_text=False)
            if i % 2 == 1:
                c.fill = alt_fill

    # ── Column widths ─────────────────────────────────────────────────────────
    for i, w in enumerate([14, 26, 16, 28, 18, 24, 13, 13, 11, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Signing area (shared official sign-off) ───────────────────────────────
    excel_signature_block(ws, ws.max_row + 2,
                          officer_title="Industrial Liaison Officer")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"attachment_report_{label.replace('–', '-')}_{year}.xlsx"
    return Response(buf.read(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


# ── Department Notices / Memos ─────────────────────────────────────────────────

@dept_admin_bp.route("/notices", methods=["GET"])
@dept_admin_required
def notices():
    """List all notices sent by this dept admin."""
    db = get_service_client()
    dept_id = _dept_id()
    dept = db.table("departments").select("name").eq("id", dept_id).single().execute().data or {}
    notices_list = []
    classes_list = []
    try:
        notices_list = (db.table("dept_notices")
                         .select("*")
                         .eq("department_id", dept_id)
                         .order("sent_at", desc=True)
                         .limit(50)
                         .execute().data or [])
        # Enrich with class name
        classes_list = (db.table("classes")
                         .select("id, name")
                         .eq("department_id", dept_id)
                         .order("name")
                         .execute().data or [])
        classes_map = {c["id"]: c["name"] for c in classes_list}
        for n in notices_list:
            cid = n.get("class_id")
            n["class_name"] = classes_map.get(cid, "All Trainees") if cid else "All Trainees"
    except Exception as e:
        flash(f"Error loading notices: {e}", "danger")
    return render_template("dept_admin/notices.html",
                           notices=notices_list,
                           classes=classes_list,
                           dept=dept)


@dept_admin_bp.route("/notices/send", methods=["POST"])
@dept_admin_required
def send_notice():
    """Send an official notice/memo to trainees."""
    dept_id = _dept_id()
    user = current_user()

    title   = (request.form.get("title") or "").strip()
    message = (request.form.get("message") or "").strip()
    ntype   = request.form.get("notice_type", "info")
    class_id = request.form.get("class_id") or None

    if not title or not message:
        flash("Title and message are required.", "warning")
        return redirect(url_for("dept_admin.notices"))
    if ntype not in ("info", "warning", "success", "error"):
        ntype = "info"

    try:
        db = get_service_client()
        # Persist the notice record
        inserted = db.table("dept_notices").insert({
            "department_id": dept_id,
            "sent_by": user["id"],
            "title": title,
            "message": message,
            "notice_type": ntype,
            "class_id": class_id,
            "sent_at": datetime.now().isoformat()
        }).execute()
        notice_id = (inserted.data or [{}])[0].get("id")

        # Push notifications to trainees
        notif_title = f"[Notice] {title}"
        count = notify_dept_notice(
            department_id=dept_id,
            title=notif_title,
            message=message,
            notice_type=ntype,
            action_url="/notifications",
            class_id=class_id,
            sender_id=user["id"],
            notice_id=notice_id,
        )
        write_audit_log("send_dept_notice",
                        target=f"dept:{dept_id}",
                        detail={"title": title, "recipients": count, "notice_id": notice_id})
        flash(f"Notice sent successfully to {count} trainee(s).", "success")
    except Exception as e:
        flash(f"Failed to send notice: {e}", "danger")

    return redirect(url_for("dept_admin.notices"))


@dept_admin_bp.route("/notices/<notice_id>/delete", methods=["POST"])
@dept_admin_required
def delete_notice(notice_id):
    """Sender deletes a sent notice and recalls recipient notifications."""
    db = get_service_client()
    user = current_user()
    dept_id = _dept_id()

    try:
        notice = (db.table("dept_notices")
                    .select("*")
                    .eq("id", notice_id)
                    .single()
                    .execute().data)
        if not notice:
            flash("Notice not found.", "error")
            return redirect(url_for("dept_admin.notices"))
        if notice.get("department_id") != dept_id and notice.get("sent_by") != user["id"]:
            flash("You can only delete notices you sent for your department.", "error")
            return redirect(url_for("dept_admin.notices"))
        if notice.get("sent_by") and notice.get("sent_by") != user["id"]:
            flash("Only the original sender can delete this notice.", "error")
            return redirect(url_for("dept_admin.notices"))

        title = notice.get("title") or ""
        message = notice.get("message") or ""
        deleted = delete_notifications_for_notice(
            notice_id=notice_id,
            title=f"[Notice] {title}",
            message=message,
            sender_id=user["id"],
        )
        db.table("dept_notices").delete().eq("id", notice_id).execute()
        write_audit_log("delete_dept_notice",
                        target=f"notice:{notice_id}",
                        detail={"title": title, "notifications_removed": deleted})
        flash(f"Notice deleted. Recalled {deleted} notification(s) from recipients.", "success")
    except Exception as e:
        flash(f"Failed to delete notice: {e}", "danger")

    return redirect(url_for("dept_admin.notices"))


# ── Fingerprint Registration ───────────────────────────────────────────────────

@dept_admin_bp.route("/fingerprint-registration")
@dept_admin_required
def fingerprint_registration():
    from datetime import datetime as _dt
    db      = get_service_client()
    dept_id = _dept_id()

    class_filter = request.args.get("class_id", "").strip()
    search_q     = request.args.get("q", "").strip().lower()

    classes   = (db.table("classes")
                   .select("id, name")
                   .eq("department_id", dept_id)
                   .order("name")
                   .execute().data or [])
    class_ids = [c["id"] for c in classes]

    students = []
    if class_ids:
        target_ids = [class_filter] if (class_filter and class_filter in class_ids) else class_ids
        rows = (db.table("enrollments")
                  .select("student_id, class_id, classes(name), "
                          "user_profiles!enrollments_student_id_fkey"
                          "(id, full_name, admission_no, mobile_number, biometric_id)")
                  .in_("class_id", target_ids)
                  .execute().data or [])
        seen = set()
        for r in rows:
            p   = r.get("user_profiles") or {}
            sid = p.get("id") or r.get("student_id")
            if not sid or sid in seen:
                continue
            seen.add(sid)
            name = p.get("full_name") or ""
            adm  = p.get("admission_no") or ""
            if search_q and search_q not in name.lower() and search_q not in adm.lower():
                continue
            cls = r.get("classes") or {}
            bio = (p.get("biometric_id") or "").strip()
            students.append({
                "id":           sid,
                "name":         name,
                "admission_no": adm,
                "mobile":       p.get("mobile_number") or "",
                "class_name":   cls.get("name") or "",
                "class_id":     r.get("class_id") or "",
                "biometric_id": bio,
                "registered":   bool(bio),
            })
        students.sort(key=lambda s: (s["class_name"], s["name"]))

    from routes.biometric_attendance import active_enrollment, enrollment_lock
    with enrollment_lock:
        active_enroll = dict(active_enrollment)

    stats = {
        "total":        len(students),
        "registered":   sum(1 for s in students if s["registered"]),
        "unregistered": sum(1 for s in students if not s["registered"]),
    }

    return render_template(
        "dept_admin/fingerprint_registration.html",
        students=students,
        classes=classes,
        class_filter=class_filter,
        search_q=search_q,
        stats=stats,
        active_enroll=active_enroll,
    )


@dept_admin_bp.route("/fingerprint-registration/assign", methods=["POST"])
@dept_admin_required
def fingerprint_assign():
    db           = get_service_client()
    dept_id      = _dept_id()
    student_id   = (request.form.get("student_id")   or "").strip()
    biometric_id = (request.form.get("biometric_id") or "").strip()
    redirect_cls = (request.form.get("class_filter") or "").strip()
    redirect_q   = (request.form.get("q")           or "").strip()

    if not student_id:
        flash("Student ID missing.", "error")
        return redirect(url_for("dept_admin.fingerprint_registration"))

    check = (db.table("enrollments")
               .select("student_id, classes!inner(department_id)")
               .eq("student_id", student_id)
               .eq("classes.department_id", dept_id)
               .limit(1)
               .execute().data or [])
    if not check:
        flash("Student not found in your department.", "error")
        return redirect(url_for("dept_admin.fingerprint_registration"))

    try:
        db.table("user_profiles").update(
            {"biometric_id": biometric_id if biometric_id else None}
        ).eq("id", student_id).execute()
        write_audit_log("fingerprint_assign",
                        target=f"student:{student_id},bio_id:{biometric_id or 'cleared'}")
        flash(f'Fingerprint ID "{biometric_id}" saved.' if biometric_id else "Fingerprint ID cleared.", "success")
    except Exception as e:
        flash(f"Error saving fingerprint ID: {e}", "error")

    return redirect(url_for("dept_admin.fingerprint_registration",
                            class_id=redirect_cls, q=redirect_q))


@dept_admin_bp.route("/fingerprint-registration/remove", methods=["POST"])
@dept_admin_required
def fingerprint_remove():
    db         = get_service_client()
    dept_id    = _dept_id()
    student_id = (request.form.get("student_id") or "").strip()
    if not student_id:
        flash("Student ID missing.", "error")
        return redirect(url_for("dept_admin.fingerprint_registration"))
    check = (db.table("enrollments")
               .select("student_id, classes!inner(department_id)")
               .eq("student_id", student_id)
               .eq("classes.department_id", dept_id)
               .limit(1)
               .execute().data or [])
    if not check:
        flash("Student not found in your department.", "error")
        return redirect(url_for("dept_admin.fingerprint_registration"))
    try:
        db.table("user_profiles").update({"biometric_id": None}).eq("id", student_id).execute()
        write_audit_log("fingerprint_remove", target=f"student:{student_id}")
        flash("Fingerprint registration removed.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("dept_admin.fingerprint_registration"))


@dept_admin_bp.route("/fingerprint-registration/start-enroll", methods=["POST"])
@dept_admin_required
def fingerprint_start_enroll():
    from datetime import datetime as _dt
    from routes.biometric_attendance import active_enrollment, enrollment_lock
    data         = request.get_json(silent=True) or {}
    student_id   = (data.get("student_id")   or "").strip()
    student_name = (data.get("student_name") or "").strip()
    if not student_id:
        return jsonify({"success": False, "error": "student_id required"}), 400
    dept_id = _dept_id()
    with enrollment_lock:
        active_enrollment.clear()
        active_enrollment.update({
            "student_id":   student_id,
            "student_name": student_name,
            "dept_id":      dept_id,
            "started_at":   _dt.now().isoformat(),
            "biometric_id": None,
            "status":       "waiting",
        })
    write_audit_log("fingerprint_enroll_start", target=f"student:{student_id}")
    return jsonify({"success": True,
                    "message": f"Enrollment started for {student_name}."})


@dept_admin_bp.route("/fingerprint-registration/enroll-status")
@dept_admin_required
def fingerprint_enroll_status():
    from routes.biometric_attendance import active_enrollment, enrollment_lock
    with enrollment_lock:
        session = dict(active_enrollment)
    return jsonify(session)


@dept_admin_bp.route("/fingerprint-registration/cancel-enroll", methods=["POST"])
@dept_admin_required
def fingerprint_cancel_enroll():
    from routes.biometric_attendance import active_enrollment, enrollment_lock
    with enrollment_lock:
        active_enrollment.clear()
    write_audit_log("fingerprint_enroll_cancel")
    return jsonify({"success": True})


# ── Industrial Attachment Marks ───────────────────────────────────────────────

from routes.attachment_helpers import (
    get_grading_config, compute_weighted_grade, score_to_cdacc, section_max
)

@dept_admin_bp.route("/attachment-marks")
@dept_admin_required
def attachment_marks():
    db      = get_service_client()
    dept_id = _dept_id()

    student_ids = _dept_student_ids(db, dept_id)
    attachments = []
    grades_map  = {}
    config      = get_grading_config(db, dept_id)

    if student_ids:
        raw = (db.table("industrial_attachments")
                 .select("id, student_id, start_date, end_date, status, "
                         "companies(name), "
                         "student:user_profiles!industrial_attachments_student_id_fkey"
                         "(full_name, admission_no)")
                 .in_("student_id", student_ids)
                 .order("created_at", desc=True)
                 .execute().data or [])

        att_ids = [a["id"] for a in raw]
        if att_ids:
            grade_rows = (db.table("attachment_grades")
                            .select("*")
                            .in_("attachment_id", att_ids)
                            .execute().data or [])
            grades_map = {g["attachment_id"]: g for g in grade_rows}

        for a in raw:
            a["_grade"] = grades_map.get(a["id"])
        attachments = raw

    # stats
    total   = len(attachments)
    graded  = sum(1 for a in attachments if a.get("_grade"))
    pending = total - graded
    scores  = [float(a["_grade"]["weighted_total"]) for a in attachments
               if a.get("_grade") and a["_grade"].get("weighted_total") is not None]
    avg     = round(sum(scores) / len(scores), 1) if scores else 0

    return render_template(
        "dept_admin/attachment_marks.html",
        attachments=attachments,
        config=config,
        stats={"total": total, "graded": graded, "pending": pending, "avg": avg},
    )


@dept_admin_bp.route("/attachment-marks/<att_id>/save", methods=["POST"])
@dept_admin_required
def save_attachment_marks(att_id):
    db      = get_service_client()
    dept_id = _dept_id()
    user    = current_user()

    # Verify this attachment belongs to a dept student
    student_ids = _dept_student_ids(db, dept_id)
    att_row = (db.table("industrial_attachments")
                 .select("id, student_id")
                 .eq("id", att_id)
                 .limit(1)
                 .execute().data or [])
    if not att_row or att_row[0].get("student_id") not in student_ids:
        return jsonify({"ok": False, "error": "Not authorised"}), 403

    def _f(name):
        try:    return float(request.form.get(name) or 0)
        except: return 0.0

    scores = {
        "score_gps_attendance":   _f("score_gps_attendance"),
        "score_logbook":          _f("score_logbook"),
        "score_mentor_eval":      _f("score_mentor_eval"),
        "score_trainer_assessment":_f("score_trainer_assessment"),
        "score_final_report":     _f("score_final_report"),
    }
    config  = get_grading_config(db, dept_id)
    weights = {k: v for k, v in config.items() if k.startswith("weight_")}
    scores  = {k: min(max(v, 0.0), section_max(weights, k)) for k, v in scores.items()}
    total   = compute_weighted_grade(scores, weights)
    grade   = score_to_cdacc(total)

    payload = {**scores, "weighted_total": total, "final_grade": grade,
               "graded_by": user["id"], "graded_at": datetime.utcnow().isoformat()}

    existing = (db.table("attachment_grades")
                  .select("id")
                  .eq("attachment_id", att_id)
                  .limit(1)
                  .execute().data or [])
    if existing:
        db.table("attachment_grades").update(payload).eq("attachment_id", att_id).execute()
    else:
        payload["attachment_id"] = att_id
        db.table("attachment_grades").insert(payload).execute()

    write_audit_log("save_attachment_marks", target=f"attachment:{att_id}",
                    detail={"grade": grade, "total": total})
    return jsonify({"ok": True, "total": total, "grade": grade})


# ── Mentoring Tool / Logbook — Dept Admin view ───────────────────────────────

@dept_admin_bp.route("/mentoring-tools")
@dept_admin_required
def view_mentoring_tools():
    db      = get_service_client()
    user    = current_user()
    dept_id = user.get("department_id")
    stu_ids = _dept_student_ids(db, dept_id)
    if not stu_ids:
        return render_template("dept_admin/mentoring_tools.html", uploads=[])
    rows = (db.table("mentoring_tool_uploads")
              .select("*, student:user_profiles!student_id(full_name,admission_no,department_id), attachment:industrial_attachments!attachment_id(id,company_id,status)")
              .in_("student_id", stu_ids)
              .order("uploaded_at", desc=True)
              .execute().data or [])
    co_ids = list({r["attachment"]["company_id"] for r in rows if r.get("attachment") and r["attachment"].get("company_id")})
    co_map = {}
    if co_ids:
        for co in (db.table("companies").select("id,name").in_("id", co_ids).execute().data or []):
            co_map[co["id"]] = co["name"]
    for r in rows:
        r["_company"] = co_map.get((r.get("attachment") or {}).get("company_id", ""), "—")
    return render_template("dept_admin/mentoring_tools.html", uploads=rows)
