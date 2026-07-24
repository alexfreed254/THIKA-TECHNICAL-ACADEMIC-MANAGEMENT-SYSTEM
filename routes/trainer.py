"""
routes/trainer_merged.py — Trainer blueprint (merged system).

Combines features from both:
- Attendance capture (from original)
- Assessment review (from copy)

Trainers can only access classes and units assigned to them.
"""

from flask import (Blueprint, render_template, request,
                   redirect, url_for, flash, abort, jsonify, make_response)
from auth_utils import (trainer_required, write_audit_log, current_user)
from db import get_service_client
from notifications import get_user_notifications
from datetime import datetime, date, timedelta
from report_utils import (excel_letterhead, style_header_row,
                          excel_signature_block, pdf_letterhead,
                          pdf_header_style_cmds, pdf_signature_block)
import re
import os
import json

trainer_bp = Blueprint("trainer", __name__)


def _trainer_assigned_unit_ids(db) -> list:
    """Return unit_ids this trainer is assigned to (trainer_units ∪ class_units)."""
    user = current_user()
    if user.get("role") != "trainer":
        return []  # non-trainers: caller decides (do not treat as "see everything")
    ids = set()
    rows = (db.table("trainer_units").select("unit_id")
            .eq("trainer_id", user["id"]).execute().data or [])
    ids.update(r["unit_id"] for r in rows if r.get("unit_id"))
    cu = (db.table("class_units").select("unit_id")
          .eq("trainer_id", user["id"]).execute().data or [])
    ids.update(r["unit_id"] for r in cu if r.get("unit_id"))
    return list(ids)


def _trainer_owns_class_unit(db, class_id: str, unit_id: str) -> bool:
    """True if current trainer is assigned to this class+unit."""
    user = current_user()
    if not user or user.get("role") != "trainer":
        return user.get("role") in ("super_admin", "dept_admin") if user else False
    row = (db.table("class_units").select("id")
           .eq("class_id", class_id)
           .eq("unit_id", unit_id)
           .eq("trainer_id", user["id"])
           .limit(1)
           .execute().data or [])
    return bool(row)


def _check_unit_access(db, unit_id: str) -> bool:
    """Return True if current trainer is allowed to act on this unit."""
    user = current_user()
    if user.get("role") != "trainer":
        return True
    assigned = _trainer_assigned_unit_ids(db)
    return bool(assigned) and unit_id in assigned


def _file_slug(text: str) -> str:
    """Sanitize text into a filename-safe slug (alphanumeric + underscore)."""
    text = str(text or "").strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s]+', '_', text)
    return text.strip('_-') or 'unknown'


def _rename_script_file(db, assessment_id: str, action: str, trainer_name: str):
    """
    Rename the assessment script file in Supabase storage and update DB.

    Approval format:
        AdmNo_UnitName_Type_No_Cycle_Term_STATUS-TrainerName.pdf
    e.g.: ADM001_Introduction_to_ICT_FA_1_1_2_APPROVED-John_Doe.pdf
    """
    try:
        a = (db.table("assessments")
             .select("student_id, unit_id, assessment_type, assessment_no, "
                     "cycle, term, script_file_path, "
                     "user_profiles!assessments_student_id_fkey(admission_no), "
                     "units(name)")
             .eq("id", assessment_id).single().execute().data)
        if not a:
            return

        student_info   = a.get("user_profiles") or {}
        unit_info      = a.get("units") or {}
        admission_slug = _file_slug(student_info.get("admission_no") or a["student_id"])
        unit_slug      = _file_slug(unit_info.get("name") or a.get("unit_id", "unit"))
        atype_slug     = _file_slug(a.get("assessment_type") or "FA")
        ano            = str(a.get("assessment_no") or "1")
        cycle_str      = str(a.get("cycle") or "1")
        term_str       = str(a.get("term") or "1")
        status_str     = "APPROVED" if action == "approved" else "REJECTED"
        trainer_slug   = _file_slug(trainer_name)

        new_display = (
            f"{admission_slug}_{unit_slug}_{atype_slug}_{ano}_"
            f"{cycle_str}_{term_str}_{status_str}-{trainer_slug}.pdf"
        )

        old_path = a.get("script_file_path")
        if not old_path:
            db.table("assessments").update({"script_file_name": new_display}).eq("id", assessment_id).execute()
            return

        # Derive new storage path — keep same folder, swap filename
        folder   = old_path.rsplit("/", 1)[0] if "/" in old_path else "scripts"
        new_path = f"{folder}/{new_display}"

        # Copy to new path then remove old file
        storage    = get_service_client().storage
        file_bytes = bytes(storage.from_("assessment-scripts").download(old_path))
        storage.from_("assessment-scripts").upload(
            new_path, file_bytes, {"content-type": "application/pdf"}
        )
        try:
            storage.from_("assessment-scripts").remove([old_path])
        except Exception:
            pass

        db.table("assessments").update({
            "script_file_name": new_display,
            "script_file_path": new_path,
        }).eq("id", assessment_id).execute()

        # ── Rename all evidence files for this assessment ──────────────────
        ext_ct = {
            'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png',
            'gif': 'image/gif',  'webp': 'image/webp',
            'mp4': 'video/mp4',  'mov': 'video/quicktime', 'avi': 'video/avi',
            'mkv': 'video/x-matroska', 'webm': 'video/webm',
            'mp3': 'audio/mpeg', 'wav': 'audio/wav',  'ogg': 'audio/ogg',
            'm4a': 'audio/mp4',  'flac': 'audio/flac', 'aac': 'audio/aac',
        }
        evidence_list = (db.table("evidence")
                         .select("id, file_path, file_name")
                         .eq("assessment_id", assessment_id)
                         .order("uploaded_at").execute().data or [])
        ev_storage = get_service_client().storage
        for ev_idx, ev in enumerate(evidence_list, 1):
            old_ev = ev.get("file_path")
            if not old_ev:
                continue
            ev_ext     = old_ev.rsplit(".", 1)[-1].lower() if "." in old_ev else "jpg"
            page_sfx   = f"-ev{ev_idx}" if len(evidence_list) > 1 else ""
            new_ev_name = (
                f"{admission_slug}_{unit_slug}_{atype_slug}_{ano}_"
                f"{cycle_str}_{term_str}_{status_str}-{trainer_slug}{page_sfx}.{ev_ext}"
            )
            ev_folder  = old_ev.rsplit("/", 1)[0] if "/" in old_ev else "evidence"
            new_ev_path = f"{ev_folder}/{new_ev_name}"
            try:
                ev_bytes = bytes(ev_storage.from_("assessment-evidence").download(old_ev))
                ev_storage.from_("assessment-evidence").upload(
                    new_ev_path, ev_bytes,
                    {"content-type": ext_ct.get(ev_ext, "application/octet-stream")}
                )
                try:
                    ev_storage.from_("assessment-evidence").remove([old_ev])
                except Exception:
                    pass
                db.table("evidence").update({
                    "file_name": new_ev_name,
                    "file_path": new_ev_path,
                }).eq("id", ev["id"]).execute()
            except Exception:
                pass

    except Exception:
        pass


# ── Dashboard ─────────────────────────────────────────────────────────────────

@trainer_bp.route("/")
@trainer_bp.route("/dashboard")
@trainer_required
def dashboard():
    db = get_service_client()
    user = current_user()
    stats = {}
    
    # Analytics defaults
    att_unit_labels = att_unit_present = att_unit_absent = []
    assess_unit_labels = assess_unit_pending = assess_unit_approved = assess_unit_rejected = []
    trend_labels = trend_present = trend_absent = []
    pending_assessments = []
    units_list = []

    try:
        assigned_unit_ids = _trainer_assigned_unit_ids(db)

        # ── Assessment stats (exact counts) ────────────────────────────────────
        from stats_utils import count_in, count_table, exact_count

        if assigned_unit_ids:
            stats["total"]    = count_in(db, "assessments", "unit_id", assigned_unit_ids)
            stats["pending"]  = count_in(db, "assessments", "unit_id", assigned_unit_ids, status="pending")
            stats["approved"] = count_in(db, "assessments", "unit_id", assigned_unit_ids, status="approved")
            stats["rejected"] = count_in(db, "assessments", "unit_id", assigned_unit_ids, status="rejected")
        else:
            stats["total"] = stats["pending"] = stats["approved"] = stats["rejected"] = 0

        # Trips uploaded by this trainer + pending clearance approvals
        try:
            stats["trips_uploaded"] = count_table(db, "academic_trips", uploaded_by=user["id"])
        except Exception:
            stats["trips_uploaded"] = 0
        try:
            stats["clearance_pending"] = exact_count(
                db.table("clearance_approvals").select("id", count="exact")
                .eq("approver_id", user["id"]).eq("status", "pending")
            )
        except Exception:
            stats["clearance_pending"] = 0
        try:
            stats["summative_nyc"] = count_in(
                db, "summative_competences", "unit_id", assigned_unit_ids or [],
                competence="not_yet_competent",
            ) if assigned_unit_ids else 0
        except Exception:
            stats["summative_nyc"] = 0

        # ── Pending assessments (for table) ───────────────────────────────────
        if assigned_unit_ids:
            pending_assessments = (db.table("assessments")
                .select("*, user_profiles!assessments_student_id_fkey(full_name, admission_no, mobile_number), units(name), classes(name)")
                .eq("status", "pending")
                .in_("unit_id", assigned_unit_ids)
                .order("uploaded_at", desc=True)
                .limit(15).execute().data or [])

        # ── Assigned units list ────────────────────────────────────────────────
        if assigned_unit_ids:
            units_list = (db.table("units").select("*")
                          .in_("id", assigned_unit_ids).order("name").execute().data or [])

        # ── Analytics: Attendance per unit ─────────────────────────────────────
        if assigned_unit_ids:
            att_rows = (db.table("attendance")
                        .select("status, units!inner(name)")
                        .in_("unit_id", assigned_unit_ids)
                        .execute().data or [])
            att_map = {}
            for row in att_rows:
                uname = (row.get("units") or {}).get("name", "Unknown")
                b = att_map.setdefault(uname, {"present": 0, "absent": 0})
                if row.get("status") == "present":
                    b["present"] += 1
                else:
                    b["absent"] += 1
            att_unit_labels   = list(att_map.keys())
            att_unit_present  = [att_map[u]["present"] for u in att_unit_labels]
            att_unit_absent   = [att_map[u]["absent"]  for u in att_unit_labels]

        # ── Analytics: 7-day attendance trend (exact) ─────────────────────────
        for i in range(6, -1, -1):
            day = (date.today() - timedelta(days=i)).isoformat()
            trend_labels.append(day[5:])
            if assigned_unit_ids:
                present_n = total_n = 0
                for ci in range(0, len(assigned_unit_ids), 100):
                    chunk = assigned_unit_ids[ci:ci + 100]
                    present_n += exact_count(
                        db.table("attendance").select("id", count="exact")
                        .in_("unit_id", chunk).eq("attendance_date", day).eq("status", "present")
                    )
                    total_n += exact_count(
                        db.table("attendance").select("id", count="exact")
                        .in_("unit_id", chunk).eq("attendance_date", day)
                    )
                trend_present.append(present_n)
                trend_absent.append(max(0, total_n - present_n))
            else:
                trend_present.append(0)
                trend_absent.append(0)

        # ── Analytics: Assessments per unit breakdown ──────────────────────────
        if assigned_unit_ids:
            a_rows = (db.table("assessments")
                      .select("status, units!inner(name)")
                      .in_("unit_id", assigned_unit_ids)
                      .execute().data or [])
            a_map = {}
            for row in a_rows:
                uname = (row.get("units") or {}).get("name", "Unknown")
                b = a_map.setdefault(uname, {"pending": 0, "approved": 0, "rejected": 0})
                s = row.get("status", "pending")
                b[s] = b.get(s, 0) + 1
            assess_unit_labels   = list(a_map.keys())
            assess_unit_pending  = [a_map[u]["pending"]  for u in assess_unit_labels]
            assess_unit_approved = [a_map[u]["approved"] for u in assess_unit_labels]
            assess_unit_rejected = [a_map[u]["rejected"] for u in assess_unit_labels]

    except Exception as e:
        flash(f"Error loading dashboard: {e}", "danger")

    unread_notifications = get_user_notifications(user["id"], unread_only=True, limit=5)

    return render_template("trainer/dashboard_enhanced.html",
                          stats=stats,
                          pending_assessments=pending_assessments,
                          units_list=units_list,
                          unread_notifications=unread_notifications,
                          current_month=datetime.now().strftime("%B %Y"),
                          # analytics
                          att_unit_labels=json.dumps(att_unit_labels),
                          att_unit_present=json.dumps(att_unit_present),
                          att_unit_absent=json.dumps(att_unit_absent),
                          trend_labels=json.dumps(trend_labels),
                          trend_present=json.dumps(trend_present),
                          trend_absent=json.dumps(trend_absent),
                          assess_unit_labels=json.dumps(assess_unit_labels),
                          assess_unit_pending=json.dumps(assess_unit_pending),
                          assess_unit_approved=json.dumps(assess_unit_approved),
                          assess_unit_rejected=json.dumps(assess_unit_rejected))


# ── Attendance Capture ─────────────────────────────────────────────────────────

@trainer_bp.route("/attendance", methods=["GET", "POST"])
@trainer_required
def attendance():
    db = get_service_client()
    user = current_user()
    dept_id = user.get("department_id")
    dept_name = "Department"
    if dept_id:
        dept = db.table("departments").select("name").eq("id", dept_id).single().execute().data
        if dept:
            dept_name = dept["name"]

    # Classes assigned to this trainer
    cu_rows = (db.table("class_units")
                 .select("class_id")
                 .eq("trainer_id", user["id"])
                 .execute().data or [])
    class_ids = list({r["class_id"] for r in cu_rows})

    class_list = []
    if class_ids:
        class_list = (db.table("classes")
                        .select("*")
                        .in_("id", class_ids)
                        .eq("department_id", dept_id)
                        .order("name")
                        .execute().data or [])

    class_id = request.args.get("class_id", "")
    unit_id  = request.args.get("unit_id", "")
    week     = request.args.get("week", 0, type=int)   # 0 = not yet selected
    lesson   = request.args.get("lesson", "")           # "" = not yet selected
    year     = request.args.get("year", datetime.now().year, type=int)
    term     = request.args.get("term", 1, type=int)

    units_list = []
    students_list = []
    attendance_submitted = False
    active_event = None

    if class_id:
        units_list = (db.table("class_units")
                        .select("*, units(id, code, name)")
                        .eq("class_id", class_id)
                        .eq("trainer_id", user["id"])
                        .execute().data or [])

        students_list = (db.table("enrollments")
                           .select("*, user_profiles(full_name, admission_no)")
                           .eq("class_id", class_id)
                           .execute().data or [])

        if unit_id and week and lesson:
            existing = (db.table("attendance")
                          .select("id", count="exact")
                          .eq("unit_id", unit_id)
                          .eq("trainer_id", user["id"])
                          .eq("week", week)
                          .eq("lesson", lesson)
                          .eq("year", year)
                          .eq("term", term)
                          .execute())
            attendance_submitted = (existing.count or 0) > 0

            event_row = (db.table("class_events")
                           .select("*")
                           .eq("class_id", class_id)
                           .eq("trainer_id", user["id"])
                           .eq("week", week)
                           .eq("lesson", lesson)
                           .eq("year", year)
                           .eq("term", term)
                           .execute().data or [])
            if event_row:
                active_event = event_row[0]

    if request.method == "POST":
        action = request.form.get("action")
        
        if action == "submit_attendance":
            if not class_id or not unit_id:
                flash("Class and unit are required.", "error")
            elif not _trainer_owns_class_unit(db, class_id, unit_id):
                abort(403)
            else:
                try:
                    for student in students_list:
                        status = request.form.get(f"status_{student['student_id']}", "absent")
                        db.table("attendance").insert({
                            "student_id": student["student_id"],
                            "unit_id": unit_id,
                            "unit_code": request.form.get("unit_code", ""),
                            "trainer_id": user["id"],
                            "lesson": lesson,
                            "week": week,
                            "year": year,
                            "term": term,
                            "status": status
                        }).execute()
                    
                    write_audit_log("submit_attendance", target=f"class:{class_id},unit:{unit_id}")
                    flash("Attendance submitted successfully.", "success")
                    return redirect(url_for("trainer.attendance", class_id=class_id, unit_id=unit_id, week=week, lesson=lesson, year=year, term=term))
                except Exception:
                    flash("Error submitting attendance. Please try again.", "error")
        
        elif action in ("mark_holiday", "mark_academic_trip"):
            if not class_id or not unit_id:
                flash("Class and unit are required.", "error")
            elif not _trainer_owns_class_unit(db, class_id, unit_id):
                abort(403)
            else:
                try:
                    # Create class_event
                    event_type = "holiday" if action == "mark_holiday" else "academic_trip"
                    label = "Holiday" if action == "mark_holiday" else "Academic Trip"
                    db.table("class_events").insert({
                        "class_id": class_id,
                        "unit_id": unit_id,
                        "trainer_id": user["id"],
                        "event_type": event_type,
                        "week": week,
                        "lesson": lesson,
                        "year": year,
                        "term": term,
                        "note": request.form.get("note", f"Marked as {label}")
                    }).execute()
                    # Mark all students with absent status (event record explains why)
                    for student in students_list:
                        existing = db.table("attendance").select("id").eq("student_id", student["student_id"]).eq("unit_id", unit_id).eq("week", week).eq("lesson", lesson).eq("year", year).eq("term", term).execute().data
                        if not existing:
                            db.table("attendance").insert({
                                "student_id": student["student_id"],
                                "unit_id": unit_id,
                                "unit_code": request.form.get("unit_code", ""),
                                "trainer_id": user["id"],
                                "lesson": lesson,
                                "week": week,
                                "year": year,
                                "term": term,
                                "status": "absent"
                            }).execute()
                    write_audit_log(action, target=f"class:{class_id},unit:{unit_id}")
                    flash(f"Marked as {label} successfully.", "success")
                    return redirect(url_for("trainer.attendance", class_id=class_id, unit_id=unit_id, week=week, lesson=lesson, year=year, term=term))
                except Exception as e:
                    flash(f"Error marking {label}: {e}", "error")

        elif action == "add_event":
            event_type = request.form.get("event_type", "holiday")
            note       = request.form.get("note", "")
            label      = "Holiday" if event_type == "holiday" else "Academic Trip"
            if not class_id or not unit_id or not week or not lesson:
                flash("Class, Unit, Week and Lesson are required.", "error")
            else:
                # Guard against duplicate event for this session
                dup = (db.table("class_events").select("id")
                         .eq("class_id", class_id).eq("unit_id", unit_id)
                         .eq("trainer_id", user["id"]).eq("week", week)
                         .eq("lesson", lesson).eq("year", year).eq("term", term)
                         .execute().data or [])
                if dup:
                    flash(f"An event is already recorded for this session.", "warning")
                else:
                    try:
                        db.table("class_events").insert({
                            "class_id":   class_id,
                            "unit_id":    unit_id,
                            "trainer_id": user["id"],
                            "event_type": event_type,
                            "week":       week,
                            "lesson":     lesson,
                            "year":       year,
                            "term":       term,
                            "note":       note or f"Marked as {label}"
                        }).execute()
                        # Mark every enrolled student absent for this session
                        for student in students_list:
                            already = (db.table("attendance").select("id")
                                         .eq("student_id", student["student_id"])
                                         .eq("unit_id", unit_id)
                                         .eq("week", week).eq("lesson", lesson)
                                         .eq("year", year).eq("term", term)
                                         .execute().data or [])
                            if not already:
                                db.table("attendance").insert({
                                    "student_id": student["student_id"],
                                    "unit_id":    unit_id,
                                    "unit_code":  "",
                                    "trainer_id": user["id"],
                                    "lesson":     lesson,
                                    "week":       week,
                                    "year":       year,
                                    "term":       term,
                                    "status":     "absent"
                                }).execute()
                        write_audit_log("add_class_event", target=f"class:{class_id},type:{event_type}")
                        flash(f"Marked as {label} successfully.", "success")
                        return redirect(url_for("trainer.attendance", class_id=class_id, unit_id=unit_id, week=week, lesson=lesson, year=year, term=term))
                    except Exception as e:
                        flash(f"Error marking {label}: {e}", "error")

        elif action == "delete_event":
            event_id = request.form.get("event_id")
            try:
                db.table("class_events").delete().eq("id", event_id).execute()
                write_audit_log("delete_class_event", target=f"event:{event_id}")
                flash("Event removed.", "success")
                return redirect(url_for("trainer.attendance", class_id=class_id, unit_id=unit_id, week=week, lesson=lesson, year=year, term=term))
            except Exception as e:
                flash(f"Error removing event: {e}", "error")

    return render_template("trainer/attendance.html",
                          class_list=class_list,
                          class_id=class_id,
                          units_list=units_list,
                          unit_id=unit_id,
                          students_list=students_list,
                          week=week,
                          lesson=lesson,
                          year=year,
                          term=term,
                          attendance_submitted=attendance_submitted,
                          active_event=active_event,
                          dept_name=dept_name)


# ── Assessment Review ─────────────────────────────────────────────────────────

def _normalize_poe_assessment_type(assessment_type: str) -> str:
    """Map POE upload type to formative Oral/Practical/Theory."""
    t = (assessment_type or "").strip().lower()
    if t == "oral":
        return "Oral"
    if t == "practical":
        return "Practical"
    if t in ("theory", "written"):
        return "Theory"
    return (assessment_type or "").strip().title() or "Theory"


def _match_formative_assessment(formative_list, poe_type_norm, assessment_no):
    """Find the formative assessment that corresponds to a POE upload."""
    if not formative_list:
        return None
    ano = int(assessment_no or 1)
    typed = [a for a in formative_list if a.get("assessment_type") == poe_type_norm]
    if not typed:
        return None

    name_patterns = {
        f"{poe_type_norm} {ano}",
        f"{poe_type_norm.lower()} {ano}",
    }
    if poe_type_norm == "Theory":
        name_patterns.update({
            f"Theory {ano}",
            f"Written Assessment {ano}",
        })

    lowered = {p.lower() for p in name_patterns}
    for fa in typed:
        if (fa.get("assessment_name") or "").strip().lower() in lowered:
            return fa

    for fa in typed:
        name = (fa.get("assessment_name") or "").strip()
        m = re.search(r"(\d+)\s*$", name)
        if m and int(m.group(1)) == ano:
            return fa

    ordered = sorted(typed, key=lambda x: x.get("created_at") or "")
    if 1 <= ano <= len(ordered):
        return ordered[ano - 1]
    return None


def _bulk_formative_marks_for_poe(db, poe_assessments):
    """Attach marks_obtained/max_marks from formative_marks (Marks Entry source)."""
    if not poe_assessments:
        return poe_assessments

    scopes = {}
    for a in poe_assessments:
        key = (a.get("class_id"), a.get("unit_id"), a.get("year"), int(a.get("term") or 1))
        scopes.setdefault(key, []).append(a)

    fa_by_scope = {}
    for class_id, unit_id, year, term in scopes:
        if not (class_id and unit_id):
            continue
        rows = (db.table("formative_assessments")
                  .select("id, assessment_type, assessment_name, max_marks, created_at")
                  .eq("class_id", class_id)
                  .eq("unit_id", unit_id)
                  .eq("year", year)
                  .eq("term", term)
                  .execute().data or [])
        fa_by_scope[(class_id, unit_id, year, term)] = rows

    all_fa_ids = [fa["id"] for rows in fa_by_scope.values() for fa in rows]
    marks_lookup = {}
    if all_fa_ids:
        for m in (db.table("formative_marks")
                    .select("assessment_id, student_id, marks_obtained")
                    .in_("assessment_id", all_fa_ids)
                    .execute().data or []):
            marks_lookup[(m["student_id"], m["assessment_id"])] = m.get("marks_obtained")

    fa_max = {fa["id"]: float(fa.get("max_marks") or 100)
              for rows in fa_by_scope.values() for fa in rows}

    for a in poe_assessments:
        key = (a.get("class_id"), a.get("unit_id"), a.get("year"), int(a.get("term") or 1))
        formative_list = fa_by_scope.get(key, [])
        poe_type = _normalize_poe_assessment_type(a.get("assessment_type"))
        fa = _match_formative_assessment(formative_list, poe_type, a.get("assessment_no"))
        if fa:
            mark = marks_lookup.get((a.get("student_id"), fa["id"]))
            a["marks_obtained"] = float(mark) if mark is not None else 0
            a["max_marks"] = fa_max.get(fa["id"], 100)
        else:
            a["marks_obtained"] = 0
            a["max_marks"] = 100

    return poe_assessments


def _formative_mark_for_poe(db, poe_assessment):
    """Return (marks_obtained, max_marks) for a single POE assessment."""
    class_id = poe_assessment.get("class_id")
    unit_id = poe_assessment.get("unit_id")
    year = poe_assessment.get("year")
    term = int(poe_assessment.get("term") or 1)
    if not (class_id and unit_id):
        return 0, 100

    formative_list = (db.table("formative_assessments")
                        .select("id, assessment_type, assessment_name, max_marks, created_at")
                        .eq("class_id", class_id)
                        .eq("unit_id", unit_id)
                        .eq("year", year)
                        .eq("term", term)
                        .execute().data or [])
    poe_type = _normalize_poe_assessment_type(poe_assessment.get("assessment_type"))
    fa = _match_formative_assessment(formative_list, poe_type, poe_assessment.get("assessment_no"))
    if not fa:
        return 0, 100

    row = (db.table("formative_marks")
             .select("marks_obtained")
             .eq("assessment_id", fa["id"])
             .eq("student_id", poe_assessment["student_id"])
             .limit(1)
             .execute().data or [])
    max_m = float(fa.get("max_marks") or 100)
    if row:
        return float(row[0].get("marks_obtained", 0)), max_m
    return 0, max_m


@trainer_bp.route("/assessments")
@trainer_required
def assessments():
    db = get_service_client()
    user = current_user()
    assigned_unit_ids = _trainer_assigned_unit_ids(db)

    # Never omit the unit filter — empty assignment must mean empty results, not "all".
    if not assigned_unit_ids:
        assessments_list = []
    else:
        assessments_list = (db.table("assessments").select(
            "*, "
            "user_profiles!assessments_student_id_fkey(full_name, admission_no), "
            "reviewer:user_profiles!assessments_reviewed_by_fkey(full_name), "
            "units(name, code), "
            "classes(id, name)"
        ).in_("unit_id", assigned_unit_ids)
         .order("uploaded_at", desc=True)
         .execute().data or [])

    # Same marks source as Marks Entry (formative_assessments + formative_marks)
    _bulk_formative_marks_for_poe(db, assessments_list)

    # Build class/unit hierarchy for drill-down
    classes_map = {}
    units_map = {}
    status_counts = {"total": 0, "pending": 0, "approved": 0, "rejected": 0}
    for a in assessments_list:
        status_counts["total"] += 1
        s = a.get("status", "pending")
        if s in status_counts:
            status_counts[s] += 1

        cls = a.get("classes") or {}
        cid = cls.get("id")
        if cid:
            if cid not in classes_map:
                classes_map[cid] = {"id": cid, "name": cls.get("name", ""), "units": {}}
            u = a.get("units") or {}
            uid = a.get("unit_id")
            if uid:
                if uid not in classes_map[cid]["units"]:
                    classes_map[cid]["units"][uid] = {
                        "id": uid,
                        "name": u.get("name", ""),
                        "code": u.get("code", ""),
                        "total": 0, "pending": 0, "approved": 0, "rejected": 0
                    }
                classes_map[cid]["units"][uid]["total"] += 1
                classes_map[cid]["units"][uid][s] += 1
                classes_map[cid]["units"][uid]["assessments"] = classes_map[cid]["units"][uid].get("assessments", []) + [a]

    # Also get assigned units for the class list
    class_list = []
    for cid, cdata in classes_map.items():
        unit_list = list(cdata["units"].values())
        unit_pending = sum(u["pending"] for u in unit_list)
        class_list.append({
            "id": cid,
            "name": cdata["name"],
            "units": sorted(unit_list, key=lambda u: u["name"]),
            "unit_count": len(unit_list),
            "pending": unit_pending
        })
    class_list.sort(key=lambda c: c["name"])

    return render_template("trainer/assessments.html",
                          classes=class_list,
                          status_counts=status_counts)


@trainer_bp.route("/assessment/<assessment_id>/review", methods=["GET", "POST"])
@trainer_required
def review_assessment(assessment_id):
    db = get_service_client()
    user = current_user()
    
    # Fetch assessment first so we can check unit access with the correct unit_id
    assessment = (db.table("assessments")
                  .select("*, user_profiles!assessments_student_id_fkey(full_name, admission_no, mobile_number), units(name, code), classes(name)")
                  .eq("id", assessment_id)
                  .limit(1)
                  .execute().data or [None])[0]
    
    if not assessment:
        abort(404)
    
    if not _check_unit_access(db, assessment["unit_id"]):
        abort(403)
    
    obtained, max_m = _formative_mark_for_poe(db, assessment)
    assessment["marks_obtained"] = obtained
    assessment["max_marks"] = max_m
    
    # Get evidence
    evidence = db.table("evidence").select("*").eq("assessment_id", assessment_id).execute().data or []
    
    if request.method == "POST":
        action = request.form.get("action")
        review_note = request.form.get("review_note", "").strip()

        if action in ["approve", "reject"]:
            new_status = "approved" if action == "approve" else "rejected"
            trainer_name = user.get("full_name", "Trainer")

            # ── Critical DB update (always attempt redirect after) ──────────
            try:
                db.table("assessments").update({
                    "status":      new_status,
                    "reviewed_by": user["id"],
                    "reviewed_at": datetime.now().isoformat(),
                    "review_note": review_note,
                }).eq("id", assessment_id).execute()
                flash(f"Assessment {new_status} by {trainer_name}.", "success")
            except Exception as e:
                flash(f"Error updating assessment: {e}", "error")

            # ── Non-critical: cosmetic file rename & audit (never block redirect)
            try:
                _rename_script_file(db, assessment_id, new_status, trainer_name)
            except Exception:
                pass
            try:
                write_audit_log(f"review_assessment_{action}", target=f"assessment:{assessment_id}")
            except Exception:
                pass

            # Always redirect — even if something failed
            return redirect(url_for("trainer.assessments"))

    # ── Fetch reviewer profile for display ────────────────────────────────────
    reviewer = None
    if assessment.get("reviewed_by"):
        try:
            reviewer_rows = (db.table("user_profiles")
                             .select("full_name")
                             .eq("id", assessment["reviewed_by"])
                             .limit(1)
                             .execute().data or [])
            reviewer = reviewer_rows[0] if reviewer_rows else None
        except Exception:
            pass

    return render_template("trainer/review_assessment.html",
                          assessment=assessment,
                          evidence=evidence,
                          reviewer=reviewer)


@trainer_bp.route("/assessment/<assessment_id>/delete", methods=["POST"])
@trainer_required
def delete_assessment(assessment_id):
    """Delete an assessment and its evidence."""
    db = get_service_client()
    user = current_user()

    assessment = db.table("assessments").select("*").eq("id", assessment_id).single().execute().data
    if not assessment:
        abort(404)
    if not _check_unit_access(db, assessment["unit_id"]):
        abort(403)

    try:
        evidence_list = db.table("evidence").select("*").eq("assessment_id", assessment_id).execute().data or []

        # Delete evidence files from storage (file_path is a relative storage path)
        svc = get_service_client()
        for e in evidence_list:
            try:
                if e.get("file_path"):
                    svc.storage.from_("assessment-evidence").remove([e["file_path"]])
            except Exception:
                pass

        # Delete script file from storage
        try:
            sfp = assessment.get("script_file_path", "")
            if sfp:
                svc.storage.from_("assessment-scripts").remove([sfp])
        except Exception:
            pass

        # Delete evidence records
        if evidence_list:
            eids = [e["id"] for e in evidence_list]
            db.table("evidence").delete().in_("id", eids).execute()

        # Delete assessment record
        db.table("assessments").delete().eq("id", assessment_id).execute()

        write_audit_log("delete_assessment", target=f"assessment:{assessment_id}")
        return jsonify({"success": True, "message": "Assessment deleted successfully."})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ── View & Download Attendance ───────────────────────────────────────────────

@trainer_bp.route("/attendance-history")
@trainer_required
def attendance_history():
    """View and download attendance — summary or per-session detail."""
    db   = get_service_client()
    user = current_user()

    cu_rows, class_list = _marks_class_unit_data(db, user)

    class_id = request.args.get("class_id", "")
    unit_id  = request.args.get("unit_id", "")
    week     = request.args.get("week", 0, type=int)
    lesson   = request.args.get("lesson", "")
    year     = request.args.get("year", datetime.now().year, type=int)
    term     = request.args.get("term", 1, type=int)

    units_list   = []
    sessions     = []   # list of {week, lesson, present, absent, total, date, event}
    records      = []   # per-student rows for one specific session
    active_event = None
    cls_obj      = {}
    unit_obj     = {}

    if class_id:
        units_list = [r for r in cu_rows
                      if (r.get("classes") or {}).get("id") == class_id]

    if class_id and unit_id:
        cls_obj  = (db.table("classes").select("name")
                      .eq("id", class_id).single().execute().data or {})
        unit_obj = (db.table("units").select("code, name")
                      .eq("id", unit_id).single().execute().data or {})

        if week and lesson:
            # ── Detail view: one session ──────────────────────────────────
            records = (db.table("attendance")
                         .select("*, user_profiles:student_id(full_name, admission_no)")
                         .eq("unit_id", unit_id)
                         .eq("trainer_id", user["id"])
                         .eq("week", week)
                         .eq("lesson", lesson)
                         .eq("year", year)
                         .eq("term", term)
                         .order("attendance_date")
                         .execute().data or [])

            event_row = (db.table("class_events").select("*")
                           .eq("class_id", class_id)
                           .eq("week", week)
                           .eq("lesson", lesson)
                           .eq("year", year)
                           .eq("term", term)
                           .execute().data or [])
            active_event = event_row[0] if event_row else None

        else:
            # ── Summary view: all sessions for class/unit/year/term ───────
            att_rows = (db.table("attendance")
                          .select("week, lesson, status, attendance_date")
                          .eq("unit_id", unit_id)
                          .eq("trainer_id", user["id"])
                          .eq("year", year)
                          .eq("term", term)
                          .execute().data or [])

            from collections import defaultdict
            session_map = defaultdict(lambda: {"present": 0, "absent": 0, "date": ""})
            for r in att_rows:
                key = (r["week"], r["lesson"])
                session_map[key][r["status"]] = session_map[key].get(r["status"], 0) + 1
                if r.get("attendance_date") and not session_map[key]["date"]:
                    session_map[key]["date"] = r["attendance_date"][:10]

            # Events for this class/unit/year/term
            event_rows = (db.table("class_events")
                            .select("week, lesson, event_type, note")
                            .eq("class_id", class_id)
                            .eq("year", year)
                            .eq("term", term)
                            .execute().data or [])
            event_map = {(e["week"], e["lesson"]): e for e in event_rows}

            for (wk, les), counts in sorted(session_map.items()):
                p = counts.get("present", 0)
                a = counts.get("absent", 0)
                sessions.append({
                    "week":    wk,
                    "lesson":  les,
                    "present": p,
                    "absent":  a,
                    "total":   p + a,
                    "date":    counts["date"],
                    "event":   event_map.get((wk, les)),
                })

    return render_template("trainer/attendance_history.html",
                           class_list=class_list,
                           units_list=units_list,
                           cls=cls_obj, unit=unit_obj,
                           sessions=sessions,
                           records=records,
                           active_event=active_event,
                           class_id=class_id, unit_id=unit_id,
                           week=week, lesson=lesson,
                           year=year, term=term)


# ── Session View (filtered attendance for one session) ───────────────────────

@trainer_bp.route("/view-session")
@trainer_required
def view_session():
    db   = get_service_client()
    user = current_user()

    class_id = request.args.get("class_id", "")
    unit_id  = request.args.get("unit_id", "")
    week     = request.args.get("week", 0, type=int)
    lesson   = request.args.get("lesson", "")
    year     = request.args.get("year", datetime.now().year, type=int)
    term     = request.args.get("term", 1, type=int)

    if not (class_id and unit_id and week and lesson):
        flash("Please select Class, Unit, Week and Lesson first.", "error")
        return redirect(url_for("trainer.attendance"))

    cls  = (db.table("classes").select("name")
              .eq("id", class_id).single().execute().data or {})
    unit = (db.table("units").select("code, name")
              .eq("id", unit_id).single().execute().data or {})
    dept = {}
    if user.get("department_id"):
        dept = (db.table("departments").select("name")
                  .eq("id", user["department_id"]).single().execute().data or {})

    records = (db.table("attendance")
                 .select("*, user_profiles:student_id(full_name, admission_no)")
                 .eq("unit_id", unit_id)
                 .eq("trainer_id", user["id"])
                 .eq("week", week)
                 .eq("lesson", lesson)
                 .eq("year", year)
                 .eq("term", term)
                 .order("attendance_date")
                 .execute().data or [])

    event_row = (db.table("class_events").select("*")
                   .eq("class_id", class_id)
                   .eq("week", week)
                   .eq("lesson", lesson)
                   .eq("year", year)
                   .eq("term", term)
                   .execute().data or [])
    active_event = event_row[0] if event_row else None

    return render_template("trainer/view_session.html",
                           cls=cls, unit=unit, dept=dept,
                           records=records, active_event=active_event,
                           class_id=class_id, unit_id=unit_id,
                           week=week, lesson=lesson, year=year, term=term,
                           trainer={"name": user.get("full_name", "")})


# ── Correct single attendance record ─────────────────────────────────────────

@trainer_bp.route("/attendance/<record_id>/correct", methods=["POST"])
@trainer_required
def correct_attendance(record_id):
    """Change one trainee's status for an already-submitted session without touching others."""
    db   = get_service_client()
    user = current_user()

    new_status = (request.form.get("new_status") or "").strip()
    if new_status not in ("present", "absent"):
        flash("Invalid status value.", "error")
        return redirect(request.referrer or url_for("trainer.attendance"))

    # Verify the record belongs to this trainer
    rec = (db.table("attendance")
             .select("id, trainer_id, student_id, status")
             .eq("id", record_id)
             .eq("trainer_id", user["id"])
             .limit(1)
             .execute().data or [])
    if not rec:
        flash("Record not found or you do not have permission to edit it.", "error")
        return redirect(request.referrer or url_for("trainer.attendance"))

    old_status = rec[0]["status"]
    if old_status == new_status:
        flash("Status is already set to that value — no change made.", "info")
        return redirect(request.referrer or url_for("trainer.attendance"))

    try:
        db.table("attendance").update({"status": new_status}).eq("id", record_id).execute()
        write_audit_log("correct_attendance", target=f"attendance:{record_id}",
                        detail={"old": old_status, "new": new_status,
                                "student_id": rec[0]["student_id"]})
        flash(f"Attendance corrected: {old_status} → {new_status}.", "success")
    except Exception as exc:
        flash(f"Error correcting attendance: {exc}", "error")

    return redirect(request.referrer or url_for("trainer.view_session"))


# ── 10-Week attendance export (Excel) ─────────────────────────────────────────

@trainer_bp.route("/attendance/weekly-export")
@trainer_required
def attendance_weekly_export():
    """Download a horizontal 10-week attendance sheet per unit/class as Excel."""
    import io
    from itertools import groupby as _groupby
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import pytz

    db   = get_service_client()
    user = current_user()

    class_id   = request.args.get("class_id", "")
    unit_id    = request.args.get("unit_id", "")
    year       = request.args.get("year",       datetime.now().year, type=int)
    term       = request.args.get("term",       1,                   type=int)
    week_start = request.args.get("week_start", 1,                   type=int)
    week_end   = request.args.get("week_end",   10,                  type=int)

    if not (class_id and unit_id):
        flash("Class and Unit are required for export.", "error")
        return redirect(url_for("trainer.attendance"))

    # Access check — use same logic as every other trainer route
    if not _check_unit_access(db, unit_id):
        flash("Access denied.", "error")
        return redirect(url_for("trainer.attendance"))

    cls  = db.table("classes").select("name").eq("id", class_id).single().execute().data or {}
    unit = db.table("units").select("code, name").eq("id", unit_id).single().execute().data or {}

    students = (db.table("enrollments")
                  .select("student_id, "
                          "user_profiles!enrollments_student_id_fkey(full_name, admission_no)")
                  .eq("class_id", class_id)
                  .order("student_id")
                  .execute().data or [])
    student_ids = [s["student_id"] for s in students]

    att_rows = []
    if student_ids:
        att_rows = (db.table("attendance")
                      .select("student_id, week, lesson, status, attendance_date")
                      .eq("unit_id", unit_id)
                      .in_("student_id", student_ids)
                      .eq("year", year).eq("term", term)
                      .gte("week", week_start).lte("week", week_end)
                      .order("week").order("lesson")
                      .execute().data or [])

    def _nl(l):
        s = str(l).strip()
        return f"L{s}" if s in ("1", "2", "3", "4") else s

    # All distinct sessions ordered (week, lesson) — normalize lesson to L1/L2/L3/L4
    sessions = sorted({(r["week"], _nl(r["lesson"])) for r in att_rows}, key=lambda x: (x[0], x[1]))
    if not sessions:
        sessions = [(w, "L1") for w in range(week_start, min(week_end, week_start + 9) + 1)]

    # Lookup {student_id: {(week,lesson): (status, ts_str)}}
    EAT = pytz.timezone("Africa/Nairobi")
    att = {}
    for r in att_rows:
        ts_str = ""
        if r.get("attendance_date"):
            try:
                from datetime import datetime as _dt
                dt = _dt.fromisoformat(str(r["attendance_date"]).replace("Z", "+00:00"))
                ts_str = dt.astimezone(EAT).strftime("%H:%M")
            except Exception:
                ts_str = str(r["attendance_date"])[11:16]
        att.setdefault(r["student_id"], {})[(r["week"], _nl(r["lesson"]))] = (r["status"], ts_str)

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    GREEN   = PatternFill("solid", fgColor="DCFCE7")
    RED     = PatternFill("solid", fgColor="FEE2E2")
    ALT     = PatternFill("solid", fgColor="EFF6FF")
    thin    = Side(style="thin", color="B0C4D8")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_fix = 3          # S/N, Adm No, Name
    n_ses = len(sessions)
    n_sum = 3          # Total P, Total A, %
    n_tot = n_fix + n_ses + n_sum

    def hdr_cell(row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        return c

    # Common letterhead (logo, institute, document title, department, meta)
    dept_name = ""
    if user.get("department_id"):
        _dept = (db.table("departments").select("name")
                   .eq("id", user["department_id"]).single().execute().data or {})
        dept_name = _dept.get("name", "")

    week_row = excel_letterhead(
        ws, "Weekly Attendance Register", n_tot, dept_name=dept_name,
        meta_lines=[
            f"Class: {cls.get('name','')}  |  Unit: {unit.get('code','')} — {unit.get('name','')}",
            f"Year: {year}  |  Term: {term}  |  Weeks {week_start}–{week_end}",
            f"Trainer: {user.get('full_name','')}  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        ],
    )

    # Week group headers (merged per week)
    for ci in range(1, n_fix + 1):
        hdr_cell(week_row, ci, "")
    # Merge cells per week
    col = n_fix + 1
    for wk, grp in _groupby(sessions, key=lambda x: x[0]):
        grp_list = list(grp)
        span = len(grp_list)
        if span > 1:
            ws.merge_cells(start_row=week_row, start_column=col,
                           end_row=week_row, end_column=col + span - 1)
        hdr_cell(week_row, col, f"WEEK {wk}")
        col += span
    for i, lbl in enumerate(["Total P", "Total A", "%"], col):
        hdr_cell(week_row, i, lbl)
    ws.row_dimensions[week_row].height = 22

    # Sub-header (S/N, Adm, Name, lesson labels, summary)
    sub_row = week_row + 1
    hdr_cell(sub_row, 1, "S/N")
    hdr_cell(sub_row, 2, "Adm. No.")
    hdr_cell(sub_row, 3, "Student Name")
    for i, (w, l) in enumerate(sessions, n_fix + 1):
        hdr_cell(sub_row, i, l)
    for i, lbl in enumerate(["Total P", "Total A", "%"], n_fix + n_ses + 1):
        hdr_cell(sub_row, i, lbl)
    ws.row_dimensions[sub_row].height = 22

    # Light, visible styling for both stacked header rows
    style_header_row(ws, week_row, n_tot)
    style_header_row(ws, sub_row, n_tot)

    # Data rows
    for idx, stu in enumerate(students, 1):
        sid  = stu["student_id"]
        prof = stu.get("user_profiles") or {}
        row  = [idx, prof.get("admission_no", ""), prof.get("full_name", "")]
        tot_p = tot_a = 0

        for w, l in sessions:
            entry = att.get(sid, {}).get((w, l))
            if entry:
                st, ts = entry
                sym = "✔" if st == "present" else "✘"  # ✔ or ✘
                row.append(f"{sym}\n{ts}" if ts else sym)
                if st == "present": tot_p += 1
                else: tot_a += 1
            else:
                row.append("")

        total = tot_p + tot_a
        row += [tot_p, tot_a, f"{round(tot_p/total*100)}%" if total else "—"]
        ws.append(row)

        dr = ws.max_row
        ws.row_dimensions[dr].height = 30

        for ci in range(1, n_tot + 1):
            c = ws.cell(row=dr, column=ci)
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # Colour status cells
            if n_fix < ci <= n_fix + n_ses:
                v = str(c.value or "")
                if v.startswith("✔"):
                    c.fill = GREEN
                    c.font = Font(bold=True, size=11, color="15803D")
                elif v.startswith("✘"):
                    c.fill = RED
                    c.font = Font(bold=True, size=11, color="B91C1C")
            elif ci <= n_fix and idx % 2 == 0:
                c.fill = ALT

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 26
    for ci in range(n_fix + 1, n_fix + n_ses + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    for ci in range(n_fix + n_ses + 1, n_tot + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 9

    ws.freeze_panes = ws.cell(row=sub_row + 1, column=n_fix + 1)

    # Official sign-off block
    excel_signature_block(ws, ws.max_row + 2, officer_title="Trainer",
                          officer_name=user.get("full_name", ""),
                          extra_officers=("Head of Department",))

    try:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as exc:
        flash(f"Excel build error: {exc}", "error")
        return redirect(url_for("trainer.attendance"))

    safe_cls  = (cls.get("name","Class") or "Class").replace("/", "-").replace(" ", "_")
    safe_unit = (unit.get("code","Unit") or "Unit").replace("/", "-")
    fname = f"Attendance_{safe_cls}_{safe_unit}_Wk{week_start}-{week_end}_Yr{year}T{term}.xlsx"

    from flask import Response as _R
    return _R(buf.read(),
              mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
              headers={"Content-Disposition": f"attachment; filename={fname}"})


# ── Session PDF (print-ready attendance for one session) ─────────────────────

@trainer_bp.route("/session-pdf")
@trainer_required
def session_pdf():
    db   = get_service_client()
    user = current_user()

    class_id = request.args.get("class_id", "")
    unit_id  = request.args.get("unit_id", "")
    week     = request.args.get("week", 0, type=int)
    lesson   = request.args.get("lesson", "")
    year     = request.args.get("year", datetime.now().year, type=int)
    term     = request.args.get("term", 1, type=int)

    if not (class_id and unit_id and week and lesson):
        flash("Select a complete session to download.", "error")
        return redirect(url_for("trainer.attendance"))

    cls  = (db.table("classes").select("name")
              .eq("id", class_id).single().execute().data or {})
    unit = (db.table("units").select("code, name")
              .eq("id", unit_id).single().execute().data or {})
    dept = {}
    if user.get("department_id"):
        dept = (db.table("departments").select("name")
                  .eq("id", user["department_id"]).single().execute().data or {})

    records = (db.table("attendance")
                 .select("*, user_profiles:student_id(full_name, admission_no)")
                 .eq("unit_id", unit_id)
                 .eq("trainer_id", user["id"])
                 .eq("week", week)
                 .eq("lesson", lesson)
                 .eq("year", year)
                 .eq("term", term)
                 .order("attendance_date")
                 .execute().data or [])

    event_row = (db.table("class_events").select("*")
                   .eq("class_id", class_id)
                   .eq("week", week)
                   .eq("lesson", lesson)
                   .eq("year", year)
                   .eq("term", term)
                   .execute().data or [])
    active_event = event_row[0] if event_row else None

    generated = datetime.now().strftime("%d %b %Y %H:%M")

    return render_template("trainer/session_pdf.html",
                           cls=cls, unit=unit, dept=dept,
                           records=records, active_event=active_event,
                           week=week, lesson=lesson, year=year, term=term,
                           trainer={"name": user.get("full_name", "")},
                           generated=generated)


@trainer_bp.route("/unit-attendance-pdf")
@trainer_required
def unit_attendance_pdf():
    """
    Landscape attendance register for one unit — all weeks/lessons the trainer
    has taught in the selected year/term, as a single printable page.
    Column headers show the real date/time attendance was taken.
    """
    from unit_attendance_register import build_unit_attendance_register

    db   = get_service_client()
    user = current_user()

    class_id = request.args.get("class_id", "")
    unit_id  = request.args.get("unit_id", "")
    year     = request.args.get("year", datetime.now().year, type=int)
    term     = request.args.get("term", 1, type=int)

    if not (class_id and unit_id):
        flash("Select a Class and Unit to download the unit attendance register.", "error")
        return redirect(url_for("trainer.attendance_history"))

    assigned = (db.table("class_units")
                  .select("id")
                  .eq("class_id", class_id)
                  .eq("unit_id", unit_id)
                  .eq("trainer_id", user["id"])
                  .limit(1).execute().data or [])
    if not assigned:
        flash("You are not assigned to this class/unit.", "error")
        return redirect(url_for("trainer.attendance_history"))

    data = build_unit_attendance_register(
        db,
        class_id=class_id,
        unit_id=unit_id,
        year=year,
        term=term,
        trainer_id=user["id"],
    )
    if not data:
        flash("No attendance records found for this unit in the selected period.", "warning")
        return redirect(url_for("trainer.attendance_history",
                                class_id=class_id, unit_id=unit_id, year=year, term=term))

    return render_template("shared/unit_attendance_pdf.html", **data)


# ── Formative Assessment Marks ───────────────────────────────────────────────

def _marks_class_unit_data(db, user):
    """Return cu_rows, class_list for this trainer."""
    cu_rows = (db.table("class_units")
                 .select("class_id, unit_id, units(id,code,name), classes(id,name)")
                 .eq("trainer_id", user["id"])
                 .execute().data or [])
    class_map = {}
    for r in cu_rows:
        c = r.get("classes") or {}
        if c.get("id"):
            class_map[c["id"]] = c["name"]
    class_list = sorted([{"id": k, "name": v} for k, v in class_map.items()],
                        key=lambda x: x["name"])
    return cu_rows, class_list


def _load_assessments_and_marks(db, unit_id, class_id, trainer_id, year, term):
    """Return (assessments, marks_map {sid: {aid: marks}})."""
    assessments = (db.table("formative_assessments")
                    .select("*")
                    .eq("unit_id",    unit_id)
                    .eq("class_id",   class_id)
                    .eq("trainer_id", trainer_id)
                    .eq("year",       year)
                    .eq("term",       term)
                    .order("assessment_type")
                    .order("created_at")
                    .execute().data or [])
    marks_map = {}
    if assessments:
        a_ids = [a["id"] for a in assessments]
        for m in (db.table("formative_marks")
                    .select("assessment_id, student_id, marks_obtained")
                    .in_("assessment_id", a_ids)
                    .execute().data or []):
            sid = m["student_id"]
            aid = m["assessment_id"]
            if sid not in marks_map:
                marks_map[sid] = {}
            marks_map[sid][aid] = m["marks_obtained"]
    return assessments, marks_map


@trainer_bp.route("/marks-entry")
@trainer_required
def marks_entry():
    """Formative assessment marks entry — dynamic grid view."""
    db   = get_service_client()
    user = current_user()
    cu_rows, class_list = _marks_class_unit_data(db, user)

    class_id = request.args.get("class_id", "")
    unit_id  = request.args.get("unit_id", "")
    year     = request.args.get("year",  datetime.now().year, type=int)
    term     = request.args.get("term",  1, type=int)

    units_list    = []
    students_list = []
    assessments   = []
    marks_map     = {}

    if class_id:
        units_list = [r for r in cu_rows if (r.get("classes") or {}).get("id") == class_id]

    if class_id and unit_id:
        raw = (db.table("enrollments")
                 .select("student_id, user_profiles(full_name, admission_no)")
                 .eq("class_id", class_id).execute().data or [])
        students_list = sorted(raw, key=lambda s: (s.get("user_profiles") or {}).get("full_name", ""))
        assessments, marks_map = _load_assessments_and_marks(db, unit_id, class_id, user["id"], year, term)

    oral_list      = [a for a in assessments if a["assessment_type"] == "Oral"]
    practical_list = [a for a in assessments if a["assessment_type"] == "Practical"]
    theory_list    = [a for a in assessments if a["assessment_type"] == "Theory"]

    return render_template("trainer/marks_entry.html",
                           class_list=class_list,
                           units_list=units_list,
                           students_list=students_list,
                           assessments=assessments,
                           oral_list=oral_list,
                           practical_list=practical_list,
                           theory_list=theory_list,
                           marks_map=marks_map,
                           class_id=class_id,
                           unit_id=unit_id,
                           year=year, term=term)


@trainer_bp.route("/marks-entry/add-assessment", methods=["POST"])
@trainer_required
def add_assessment():
    """AJAX — create a new formative assessment definition."""
    db   = get_service_client()
    user = current_user()
    data = request.get_json() or {}

    unit_id         = data.get("unit_id", "").strip()
    class_id        = data.get("class_id", "").strip()
    assessment_type = data.get("assessment_type", "").strip()
    assessment_name = data.get("assessment_name", "").strip()
    max_marks       = data.get("max_marks", 100)
    year            = int(data.get("year", datetime.now().year))
    term            = int(data.get("term", 1))

    if not all([unit_id, class_id, assessment_type, assessment_name]):
        return jsonify({"success": False, "message": "All fields are required."}), 400
    if assessment_type not in ("Oral", "Practical", "Theory"):
        return jsonify({"success": False, "message": "Invalid type."}), 400

    try:
        max_marks_val = float(max_marks)
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "Maximum marks must be a number."}), 400
    if max_marks_val < 1 or max_marks_val > 100:
        return jsonify({
            "success": False,
            "message": "Maximum marks must be between 1 and 100. Scores convert to % out of 100.",
        }), 400

    if not _trainer_owns_class_unit(db, class_id, unit_id):
        return jsonify({"success": False, "message": "You are not assigned to this class/unit."}), 403

    dup = (db.table("formative_assessments").select("id")
             .eq("unit_id", unit_id).eq("class_id", class_id)
             .eq("trainer_id", user["id"])
             .eq("assessment_name", assessment_name)
             .eq("year", year).eq("term", term)
             .execute().data or [])
    if dup:
        return jsonify({"success": False, "message": f"'{assessment_name}' already exists."}), 400

    try:
        result = db.table("formative_assessments").insert({
            "unit_id": unit_id, "class_id": class_id,
            "trainer_id": user["id"],
            "assessment_type": assessment_type,
            "assessment_name": assessment_name,
            "max_marks": max_marks_val,
            "year": year, "term": term
        }).execute()
        write_audit_log("add_formative_assessment",
                        target=f"unit:{unit_id},{assessment_type}:{assessment_name}")
        return jsonify({"success": True, "assessment": result.data[0] if result.data else {}})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@trainer_bp.route("/marks-entry/delete-assessment", methods=["POST"])
@trainer_required
def delete_formative_assessment():
    """Delete a formative assessment and all its marks."""
    db   = get_service_client()
    user = current_user()
    assessment_id = request.form.get("assessment_id", "")
    class_id = request.form.get("class_id", "")
    unit_id  = request.form.get("unit_id", "")
    year     = request.form.get("year", "")
    term     = request.form.get("term", "")
    try:
        rec = (db.table("formative_assessments")
                 .select("trainer_id, assessment_name")
                 .eq("id", assessment_id).single().execute().data)
        if not rec or rec["trainer_id"] != user["id"]:
            flash("Access denied.", "error")
        else:
            db.table("formative_assessments").delete().eq("id", assessment_id).execute()
            write_audit_log("delete_formative_assessment", target=f"assessment:{assessment_id}")
            flash(f"'{rec['assessment_name']}' deleted.", "success")
    except Exception as e:
        flash(f"Error: {e}", "error")
    return redirect(url_for("trainer.marks_entry",
                            class_id=class_id, unit_id=unit_id, year=year, term=term))


@trainer_bp.route("/marks-entry/save-mark", methods=["POST"])
@trainer_required
def save_mark():
    """AJAX — upsert a single mark."""
    db   = get_service_client()
    user = current_user()
    data = request.get_json() or {}

    assessment_id = data.get("assessment_id", "")
    student_id    = data.get("student_id", "")
    marks_str     = data.get("marks", "")

    if not assessment_id or not student_id:
        return jsonify({"success": False, "message": "Missing fields."}), 400

    rec = (db.table("formative_assessments")
             .select("trainer_id, max_marks")
             .eq("id", assessment_id).single().execute().data)
    if not rec or rec["trainer_id"] != user["id"]:
        return jsonify({"success": False, "message": "Access denied."}), 403

    # Empty → delete record
    if marks_str == "" or marks_str is None:
        try:
            (db.table("formative_marks").delete()
               .eq("assessment_id", assessment_id)
               .eq("student_id", student_id).execute())
            return jsonify({"success": True, "cleared": True})
        except Exception as e:
            return jsonify({"success": False, "message": str(e)}), 500

    try:
        marks_val = float(marks_str)
    except (ValueError, TypeError):
        return jsonify({"success": False, "message": "Marks must be a number."}), 400

    max_m = float(rec.get("max_marks", 100))
    if marks_val < 0:
        return jsonify({"success": False, "message": "Marks cannot be negative."}), 400
    if marks_val > max_m:
        return jsonify({"success": False, "message": f"Cannot exceed {int(max_m)}."}), 400

    try:
        existing = (db.table("formative_marks").select("id")
                      .eq("assessment_id", assessment_id)
                      .eq("student_id", student_id).execute().data or [])
        if existing:
            db.table("formative_marks").update({
                "marks_obtained": marks_val,
                "uploaded_by":    user["id"],
                "updated_at":     datetime.now().isoformat()
            }).eq("id", existing[0]["id"]).execute()
        else:
            db.table("formative_marks").insert({
                "assessment_id": assessment_id,
                "student_id":    student_id,
                "marks_obtained": marks_val,
                "uploaded_by":   user["id"]
            }).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@trainer_bp.route("/marks-entry/marks-pdf")
@trainer_required
def marks_pdf():
    """Official formative marks PDF (download, not webpage print)."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    db   = get_service_client()
    user = current_user()
    class_id = request.args.get("class_id", "")
    unit_id  = request.args.get("unit_id", "")
    year     = request.args.get("year", datetime.now().year, type=int)
    term     = request.args.get("term", 1, type=int)

    if not (class_id and unit_id):
        flash("Select class and unit.", "error")
        return redirect(url_for("trainer.marks_entry"))

    cls  = (db.table("classes").select("name").eq("id", class_id).single().execute().data or {})
    unit = (db.table("units").select("code,name").eq("id", unit_id).single().execute().data or {})
    dept = {}
    if user.get("department_id"):
        dept = (db.table("departments").select("name")
                  .eq("id", user["department_id"]).single().execute().data or {})

    raw = (db.table("enrollments")
             .select("student_id, user_profiles(full_name, admission_no)")
             .eq("class_id", class_id).execute().data or [])
    students_list = sorted(raw, key=lambda s: (s.get("user_profiles") or {}).get("full_name", ""))
    assessments, marks_map = _load_assessments_and_marks(db, unit_id, class_id, user["id"], year, term)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    cell_s = ParagraphStyle("c", parent=styles["Normal"], fontSize=7, leading=9)
    small = ParagraphStyle(
        "sm", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#64748B"),
        alignment=TA_LEFT, spaceBefore=8,
    )

    story = pdf_letterhead(
        "Official Formative Assessment Marks Sheet", doc.width,
        dept_name=dept.get("name", ""),
        meta_lines=[
            f"Class: {cls.get('name') or '—'} &nbsp;|&nbsp; "
            f"Unit: {unit.get('code') or ''} — {unit.get('name') or '—'} &nbsp;|&nbsp; "
            f"Year: {year} &nbsp;|&nbsp; Term: {term}",
            f"Trainer: {user.get('full_name') or '—'} &nbsp;·&nbsp; "
            f"Trainees: {len(students_list)} &nbsp;·&nbsp; "
            f"Assessments: {len(assessments)} &nbsp;·&nbsp; "
            f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        ],
    )

    from grading_utils import cdacc_code

    headers = ["#", "Adm. No.", "Trainee Name"] + [
        Paragraph(
            f"<b>{a['assessment_name'][:12]}</b><br/>/{int(a.get('max_marks') or 100)}",
            cell_s,
        )
        for a in assessments
    ] + ["Total", "Avg %", "Grade"]
    data = [headers]

    for i, student in enumerate(students_list, 1):
        p = student.get("user_profiles") or {}
        sid = student["student_id"]
        sm = marks_map.get(sid, {})
        total, count, max_total = 0.0, 0, 0.0
        row = [str(i), p.get("admission_no") or "—", Paragraph(p.get("full_name") or "—", cell_s)]
        for a in assessments:
            m = sm.get(a["id"])
            if m is not None:
                total += float(m)
                count += 1
                max_total += float(a.get("max_marks") or 100)
                row.append(f"{float(m):.1f}".rstrip("0").rstrip("."))
            else:
                row.append("—")
        pct = round(total / max_total * 100, 1) if max_total else None
        row.append(f"{round(total, 1)}" if count else "—")
        row.append(f"{pct}" if count and pct is not None else "—")
        row.append(cdacc_code(pct) if count and pct is not None else "—")
        data.append(row)

    name_w = 42 * mm
    assess_w = max(14 * mm, min(22 * mm, (148 * mm) / max(len(assessments), 1)))
    col_widths = [8 * mm, 22 * mm, name_w] + [assess_w] * len(assessments) + [14 * mm, 14 * mm, 14 * mm]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(pdf_header_style_cmds(0) + [
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (2, 1), (2, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("FONTNAME", (-3, 1), (-1, -1), "Helvetica-Bold"),
    ]))
    story.append(table)
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        "TVET CDACC Grading: <b>M</b> = Mastery (80–100%) &nbsp;|&nbsp; <b>P</b> = Proficient (65–79%) &nbsp;|&nbsp; "
        "<b>C</b> = Competent (50–64%) &nbsp;|&nbsp; <b>NYC</b> = Not Yet Competent (0–49%) &nbsp;|&nbsp; "
        "<b>CRNM</b> = Course Requirement Not Met (any marks). Avg % = total marks obtained / total maximum marks.",
        small,
    ))
    story += pdf_signature_block(doc.width, officer_title="Trainer",
                                 officer_name=user.get("full_name", ""),
                                 extra_officers=("Head of Department",))

    doc.build(story)
    buf.seek(0)
    safe_cls = (cls.get("name") or "class").replace(" ", "_")[:36]
    fname = f"TTTI_Formative_Marks_{safe_cls}_T{term}_{year}.pdf"
    resp = make_response(buf.getvalue())
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


@trainer_bp.route("/marks-entry/export-excel")
@trainer_required
def export_marks_excel():
    """Official formative marks Excel workbook."""
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db   = get_service_client()
    user = current_user()
    class_id = request.args.get("class_id", "")
    unit_id  = request.args.get("unit_id", "")
    year     = request.args.get("year", datetime.now().year, type=int)
    term     = request.args.get("term", 1, type=int)

    if not (class_id and unit_id):
        flash("Select class and unit.", "error")
        return redirect(url_for("trainer.marks_entry"))

    cls  = (db.table("classes").select("name").eq("id", class_id).single().execute().data or {})
    unit = (db.table("units").select("code,name").eq("id", unit_id).single().execute().data or {})
    dept = {}
    if user.get("department_id"):
        dept = (db.table("departments").select("name")
                  .eq("id", user["department_id"]).single().execute().data or {})

    raw = (db.table("enrollments")
             .select("student_id, user_profiles(full_name, admission_no)")
             .eq("class_id", class_id).execute().data or [])
    students_list = sorted(raw, key=lambda s: (s.get("user_profiles") or {}).get("full_name", ""))
    assessments, marks_map = _load_assessments_and_marks(db, unit_id, class_id, user["id"], year, term)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formative Marks"

    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left", vertical="center")
    thin     = Side(style="thin", color="CBD5E1")
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
    type_color = {"Oral": "DCFCE7", "Practical": "FFEDD5", "Theory": "E0E7FF"}
    gold_fill = PatternFill("solid", fgColor="FEF3C7")

    total_cols = 3 + len(assessments) + 3  # Total, Avg %, CDACC Grade
    last_col   = get_column_letter(total_cols)

    # Common letterhead
    type_row = excel_letterhead(
        ws, "Official Formative Assessment Marks Sheet", total_cols,
        dept_name=dept.get("name", ""),
        meta_lines=[
            f"Class: {cls.get('name', '')}  |  Unit: {unit.get('code', '')} – {unit.get('name', '')}  |  "
            f"Year: {year}  |  Term: {term}  |  Trainer: {user.get('full_name') or '—'}",
            f"Trainees: {len(students_list)}  ·  Assessments: {len(assessments)}  ·  "
            f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        ],
    )

    # Type sub-header (kept light, per-type colour fills)
    for col_idx, a in enumerate(assessments, start=4):
        c = ws.cell(row=type_row, column=col_idx, value=a["assessment_type"])
        c.fill = PatternFill("solid", fgColor=type_color.get(a["assessment_type"], "E2E8F0"))
        c.font = Font(bold=True, size=8)
        c.alignment = center
        c.border = bdr

    # Column headers
    hdr_row = type_row + 1
    for col_idx, h in enumerate(
        ["#", "Admission No.", "Trainee Name"] +
        [f"{a['assessment_name']}\n(/{int(a['max_marks'])})" for a in assessments] +
        ["Total", "Avg %", "Grade"],
        start=1
    ):
        ws.cell(row=hdr_row, column=col_idx, value=h)
    style_header_row(ws, hdr_row, total_cols)
    ws.row_dimensions[hdr_row].height = 36

    for ri, student in enumerate(students_list, start=1):
        p   = student.get("user_profiles") or {}
        sid = student["student_id"]
        row = ri + hdr_row
        ws.cell(row=row, column=1, value=ri).alignment = center
        ws.cell(row=row, column=1).border = bdr
        ws.cell(row=row, column=2, value=p.get("admission_no", "")).alignment = center
        ws.cell(row=row, column=2).border = bdr
        name_cell = ws.cell(row=row, column=3, value=p.get("full_name", ""))
        name_cell.alignment = left
        name_cell.border = bdr
        sm = marks_map.get(sid, {})
        total, count, max_total = 0.0, 0, 0.0
        for ci, a in enumerate(assessments, start=4):
            m = sm.get(a["id"])
            c = ws.cell(row=row, column=ci, value=float(m) if m is not None else None)
            c.alignment = center
            c.border = bdr
            if m is not None:
                total += float(m)
                count += 1
                max_total += float(a.get("max_marks") or 100)
        from grading_utils import cdacc_code
        pct = round(total / max_total * 100, 1) if max_total else None
        grade = cdacc_code(pct) if count else None
        tc = 4 + len(assessments)
        tot_c = ws.cell(row=row, column=tc, value=round(total, 1) if count else None)
        avg_c = ws.cell(row=row, column=tc + 1, value=pct if count else None)
        grd_c = ws.cell(row=row, column=tc + 2, value=grade)
        for c in (tot_c, avg_c, grd_c):
            c.alignment = center
            c.border = bdr
            c.font = Font(bold=True, size=10)
        grade_fills = {
            "M": "EDE9FE", "P": "DCFCE7", "C": "DBEAFE", "NYC": "FEE2E2",
        }
        if grade in grade_fills:
            grd_c.fill = PatternFill("solid", fgColor=grade_fills[grade])

    legend_row = hdr_row + len(students_list) + 2
    ws.merge_cells(f"A{legend_row}:{last_col}{legend_row}")
    ws[f"A{legend_row}"] = (
        "TVET CDACC Grading: M = Mastery (80-100%) | P = Proficient (65-79%) | C = Competent (50-64%) | "
        "NYC = Not Yet Competent (0-49%) | CRNM = Course Requirement Not Met. "
        "Blank cells = not yet entered. Official formative assessment record — Thika Technical Training Institute."
    )
    ws[f"A{legend_row}"].font = Font(size=8, italic=True, color="64748B")
    ws[f"A{legend_row}"].fill = gold_fill

    ws2 = wb.create_sheet("Authentication")
    r2 = excel_letterhead(
        ws2, "Formative Marks Authentication", 4,
        dept_name=dept.get("name", ""),
        meta_lines=[
            f"Class: {cls.get('name', '')}  |  Unit: {unit.get('code', '')} – {unit.get('name', '')}  |  "
            f"Year: {year}  Term: {term}",
        ],
    )
    excel_signature_block(ws2, r2, officer_title="Trainer",
                          officer_name=user.get("full_name", ""),
                          extra_officers=("Head of Department",))
    ws2.column_dimensions["A"].width = 90

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 28
    for i in range(len(assessments) + 3):
        ws.column_dimensions[get_column_letter(4 + i)].width = 13

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_cls = (cls.get("name") or "class").replace(" ", "_")[:36]
    fname = f"TTTI_Formative_Marks_{safe_cls}_T{term}_{year}.xlsx"
    resp = make_response(buf.getvalue())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


@trainer_bp.route("/marks-import")
@trainer_required
def marks_import():
    """Formative marks Excel import page."""
    db   = get_service_client()
    user = current_user()
    cu_rows, class_list = _marks_class_unit_data(db, user)

    class_id = request.args.get("class_id", "")
    unit_id  = request.args.get("unit_id", "")
    year     = request.args.get("year", datetime.now().year, type=int)
    term     = request.args.get("term", 1, type=int)

    units_list    = []
    assessments   = []
    students_list = []

    if class_id:
        units_list = [r for r in cu_rows if (r.get("classes") or {}).get("id") == class_id]
    if class_id and unit_id:
        assessments, _ = _load_assessments_and_marks(db, unit_id, class_id, user["id"], year, term)
        raw = (db.table("enrollments")
                 .select("student_id, user_profiles(full_name, admission_no)")
                 .eq("class_id", class_id).execute().data or [])
        students_list = sorted(raw, key=lambda s: (s.get("user_profiles") or {}).get("full_name", ""))

    return render_template("trainer/marks_import.html",
                           class_list=class_list, units_list=units_list,
                           assessments=assessments, students_list=students_list,
                           class_id=class_id, unit_id=unit_id, year=year, term=term)


@trainer_bp.route("/marks-import/template")
@trainer_required
def marks_import_template():
    """Download Excel template pre-filled with students and assessment columns."""
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    db   = get_service_client()
    user = current_user()
    class_id = request.args.get("class_id", "")
    unit_id  = request.args.get("unit_id", "")
    year     = request.args.get("year", datetime.now().year, type=int)
    term     = request.args.get("term", 1, type=int)

    students_list = []
    assessments   = []
    if class_id and unit_id:
        raw = (db.table("enrollments")
                 .select("student_id, user_profiles(full_name, admission_no)")
                 .eq("class_id", class_id).execute().data or [])
        students_list = sorted(raw, key=lambda s: (s.get("user_profiles") or {}).get("full_name", ""))
        assessments, _ = _load_assessments_and_marks(db, unit_id, class_id, user["id"], year, term)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks Import"

    center   = Alignment(horizontal="center")
    type_color = {"Oral": "E8F5E9", "Practical": "FFF3E0", "Theory": "EDE7F6"}

    if assessments:
        headers = ["Admission No", "Trainee Name"] + \
                  [f"{a['assessment_name']} ({a['assessment_type']}) /{int(a['max_marks'])}"
                   for a in assessments]
    else:
        headers = ["Admission No", "Trainee Name",
                   "Assessment Name (e.g. Oral 1)",
                   "Assessment Type (Oral/Practical/Theory)", "Marks (0-100)"]

    # Type color sub-header row
    if assessments:
        for ci, a in enumerate(assessments, start=3):
            c = ws.cell(row=1, column=ci, value=a["assessment_type"])
            c.fill = PatternFill("solid", fgColor=type_color.get(a["assessment_type"], "E0E0E0"))
            c.font = Font(bold=True, size=9); c.alignment = center

    # NOTE: no letterhead rows here — the upload parser (upload_marks) expects
    # headers at row 1/2, so we only restyle the header row (light style).
    hdr_row = 2 if assessments else 1
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=hdr_row, column=ci, value=h)
    style_header_row(ws, hdr_row, len(headers))

    start_row = hdr_row + 1
    for ri, s in enumerate(students_list, start=start_row):
        p = s.get("user_profiles") or {}
        ws.cell(row=ri, column=1, value=p.get("admission_no", ""))
        ws.cell(row=ri, column=2, value=p.get("full_name", ""))

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 28
    for i in range(max(len(assessments), 3)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 22

    # Instructions sheet
    instr = wb.create_sheet("Instructions")
    instr["A1"] = "TTTI Formative Marks Import Template — Instructions"
    instr["A1"].font = Font(bold=True, size=13)
    for r, txt in enumerate([
        "1. Do NOT change column headers or Admission No values.",
        "2. Enter marks in the correct column for each assessment.",
        "3. Marks must be between 0 and the maximum shown in the header.",
        "4. Leave blank if a student was not assessed.",
        "5. Save as .xlsx and upload using the Import page.",
        "6. If no assessments exist yet, create them first on the Marks Entry page.",
    ], start=3):
        instr[f"A{r}"] = txt

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = "attachment; filename=marks_import_template.xlsx"
    return resp


@trainer_bp.route("/marks-import/upload", methods=["POST"])
@trainer_required
def upload_marks():
    """Process uploaded Excel marks file into formative_marks."""
    import openpyxl

    db   = get_service_client()
    user = current_user()
    class_id = request.form.get("class_id", "")
    unit_id  = request.form.get("unit_id", "")
    year     = request.form.get("year", str(datetime.now().year))
    term     = request.form.get("term", "1")

    redirect_url = url_for("trainer.marks_import",
                           class_id=class_id, unit_id=unit_id, year=year, term=term)

    if not all([class_id, unit_id, year, term]):
        flash("Class, Unit, Year and Term are required.", "error")
        return redirect(redirect_url)
    if "marks_file" not in request.files or request.files["marks_file"].filename == "":
        flash("No file selected.", "error")
        return redirect(redirect_url)

    try:
        wb = openpyxl.load_workbook(request.files["marks_file"], read_only=True, data_only=True)
        ws = wb["Marks Import"] if "Marks Import" in wb.sheetnames else wb.active

        rows_iter = ws.iter_rows(values_only=True)
        row1 = next(rows_iter, None)
        if row1 is None:
            flash("Empty file.", "error")
            return redirect(redirect_url)

        headers = [str(h).strip() if h else "" for h in row1]

        # If first row is the type sub-header, skip it and read real headers
        if headers and headers[0].lower() in ("oral", "practical", "theory", ""):
            row1 = next(rows_iter, None)
            if row1 is None:
                flash("Empty file.", "error")
                return redirect(redirect_url)
            headers = [str(h).strip() if h else "" for h in row1]

        if "Admission No" not in headers:
            flash("Invalid template: 'Admission No' column not found.", "error")
            return redirect(redirect_url)

        adm_idx = headers.index("Admission No")

        # Load this trainer's assessments for this filter
        assessments, _ = _load_assessments_and_marks(db, unit_id, class_id, user["id"],
                                                     int(year), int(term))
        # Map column index → assessment (match by name prefix)
        col_to_assessment = {}
        for ci, h in enumerate(headers):
            for a in assessments:
                if h.startswith(a["assessment_name"]):
                    col_to_assessment[ci] = a
                    break

        processed, errors = 0, []
        for row_vals in rows_iter:
            if not row_vals or row_vals[adm_idx] is None:
                continue
            adm_no = str(row_vals[adm_idx]).strip()
            if not adm_no:
                continue

            stu = (db.table("user_profiles").select("id")
                     .eq("admission_no", adm_no).eq("role", "student")
                     .execute().data or [])
            if not stu:
                errors.append(f"'{adm_no}' not found.")
                continue
            sid = stu[0]["id"]

            for ci, assessment in col_to_assessment.items():
                cell_val = row_vals[ci] if ci < len(row_vals) else None
                if cell_val is None or str(cell_val).strip() == "":
                    continue
                try:
                    mv = float(str(cell_val).strip())
                except (ValueError, TypeError):
                    errors.append(f"{adm_no}: bad value '{cell_val}' in {assessment['assessment_name']}.")
                    continue
                mm = float(assessment.get("max_marks", 100))
                if mv < 0 or mv > mm:
                    errors.append(f"{adm_no}: {assessment['assessment_name']} {mv} out of range (0-{mm}).")
                    continue
                ex = (db.table("formative_marks").select("id")
                        .eq("assessment_id", assessment["id"])
                        .eq("student_id", sid).execute().data or [])
                if ex:
                    db.table("formative_marks").update({
                        "marks_obtained": mv, "uploaded_by": user["id"],
                        "updated_at": datetime.now().isoformat()
                    }).eq("id", ex[0]["id"]).execute()
                else:
                    db.table("formative_marks").insert({
                        "assessment_id": assessment["id"],
                        "student_id": sid, "marks_obtained": mv,
                        "uploaded_by": user["id"]
                    }).execute()
                processed += 1

        wb.close()
        write_audit_log("import_formative_marks", target=f"class:{class_id},unit:{unit_id}")
        msg = f"Imported {processed} mark(s) successfully."
        if errors:
            msg += f" Errors ({len(errors)}): " + "; ".join(errors[:3])
            flash(msg, "warning")
        else:
            flash(msg, "success")
    except Exception as e:
        flash(f"Error processing file: {e}", "error")

    return redirect(redirect_url)


# ── Portfolio / Documents ───────────────────────────────────────────────────

@trainer_bp.route("/portfolio")
@trainer_required
def portfolio():
    """Trainer portfolio document management."""
    db = get_service_client()
    user = current_user()
    
    # Get trainer's assigned units with full unit details for the upload form
    assigned_unit_ids = _trainer_assigned_unit_ids(db)
    assigned_units = []
    if assigned_unit_ids:
        unit_rows = (db.table("units")
                      .select("id, name, code")
                      .in_("id", assigned_unit_ids)
                      .order("name")
                      .execute().data or [])
        # Wrap each row to match the template's {{ unit.units.id }} pattern
        assigned_units = [{"units": u} for u in unit_rows]

    # Get all documents uploaded by trainer
    documents = (db.table("trainer_documents")
                .select("*, units(name, code), classes(name)")
                .eq("trainer_id", user["id"])
                .order("created_at", desc=True)
                .execute().data or [])
    
    # Get filter parameters
    document_type = request.args.get("document_type", "").strip()
    year = request.args.get("year", str(datetime.now().year))
    term = request.args.get("term", "").strip()
    
    # Apply filters
    if document_type:
        documents = [d for d in documents if d["document_type"] == document_type]
    if year:
        documents = [d for d in documents if d.get("academic_year") == int(year)]
    if term:
        documents = [d for d in documents if d.get("term") == term]
    
    return render_template("trainer/portfolio.html",
                          documents=documents,
                          assigned_units=assigned_units,
                          document_type=document_type,
                          year=year,
                          term=term)


@trainer_bp.route("/portfolio/upload", methods=["POST"])
@trainer_required
def upload_document():
    """Upload a document to portfolio."""
    db = get_service_client()
    user = current_user()
    
    document_type = request.form.get("document_type")
    document_name = request.form.get("document_name")
    unit_id = request.form.get("unit_id")
    class_id = request.form.get("class_id")
    academic_year = request.form.get("academic_year")
    term = request.form.get("term")
    description = request.form.get("description", "")
    
    if not all([document_type, document_name, academic_year]):
        flash("Missing required fields.", "error")
        return redirect(url_for("trainer.portfolio"))
    
    if 'document_file' not in request.files:
        flash("No file uploaded.", "error")
        return redirect(url_for("trainer.portfolio"))
    
    file = request.files['document_file']
    if file.filename == '':
        flash("No file selected.", "error")
        return redirect(url_for("trainer.portfolio"))
    
    try:
        # Upload file to Supabase Storage
        import uuid
        file_extension = file.filename.split('.')[-1]
        unique_filename = f"{uuid.uuid4()}.{file_extension}"
        storage_path = f"trainer_documents/{user['id']}/{unique_filename}"
        
        # Read file once — used for both upload and size calculation
        file_data = file.read()
        
        # Upload to storage bucket using the service client (already configured)
        svc = get_service_client()
        svc.storage.from_("assessment-scripts").upload(
            path=storage_path,
            file=file_data,
            file_options={"content-type": file.content_type, "content-disposition": "inline"}
        )

        # Get public URL
        file_url = f"{os.getenv('SUPABASE_URL')}/storage/v1/object/public/assessment-scripts/{storage_path}"
        
        # Save document record
        db.table("trainer_documents").insert({
            "trainer_id": user["id"],
            "unit_id": unit_id if unit_id else None,
            "class_id": class_id if class_id else None,
            "document_type": document_type,
            "document_name": document_name,
            "file_url": file_url,
            "file_name": file.filename,
            "file_size": len(file_data),
            "file_type": file.content_type,
            "description": description,
            "academic_year": int(academic_year),
            "term": term
        }).execute()
        
        write_audit_log("upload_trainer_document", target=f"type:{document_type}")
        flash("Document uploaded successfully.", "success")
    except Exception as e:
        flash(f"Error uploading document: {e}", "error")
    
    return redirect(url_for("trainer.portfolio"))


@trainer_bp.route("/portfolio/view/<document_id>")
@trainer_required
def view_document(document_id):
    db = get_service_client()
    user = current_user()
    result = db.table("trainer_documents").select("*").eq("id", document_id).eq("trainer_id", user["id"]).execute()
    doc = result.data[0] if result.data else None
    if not doc:
        abort(404)
    file_url = doc.get("file_url", "")
    bucket = "assessment-scripts" if "/assessment-scripts/" in file_url else "documents"
    split_key = f"/{bucket}/"
    storage_path = file_url.split(split_key)[-1] if split_key in file_url else None
    if not storage_path:
        return redirect(file_url)
    try:
        raw = bytes(db.storage.from_(bucket).download(storage_path))
    except Exception:
        abort(404)
    ct = doc.get("file_type") or "application/octet-stream"
    fn = doc.get("file_name") or "document"
    resp = make_response(raw)
    resp.headers["Content-Type"] = ct
    resp.headers["Content-Disposition"] = f"inline; filename=\"{fn}\""
    return resp


@trainer_bp.route("/portfolio/delete/<document_id>", methods=["POST"])
@trainer_required
def delete_document(document_id):
    """Delete a document from portfolio."""
    db = get_service_client()
    user = current_user()
    
    try:
        # Get document
        document = db.table("trainer_documents").select("*").eq("id", document_id).single().execute().data
        
        if not document or document["trainer_id"] != user["id"]:
            abort(403)
        
        # Delete from storage
        _furl = document.get("file_url", "")
        storage_path = _furl.split("/assessment-scripts/")[-1] if "/assessment-scripts/" in _furl else _furl.split("documents/")[-1]
        svc = get_service_client()
        svc.storage.from_("assessment-scripts").remove([storage_path])
        
        # Delete record
        db.table("trainer_documents").delete().eq("id", document_id).execute()
        
        write_audit_log("delete_trainer_document", target=f"document:{document_id}")
        flash("Document deleted successfully.", "success")
    except Exception as e:
        flash(f"Error deleting document: {e}", "error")
    
    return redirect(url_for("trainer.portfolio"))
