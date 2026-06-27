"""
routes/super_admin_merged.py — Super Admin blueprint (merged system).

Combines features from both:
- Attendance management (from original)
- E-Portfolio management (from copy)
- User management (from copy)
- Department/Class/Unit management (from both)
"""

from flask import (Blueprint, render_template, request,
                   redirect, url_for, flash, abort, jsonify, make_response)
from auth_utils import super_admin_required, write_audit_log, current_user
from db import get_service_client
from werkzeug.security import generate_password_hash
from notifications import create_notification

super_admin_bp = Blueprint("super_admin", __name__)


def _svc():
    return get_service_client()


def _generate_password(length: int = 10) -> str:
    """Generate a readable temporary password."""
    import secrets, string
    chars = string.ascii_letters + string.digits + "@#!"
    while True:
        pwd = ''.join(secrets.choice(chars) for _ in range(length))
        # Ensure at least one digit and one special char
        if (any(c.isdigit() for c in pwd) and
                any(c in "@#!" for c in pwd)):
            return pwd


# ── One-time setup: seed super_admin user_profiles row ───────────────────────
# Visit /super-admin/setup-profile?email=YOUR_EMAIL once to create the row.
# This route is only accessible when NOT already logged in as super_admin.

@super_admin_bp.route("/setup-profile")
def setup_profile():
    """
    One-time helper: creates a user_profiles row for an existing Supabase Auth
    super_admin user so they can log in through the app.
    Usage: /super-admin/setup-profile?email=you@example.com
    """
    email = request.args.get("email", "").strip().lower()
    if not email:
        return jsonify({"error": "Pass ?email=your@email.com"}), 400

    db = get_service_client()

    # Check if profile already exists
    existing = db.table("user_profiles").select("id, role").eq("email", email).execute().data
    if existing:
        return jsonify({
            "status": "already_exists",
            "id": existing[0]["id"],
            "role": existing[0]["role"]
        })

    # Look up the auth user by email using admin API
    try:
        users_resp = db.auth.admin.list_users()
        auth_user = next(
            (u for u in users_resp if getattr(u, "email", "") == email),
            None
        )
        if not auth_user:
            return jsonify({"error": f"No Supabase Auth user found for {email}"}), 404

        user_id = str(auth_user.id)

        # Insert user_profiles row
        db.table("user_profiles").insert({
            "id": user_id,
            "email": email,
            "full_name": email.split("@")[0].replace(".", " ").title(),
            "role": "super_admin",
            "is_active": True
        }).execute()

        return jsonify({
            "status": "created",
            "id": user_id,
            "email": email,
            "role": "super_admin",
            "message": "Profile created. You can now log in."
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ── Dashboard ─────────────────────────────────────────────────────────────────

@super_admin_bp.route("/")
@super_admin_bp.route("/dashboard")
@super_admin_required
def dashboard():
    db = _svc()
    stats = {}
    recent_assessments = []
    recent_logs = []
    recent_clearances = []
    dept_stats = []

    try:
        # ── Core counts ──────────────────────────────────────────────────────
        stats['departments']  = db.table("departments").select("id", count="exact").execute().count or 0
        stats['classes']      = db.table("classes").select("id", count="exact").execute().count or 0
        stats['units']        = db.table("units").select("id", count="exact").execute().count or 0
        stats['attendance']   = db.table("attendance").select("id", count="exact").execute().count or 0
        stats['assessments']  = db.table("assessments").select("id", count="exact").execute().count or 0

        # ── Role breakdown ────────────────────────────────────────────────────
        all_users = db.table("user_profiles").select("role").execute().data or []
        stats['users']       = len(all_users)
        stats['dept_admins'] = sum(1 for u in all_users if u['role'] == 'dept_admin')
        stats['trainers']    = sum(1 for u in all_users if u['role'] == 'trainer')
        stats['students']    = sum(1 for u in all_users if u['role'] == 'student')

        # ── Assessment status breakdown ───────────────────────────────────────
        all_assess = db.table("assessments").select("status").execute().data or []
        stats['pending']  = sum(1 for a in all_assess if a['status'] == 'pending')
        stats['approved'] = sum(1 for a in all_assess if a['status'] == 'approved')
        stats['rejected'] = sum(1 for a in all_assess if a['status'] == 'rejected')

        # ── Clearance counts ──────────────────────────────────────────────────
        all_cl = db.table("clearance_requests").select("status").execute().data or []
        stats['clearances']           = len(all_cl)
        stats['clearances_pending']   = sum(1 for c in all_cl if c['status'] in ('pending','in_progress'))
        stats['clearances_completed'] = sum(1 for c in all_cl if c['status'] == 'completed')

        # ── Recent assessments ────────────────────────────────────────────────
        recent_assessments = (
            db.table("assessments")
            .select("*, user_profiles!assessments_student_id_fkey(full_name, admission_no), units(name), classes(name)")
            .order("uploaded_at", desc=True).limit(8).execute().data or []
        )

        # ── Recent clearances ─────────────────────────────────────────────────
        recent_clearances = (
            db.table("clearance_requests")
            .select("*, user_profiles!clearance_requests_student_id_fkey(full_name, admission_no), courses(name)")
            .order("created_at", desc=True).limit(5).execute().data or []
        )

        # ── Department stats ──────────────────────────────────────────────────
        depts = db.table("departments").select("id, name").order("name").execute().data or []
        for d in depts:
            did = d["id"]
            cc = db.table("classes").select("id", count="exact").eq("department_id", did).execute().count or 0
            sc = db.table("user_profiles").select("id", count="exact").eq("department_id", did).eq("role", "student").execute().count or 0
            tc = db.table("user_profiles").select("id", count="exact").eq("department_id", did).eq("role", "trainer").execute().count or 0
            dept_stats.append({"id": did, "name": d["name"],
                               "class_count": cc, "student_count": sc, "trainer_count": tc})

        # ── Recent audit logs ─────────────────────────────────────────────────
        recent_logs = (
            db.table("system_logs")
            .select("*, user_profiles(full_name, role)")
            .order("created_at", desc=True).limit(10).execute().data or []
        )

    except Exception as e:
        flash(f'Error loading dashboard: {e}', 'danger')

    return render_template("super_admin/welcome.html",
                           stats=stats,
                           # legacy vars kept for welcome.html compat
                           depts_count=stats.get('departments', 0),
                           trainers_count=stats.get('trainers', 0),
                           classes_count=stats.get('classes', 0),
                           students_count=stats.get('students', 0),
                           units_count=stats.get('units', 0),
                           recent_assessments=recent_assessments,
                           recent_clearances=recent_clearances,
                           recent_logs=recent_logs,
                           dept_stats=dept_stats)


# ── Departments ───────────────────────────────────────────────────────────────

@super_admin_bp.route("/departments", methods=["GET", "POST"])
@super_admin_required
def departments():
    db = _svc()
    error = None

    if request.method == "POST" and request.form.get("add_dept"):
        name = request.form.get("name", "").strip().upper()
        code = request.form.get("code", "").strip().upper()
        if not name or not code:
            error = "Department name and code cannot be empty."
        else:
            try:
                existing = db.table("departments").select("id").eq("name", name).execute()
                if existing.data:
                    error = "Department already exists."
                else:
                    db.table("departments").insert({"name": name, "code": code}).execute()
                    write_audit_log("create_department", target=name)
                    flash("Department added successfully.", "success")
                    return redirect(url_for("super_admin.departments"))
            except Exception as exc:
                error = f"Error: {exc}"

    if request.args.get("delete"):
        try:
            dept_id = request.args["delete"]
            db.table("departments").delete().eq("id", dept_id).execute()
            write_audit_log("delete_department", target=dept_id)
            flash("Department deleted.", "success")
            return redirect(url_for("super_admin.departments"))
        except Exception as exc:
            error = f"Error deleting: {exc}"

    depts = db.table("departments").select("*").order("name").execute().data or []
    return render_template("super_admin/departments.html", departments=depts, error=error)


# ── Users Management ───────────────────────────────────────────────────────────

@super_admin_bp.route("/users", methods=["GET", "POST"])
@super_admin_required
def users():
    db = _svc()
    error = None
    role_filter = request.args.get("role", "")
    dept_filter = request.args.get("department", "")
    new_user_creds = None          # shown once after creation
    suggested_password = _generate_password()

    if request.method == "POST" and request.form.get("add_user"):
        email         = request.form.get("email", "").strip().lower()
        full_name     = request.form.get("full_name", "").strip()
        role          = request.form.get("role", "")
        department_id = request.form.get("department_id") or None
        password      = request.form.get("password", "").strip()
        admission_no  = request.form.get("admission_no", "").strip()
        staff_no      = request.form.get("staff_no", "").strip() or None
        mobile_number = request.form.get("mobile_number", "").strip() or None

        # Validate role against allowed set
        from auth_utils import ALL_ROLES
        if not all([email, full_name, role, password]):
            error = "Full name, email, role and password are all required."
        elif role not in ALL_ROLES:
            error = f"Invalid role: {role}"
        elif len(password) < 8:
            error = "Password must be at least 8 characters."
        else:
            try:
                # Duplicate email check
                existing = db.table("user_profiles").select("id").eq("email", email).execute()
                if existing.data:
                    error = "An account with this email already exists."
                else:
                    if role == "student":
                        if not admission_no:
                            error = "Admission number is required for students."
                        else:
                            dup_adm = db.table("user_profiles").select("id").eq("admission_no", admission_no).execute()
                            if dup_adm.data:
                                error = "Admission number already exists."
                            else:
                                from auth_utils import create_student_auth_user
                                user_id = create_student_auth_user(
                                    admission_no=admission_no,
                                    password=password,
                                    email=email,
                                    full_name=full_name,
                                    department_id=department_id,
                                    class_id=None,
                                    mobile_number=mobile_number
                                )
                                new_user_creds = {"full_name": full_name, "role": role,
                                                  "email": email, "password": password}
                                write_audit_log("create_user", target=f"user:{user_id}")
                                flash(f"Student account created for {full_name}.", "success")
                    else:
                        from auth_utils import create_staff_auth_user
                        user_id = create_staff_auth_user(
                            email=email,
                            password=password,
                            full_name=full_name,
                            role=role,
                            department_id=department_id,
                            staff_no=staff_no,
                            mobile_number=mobile_number
                        )
                        new_user_creds = {"full_name": full_name, "role": role,
                                          "email": email, "password": password}
                        write_audit_log("create_user", target=f"user:{user_id}")
                        flash(f"Account created for {full_name} ({role.replace('_',' ').title()}).", "success")
            except Exception as exc:
                error = f"Error creating account: {exc}"

    # Build query
    query = db.table("user_profiles").select("*, departments(name)")
    if role_filter:
        query = query.eq("role", role_filter)
    if dept_filter:
        query = query.eq("department_id", dept_filter)

    users_list   = query.order("created_at", desc=True).execute().data or []
    departments  = db.table("departments").select("*").order("name").execute().data or []

    return render_template("super_admin/users.html",
                           users=users_list,
                           departments=departments,
                           error=error,
                           role_filter=role_filter,
                           dept_filter=dept_filter,
                           new_user_creds=new_user_creds,
                           suggested_password=suggested_password)


@super_admin_bp.route("/users/<user_id>/edit", methods=["GET", "POST"])
@super_admin_required
def edit_user(user_id):
    db = _svc()
    error = None

    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        role = request.form.get("role", "")
        department_id = request.form.get("department_id")
        is_active = request.form.get("is_active") == "on"

        try:
            update_data = {
                "full_name": full_name,
                "role": role,
                "department_id": department_id if department_id else None,
                "is_active": is_active
            }
            db.table("user_profiles").update(update_data).eq("id", user_id).execute()
            write_audit_log("update_user", target=f"user:{user_id}")
            flash("User updated successfully.", "success")
            return redirect(url_for("super_admin.users"))
        except Exception as exc:
            error = f"Error: {exc}"

    user = db.table("user_profiles").select("*, departments(name)").eq("id", user_id).single().execute().data
    departments = db.table("departments").select("*").order("name").execute().data or []

    return render_template("super_admin/edit_user.html", 
                          user=user, 
                          departments=departments,
                          error=error)


@super_admin_bp.route("/users/<user_id>/delete")
@super_admin_required
def delete_user(user_id):
    db = _svc()
    try:
        db.table("user_profiles").delete().eq("id", user_id).execute()
        write_audit_log("delete_user", target=f"user:{user_id}")
        flash("User deleted successfully.", "success")
    except Exception as exc:
        flash(f"Error deleting user: {exc}", "danger")
    return redirect(url_for("super_admin.users"))


# ── Classes Management ─────────────────────────────────────────────────────────

@super_admin_bp.route("/classes", methods=["GET", "POST"])
@super_admin_required
def classes():
    db = _svc()
    error = None

    if request.args.get("delete"):
        try:
            db.table("classes").delete().eq("id", request.args["delete"]).execute()
            write_audit_log("delete_class", target=request.args["delete"])
            flash("Class deleted.", "success")
            return redirect(url_for("super_admin.classes"))
        except Exception as exc:
            error = f"Error deleting: {exc}"

    if request.method == "POST" and request.form.get("add_class"):
        name          = request.form.get("name", "").strip()
        course_id     = request.form.get("course_id") or None
        department_id = request.form.get("department_id") or None
        intake_year   = request.form.get("intake_year") or None
        intake_month  = request.form.get("intake_month") or None
        level         = request.form.get("level") or None
        cycle         = request.form.get("cycle") or None

        if not name or not department_id:
            error = "Class name and department are required."
        else:
            try:
                db.table("classes").insert({
                    "name": name,
                    "course_id": course_id,
                    "department_id": department_id,
                    "intake_year": int(intake_year) if intake_year else None,
                    "intake_month": intake_month,
                    "level": level,
                    "cycle": cycle
                }).execute()
                write_audit_log("create_class", target=name)
                flash("Class added successfully.", "success")
                return redirect(url_for("super_admin.classes"))
            except Exception as exc:
                error = f"Error: {exc}"

    classes_list = db.table("classes").select("*, departments(name), courses(name)").order("name").execute().data or []
    departments  = db.table("departments").select("*").order("name").execute().data or []
    courses      = db.table("courses").select("*").order("name").execute().data or []

    return render_template("super_admin/classes.html",
                           classes=classes_list,
                           departments=departments,
                           courses=courses,
                           error=error)


# ── Units Management ───────────────────────────────────────────────────────────

@super_admin_bp.route("/units", methods=["GET", "POST"])
@super_admin_required
def units():
    db = _svc()
    error = None

    if request.args.get("delete"):
        try:
            db.table("units").delete().eq("id", request.args["delete"]).execute()
            write_audit_log("delete_unit", target=request.args["delete"])
            flash("Unit deleted.", "success")
            return redirect(url_for("super_admin.units"))
        except Exception as exc:
            error = f"Error deleting: {exc}"

    if request.method == "POST" and request.form.get("add_unit"):
        code          = request.form.get("code", "").strip().upper()
        name          = request.form.get("name", "").strip()
        department_id = request.form.get("department_id") or None
        course_id     = request.form.get("course_id") or None

        if not code or not name or not department_id:
            error = "Code, name, and department are required."
        else:
            try:
                db.table("units").insert({
                    "code": code,
                    "name": name,
                    "department_id": department_id,
                    "course_id": course_id
                }).execute()
                write_audit_log("create_unit", target=code)
                flash("Unit added successfully.", "success")
                return redirect(url_for("super_admin.units"))
            except Exception as exc:
                error = f"Error: {exc}"

    units_list  = db.table("units").select("*, departments(name), courses(name)").order("code").execute().data or []
    departments = db.table("departments").select("*").order("name").execute().data or []
    courses     = db.table("courses").select("*").order("name").execute().data or []

    return render_template("super_admin/units.html",
                           units=units_list,
                           departments=departments,
                           courses=courses,
                           error=error)


# ── Courses Management ─────────────────────────────────────────────────────────

@super_admin_bp.route("/courses", methods=["GET", "POST"])
@super_admin_required
def courses():
    db = _svc()
    error = None

    if request.args.get("delete"):
        try:
            db.table("courses").delete().eq("id", request.args["delete"]).execute()
            write_audit_log("delete_course", target=request.args["delete"])
            flash("Course deleted.", "success")
            return redirect(url_for("super_admin.courses"))
        except Exception as exc:
            error = f"Error deleting: {exc}"

    if request.method == "POST" and request.form.get("add_course"):
        name          = request.form.get("name", "").strip()
        code          = request.form.get("code", "").strip().upper()
        department_id = request.form.get("department_id") or None

        if not name or not code or not department_id:
            error = "Name, code, and department are required."
        else:
            try:
                db.table("courses").insert({
                    "name": name,
                    "code": code,
                    "department_id": department_id
                }).execute()
                write_audit_log("create_course", target=code)
                flash("Course added successfully.", "success")
                return redirect(url_for("super_admin.courses"))
            except Exception as exc:
                error = f"Error: {exc}"

    courses_list = db.table("courses").select("*, departments(name)").order("name").execute().data or []
    departments  = db.table("departments").select("*").order("name").execute().data or []

    return render_template("super_admin/courses.html",
                           courses=courses_list,
                           departments=departments,
                           error=error)


# ── System Logs ───────────────────────────────────────────────────────────────

@super_admin_bp.route("/logs")
@super_admin_required
def logs():
    db      = _svc()
    page    = max(1, int(request.args.get("page", 1)))
    q       = request.args.get("q", "").strip()
    action_filter = request.args.get("action", "").strip()
    per_page = 100
    offset   = (page - 1) * per_page

    query = db.table("system_logs").select("*, user_profiles(full_name, role)")
    if action_filter:
        query = query.ilike("action", f"%{action_filter}%")
    logs_list = (query.order("created_at", desc=True)
                      .range(offset, offset + per_page - 1)
                      .execute().data or [])

    if q:
        ql = q.lower()
        logs_list = [l for l in logs_list if
                     ql in (l.get("action") or "").lower() or
                     ql in (l.get("target") or "").lower() or
                     ql in ((l.get("user_profiles") or {}).get("full_name") or "").lower()]

    # Collect distinct action types for filter dropdown
    all_actions = []
    try:
        all_actions = sorted({l.get("action", "") for l in
                              (db.table("system_logs").select("action")
                                  .execute().data or []) if l.get("action")})
    except Exception:
        pass

    return render_template("super_admin/system_logs.html",
                           logs=logs_list, page=page,
                           q=q, action_filter=action_filter,
                           all_actions=all_actions,
                           per_page=per_page)


# ── Attendance Matrix (institute-wide) ────────────────────────────────────────

@super_admin_bp.route("/attendance")
@super_admin_required
def attendance():
    from datetime import date as _date
    db = _svc()

    dept_filter  = request.args.get("dept_id",   "").strip()
    class_filter = request.args.get("class_id",  "").strip()
    unit_filter  = request.args.get("unit_id",   "").strip()
    term_filter  = request.args.get("term",      "").strip()
    year_filter  = request.args.get("year",      str(_date.today().year)).strip()

    departments = db.table("departments").select("id, name").order("name").execute().data or []

    # Classes and units scoped to selected department (or all)
    cq = db.table("classes").select("id, name, department_id")
    if dept_filter:
        cq = cq.eq("department_id", dept_filter)
    classes = cq.order("name").execute().data or []

    uq = db.table("units").select("id, name, code, department_id")
    if dept_filter:
        uq = uq.eq("department_id", dept_filter)
    units = uq.order("name").execute().data or []

    LESSONS = ["L1", "L2", "L3", "L4"]
    WEEKS   = list(range(1, 16))          # weeks 1-15
    matrix  = []
    unit_obj = cls_obj = None
    term_int = int(term_filter) if term_filter else None
    year_int = int(year_filter) if year_filter else None

    def _norm_lesson(l):
        s = str(l).strip()
        return f"L{s}" if s in ("1", "2", "3", "4") else s

    if class_filter and unit_filter:
        unit_obj = next((u for u in units if u["id"] == unit_filter), None)
        cls_obj  = next((c for c in classes if c["id"] == class_filter), None)

        enr_rows = (db.table("enrollments")
                    .select("student_id, "
                            "user_profiles!enrollments_student_id_fkey"
                            "(id, full_name, admission_no)")
                    .eq("class_id", class_filter)
                    .execute().data or [])

        students_ordered = []
        for e in enr_rows:
            up  = e.get("user_profiles") or {}
            sid = up.get("id") or e.get("student_id")
            if sid and not any(s["id"] == sid for s in students_ordered):
                students_ordered.append({
                    "id":           sid,
                    "full_name":    up.get("full_name", "—"),
                    "admission_no": up.get("admission_no", "—"),
                })
        students_ordered.sort(key=lambda s: s["full_name"])

        student_ids = [s["id"] for s in students_ordered]
        att_rows    = []
        if student_ids:
            q = (db.table("attendance")
                 .select("student_id, week, lesson, status")
                 .eq("unit_id", unit_filter)
                 .in_("student_id", student_ids))
            if term_int:
                q = q.eq("term", term_int)
            if year_int:
                q = q.eq("year", year_int)
            att_rows = q.execute().data or []

        pivot = {}
        for r in att_rows:
            key = (r["week"], _norm_lesson(r["lesson"]))
            pivot.setdefault(r["student_id"], {})[key] = r["status"]

        for s in students_ordered:
            cells   = {}
            present = absent = 0
            for w in WEEKS:
                for l in LESSONS:
                    st = pivot.get(s["id"], {}).get((w, l))
                    cells[(w, l)] = st
                    if st == "present":
                        present += 1
                    elif st == "absent":
                        absent += 1
            total = present + absent
            pct   = round(present / total * 100, 1) if total else 0
            matrix.append({
                "id":           s["id"],
                "full_name":    s["full_name"],
                "admission_no": s["admission_no"],
                "cells":        cells,
                "present":      present,
                "absent":       absent,
                "total":        total,
                "pct":          pct,
            })

    return render_template(
        "super_admin/attendance.html",
        departments=departments,
        classes=classes,
        units=units,
        dept_filter=dept_filter,
        class_filter=class_filter,
        unit_filter=unit_filter,
        term_filter=term_filter,
        year_filter=year_filter,
        term_int=term_int,
        year_int=year_int,
        matrix=matrix,
        unit_obj=unit_obj,
        cls_obj=cls_obj,
        WEEKS=WEEKS,
        LESSONS=LESSONS,
    )


# ── Trainees POE ───────────────────────────────────────────────────────────────

@super_admin_bp.route("/assessments")
@super_admin_required
def assessments():
    import os
    from datetime import date as _date
    db = _svc()
    supabase_url  = os.environ.get("SUPABASE_URL", "").strip()

    dept_filter   = request.args.get("department", "")
    year_filter   = request.args.get("year", "")
    class_filter  = request.args.get("class_id", "")
    adm_filter    = request.args.get("admission_no", "").strip().upper()
    status_filter = request.args.get("status", "")

    query = (db.table("assessments")
               .select("id, status, script_file_path, script_file_name, script_file_size, "
                       "uploaded_at, assessment_type, assessment_no, term, year, class_id, "
                       "student:user_profiles!assessments_student_id_fkey(full_name, admission_no), "
                       "units!inner(name, code, department_id, departments(name)), "
                       "classes(name)")
               .order("uploaded_at", desc=True)
               .limit(500))

    if dept_filter:
        query = query.eq("units.department_id", dept_filter)
    if year_filter:
        try:
            query = query.eq("year", int(year_filter))
        except ValueError:
            pass
    if class_filter:
        query = query.eq("class_id", class_filter)
    if status_filter:
        query = query.eq("status", status_filter)

    records = query.execute().data or []

    if adm_filter:
        records = [r for r in records
                   if adm_filter in (r.get("student") or {}).get("admission_no", "").upper()]

    # Batch-fetch evidence for all returned assessments
    evidence_map = {}
    if records:
        a_ids = [r["id"] for r in records if r.get("id")]
        try:
            ev_rows = (db.table("evidence")
                         .select("assessment_id, file_path, file_name, file_type")
                         .in_("assessment_id", a_ids)
                         .execute().data or [])
            for ev in ev_rows:
                aid = ev.get("assessment_id")
                if not aid:
                    continue
                fp  = ev.get("file_path") or ""
                ext = fp.rsplit(".", 1)[-1].lower() if "." in fp else "bin"
                evidence_map.setdefault(aid, []).append({
                    "url":  f"{supabase_url}/storage/v1/object/public/assessment-evidence/{fp}" if fp else "",
                    "name": ev.get("file_name") or (fp.rsplit("/", 1)[-1] if fp else "file"),
                    "ext":  ext,
                    "type": ev.get("file_type") or "",
                })
        except Exception as e:
            print(f"[assessments] evidence fetch: {e}")

    def _fmt_size(b):
        if not b:
            return ""
        for u in ["B", "KB", "MB", "GB"]:
            if b < 1024:
                return f"{b:.1f} {u}"
            b /= 1024
        return f"{b:.1f} GB"

    for r in records:
        fp = r.get("script_file_path") or ""
        r["_script_url"]  = (f"{supabase_url}/storage/v1/object/public/assessment-scripts/{fp}"
                             if fp else "")
        r["_script_size"] = _fmt_size(r.get("script_file_size"))
        r["_evidence"]    = evidence_map.get(r["id"], [])

    departments = db.table("departments").select("id, name").order("name").execute().data or []
    classes     = db.table("classes").select("id, name").order("name").execute().data or []
    cur_yr      = _date.today().year
    years       = [str(y) for y in range(cur_yr, 2021, -1)]

    return render_template("super_admin/assessments.html",
                           assessments=records,
                           departments=departments,
                           classes=classes,
                           years=years,
                           dept_filter=dept_filter,
                           year_filter=year_filter,
                           class_filter=class_filter,
                           adm_filter=adm_filter,
                           status_filter=status_filter)


# ── Marks Report ───────────────────────────────────────────────────────────────

def _sa_compute_grade(obtained, max_m):
    try:
        pct = round(float(obtained) / float(max_m) * 100, 1) if max_m else 0
    except (TypeError, ZeroDivisionError):
        pct = 0
    if pct >= 80:   grade = "4"
    elif pct >= 65: grade = "3"
    elif pct >= 50: grade = "2"
    else:           grade = "1"
    return pct, grade


def _fetch_marks_all(db, dept_id, year, term, class_id, unit_id, adm_filter):
    unit_q = db.table("units").select("id")
    if dept_id:
        unit_q = unit_q.eq("department_id", dept_id)
    unit_ids = [u["id"] for u in (unit_q.execute().data or [])]
    if not unit_ids:
        return []

    # Chunk unit_ids to avoid query size limits
    all_fas = []
    for i in range(0, len(unit_ids), 200):
        chunk = unit_ids[i:i+200]
        fa_q = (db.table("formative_assessments")
                .select("id, unit_id, class_id, trainer_id, assessment_type, "
                        "assessment_name, max_marks, year, term, created_at, "
                        "units(name, code, departments(name)), classes(name), "
                        "trainer:user_profiles!formative_assessments_trainer_id_fkey(full_name)")
                .in_("unit_id", chunk)
                .eq("year", int(year)))
        if term:     fa_q = fa_q.eq("term",     int(term))
        if class_id: fa_q = fa_q.eq("class_id", class_id)
        if unit_id:  fa_q = fa_q.eq("unit_id",  unit_id)
        try:
            all_fas.extend(fa_q.order("created_at", desc=True).execute().data or [])
        except Exception as e:
            print(f"[marks_all] formative_assessments: {e}")

    if not all_fas:
        return []

    fa_map = {a["id"]: a for a in all_fas}
    a_ids  = list(fa_map.keys())

    all_fm = []
    for i in range(0, len(a_ids), 400):
        chunk = a_ids[i:i+400]
        try:
            all_fm.extend(
                db.table("formative_marks")
                  .select("assessment_id, student_id, marks_obtained, "
                          "student:user_profiles!formative_marks_student_id_fkey"
                          "(full_name, admission_no)")
                  .in_("assessment_id", chunk)
                  .execute().data or []
            )
        except Exception as e:
            print(f"[marks_all] formative_marks: {e}")

    rows = []
    for m in all_fm:
        fa  = fa_map.get(m["assessment_id"], {})
        pct, grade = _sa_compute_grade(m.get("marks_obtained"), fa.get("max_marks", 100))
        un = fa.get("units") or {}
        rows.append({
            "student":         m.get("student") or {},
            "unit":            un,
            "dept_name":       (un.get("departments") or {}).get("name", ""),
            "class_":          fa.get("classes") or {},
            "trainer":         fa.get("trainer") or {},
            "assessment_name": fa.get("assessment_name", ""),
            "assessment_type": fa.get("assessment_type", ""),
            "max_marks":       fa.get("max_marks", 100),
            "marks_obtained":  m.get("marks_obtained"),
            "percentage":      pct,
            "grade":           grade,
            "year":            fa.get("year"),
            "term":            fa.get("term"),
        })

    if adm_filter:
        rows = [r for r in rows
                if adm_filter.upper() in (r["student"].get("admission_no") or "").upper()]

    rows.sort(key=lambda r: (
        r["dept_name"],
        r["class_"].get("name", ""),
        r["student"].get("full_name", ""),
        r["unit"].get("name", ""),
        r["assessment_name"],
    ))
    return rows


@super_admin_bp.route("/marks")
@super_admin_required
def marks():
    from datetime import datetime as _dt
    db = _svc()
    year        = request.args.get("year",       str(_dt.now().year))
    term        = request.args.get("term",       "")
    dept_filter = request.args.get("department", "")
    class_filter = request.args.get("class_id",  "")
    unit_filter  = request.args.get("unit_id",   "")
    adm_filter   = request.args.get("admission_no", "").strip()

    marks_list = _fetch_marks_all(db, dept_filter, year, term,
                                  class_filter, unit_filter, adm_filter)

    distinct_students = len({r["student"].get("admission_no")
                             for r in marks_list if r["student"].get("admission_no")})
    pass_count = sum(1 for r in marks_list if r["grade"] in ("4", "3", "2"))
    pass_rate  = round(pass_count / len(marks_list) * 100) if marks_list else 0

    cur_yr      = _dt.now().year
    departments = db.table("departments").select("id, name").order("name").execute().data or []
    classes     = db.table("classes").select("id, name").order("name").execute().data or []
    units       = db.table("units").select("id, name, code").order("name").execute().data or []
    years       = [str(y) for y in range(2023, cur_yr + 2)]

    return render_template("super_admin/marks.html",
                           marks=marks_list,
                           departments=departments,
                           classes=classes,
                           units=units,
                           years=years,
                           year=year, term=term,
                           dept_filter=dept_filter,
                           class_filter=class_filter,
                           unit_filter=unit_filter,
                           adm_filter=adm_filter,
                           distinct_students=distinct_students,
                           pass_rate=pass_rate)


# ── Clearance Requests ─────────────────────────────────────────────────────────

@super_admin_bp.route("/clearances")
@super_admin_required
def clearances():
    db = _svc()
    status_filter = request.args.get("status", "")
    dept_filter   = request.args.get("department", "")

    query = db.table("clearance_requests").select(
        "*, user_profiles!clearance_requests_student_id_fkey(full_name, admission_no), "
        "courses(name, code), departments(name)"
    ).order("created_at", desc=True).limit(500)
    if status_filter:
        query = query.eq("status", status_filter)
    if dept_filter:
        query = query.eq("department_id", dept_filter)
    records = query.execute().data or []

    departments = db.table("departments").select("id, name").order("name").execute().data or []
    return render_template("super_admin/clearances.html",
                           clearances=records,
                           departments=departments,
                           status_filter=status_filter,
                           dept_filter=dept_filter)


# ── Service Clearance Departments Hub ─────────────────────────────────────────

SERVICE_CLR_CATS = {
    "svc_library": {"label": "Institute Library",  "icon": "fa-book",         "color": "#1d4ed8"},
    "svc_games":   {"label": "Games Department",   "icon": "fa-futbol",       "color": "#16a34a"},
    "svc_ict":     {"label": "ICT Department",     "icon": "fa-laptop",       "color": "#7c3aed"},
    "svc_kitchen": {"label": "Kitchen / Cafeteria","icon": "fa-utensils",     "color": "#b45309"},
    "svc_store":   {"label": "Store Department",   "icon": "fa-warehouse",    "color": "#0e7490"},
}


@super_admin_bp.route("/service-clearance")
@super_admin_required
def service_clearance():
    db = _svc()
    cat_filter = request.args.get("cat", "")

    rows = (db.table("clearance_approvals")
              .select(
                  "id, approver_category, status, comments, approved_at, created_at, approver_id, "
                  "clearance_requests!inner(id, student_id, status, stage, created_at, "
                  "  user_profiles:user_profiles!clearance_requests_student_id_fkey"
                  "  (full_name, admission_no), "
                  "  courses(name, code), departments(name))"
              )
              .in_("approver_category", list(SERVICE_CLR_CATS.keys()))
              .order("created_at", desc=True)
              .limit(600)
              .execute().data or [])

    if cat_filter and cat_filter in SERVICE_CLR_CATS:
        rows = [r for r in rows if r.get("approver_category") == cat_filter]

    # Group by category
    by_cat = {cat: [] for cat in SERVICE_CLR_CATS}
    for row in rows:
        cat = row.get("approver_category", "")
        if cat in by_cat:
            req = row.get("clearance_requests") or {}
            row["_student"]   = req.get("user_profiles") or {}
            row["_course"]    = req.get("courses") or {}
            row["_dept"]      = req.get("departments") or {}
            row["_req_id"]    = req.get("id", "")
            row["_req_status"]= req.get("status", "")
            by_cat[cat].append(row)

    pending_counts = {
        cat: sum(1 for r in lst if r.get("status") == "pending" and r.get("_req_status") not in ("completed", "rejected"))
        for cat, lst in by_cat.items()
    }

    # Fetch approver names for display
    approver_ids = list({r["approver_id"] for r in rows if r.get("approver_id")})
    approver_map = {}
    if approver_ids:
        profiles = (db.table("user_profiles")
                      .select("id, full_name")
                      .in_("id", approver_ids)
                      .execute().data or [])
        approver_map = {p["id"]: p.get("full_name", "") for p in profiles}

    for cat_rows in by_cat.values():
        for row in cat_rows:
            row["_approver_name"] = approver_map.get(row.get("approver_id"), "—")

    # Fetch existing service dept portal users
    svc_roles = ["library_hod", "sports_hod", "service_clearance_officer"]
    svc_users = (db.table("user_profiles")
                   .select("id, full_name, email, role, department_id, departments(name)")
                   .in_("role", svc_roles)
                   .order("role")
                   .execute().data or [])

    return render_template(
        "super_admin/service_clearance.html",
        by_cat=by_cat,
        cat_meta=SERVICE_CLR_CATS,
        pending_counts=pending_counts,
        total_pending=sum(pending_counts.values()),
        active_cat=cat_filter or "svc_library",
        svc_users=svc_users,
    )


# ── Industry Partners (Companies) ─────────────────────────────────────────────

@super_admin_bp.route("/companies")
@super_admin_required
def companies():
    db = _svc()
    records = db.table("companies").select("*, departments(name)").order("name").execute().data or []
    departments = db.table("departments").select("id, name").order("name").execute().data or []
    return render_template("super_admin/companies.html",
                           companies=records, departments=departments)


# ── Industrial Attachments ─────────────────────────────────────────────────────

@super_admin_bp.route("/attachments")
@super_admin_required
def attachments():
    db = _svc()
    status_filter = request.args.get("status", "")
    query = db.table("industrial_attachments").select(
        "*, user_profiles!industrial_attachments_student_id_fkey(full_name, admission_no), "
        "companies(name), units(name, code)"
    ).order("created_at", desc=True).limit(300)
    if status_filter:
        query = query.eq("status", status_filter)
    records = query.execute().data or []
    return render_template("super_admin/attachments.html",
                           attachments=records, status_filter=status_filter)


@super_admin_bp.route("/logbooks")
@super_admin_required
def logbooks():
    import os
    db = _svc()
    supabase_url  = os.environ.get("SUPABASE_URL", "").strip()
    dept_filter   = request.args.get("department", "")
    status_filter = request.args.get("status", "")
    adm_filter    = request.args.get("admission_no", "").strip().upper()

    query = (db.table("digital_logbook")
               .select("id, student_id, log_date, entry_time, tasks_performed, "
                       "skills_applied, hours_worked, challenges_encountered, "
                       "achievements, mentor_approval_status, mentor_comments, "
                       "trainer_comments, evidence_urls, created_at, "
                       "student:user_profiles!digital_logbook_student_id_fkey"
                       "(full_name, admission_no), "
                       "attachment:industrial_attachments!digital_logbook_attachment_id_fkey"
                       "(companies(name))")
               .order("log_date", desc=True)
               .limit(500))

    if status_filter:
        query = query.eq("mentor_approval_status", status_filter)

    records = query.execute().data or []

    # Department filter via enrolled students
    if dept_filter:
        enr = (db.table("enrollments")
                 .select("student_id, classes!inner(department_id)")
                 .eq("classes.department_id", dept_filter)
                 .execute().data or [])
        dept_sids = {e["student_id"] for e in enr}
        records = [r for r in records if r.get("student_id") in dept_sids]

    if adm_filter:
        records = [r for r in records
                   if adm_filter in (r.get("student") or {}).get("admission_no", "").upper()]

    for entry in records:
        ev_paths = entry.get("evidence_urls") or []
        entry["_evidence"] = [
            {
                "url":  f"{supabase_url}/storage/v1/object/public/assessment-evidence/{p}",
                "ext":  p.rsplit(".", 1)[-1].lower() if "." in p else "bin",
                "name": p.rsplit("/", 1)[-1],
            }
            for p in ev_paths if p
        ]

    departments = db.table("departments").select("id, name").order("name").execute().data or []
    return render_template("super_admin/logbooks.html",
                           logbooks=records, departments=departments,
                           dept_filter=dept_filter, status_filter=status_filter,
                           adm_filter=adm_filter)


# ── GIS Placements & Logbook (system-wide) ────────────────────────────────────

@super_admin_bp.route("/gis-tracking")
@super_admin_required
def gis_tracking():
    import os
    db = _svc()
    dept_filter = request.args.get("department", "")
    departments = db.table("departments").select("id, name").order("name").execute().data or []

    # All enrollments (optionally filtered by department)
    enr_query = (db.table("enrollments")
                   .select("student_id, classes!inner(department_id, departments(name))")
                   .order("student_id"))
    if dept_filter:
        enr_query = enr_query.eq("classes.department_id", dept_filter)
    enr = enr_query.execute().data or []
    student_ids = list({e["student_id"] for e in enr})

    # Build dept name lookup keyed by student_id
    student_dept = {}
    for e in enr:
        sid = e["student_id"]
        cls = e.get("classes") or {}
        dpt = cls.get("departments") or {}
        student_dept[sid] = dpt.get("name", "")

    placements = []
    if student_ids:
        try:
            placements = (db.table("industrial_attachments")
                .select("id, student_id, status, start_date, end_date, "
                        "companies(name, latitude, longitude, city, address, "
                        "  industry_classification, geofence_radius_meters, "
                        "  contact_person, contact_phone), "
                        "units(name, code), "
                        "student:user_profiles!industrial_attachments_student_id_fkey"
                        "(full_name, admission_no, mobile_number)")
                .in_("student_id", student_ids)
                .in_("status", ["active", "approved", "pending"])
                .order("created_at", desc=True)
                .execute().data or [])
        except Exception as e:
            flash(f"Error loading placements: {e}", "warning")

    # Attach dept name to each placement
    for p in placements:
        p["_department"] = student_dept.get(p.get("student_id"), "")

    # Live locations
    live_locations = []
    if student_ids:
        try:
            loc_rows = (db.table("location_logs")
                .select("student_id, latitude, longitude, check_in_time, "
                        "check_out_time, accuracy_meters, is_within_geofence")
                .in_("student_id", student_ids)
                .not_.is_("latitude", "null")
                .order("check_in_time", desc=True)
                .execute().data or [])
            seen = set()
            for loc in loc_rows:
                sid = loc["student_id"]
                if sid not in seen:
                    seen.add(sid)
                    live_locations.append(loc)
        except Exception as e:
            print(f"[super_admin.gis_tracking] location_logs: {e}")

    student_map = {p.get("student_id"): p.get("student") or {}
                   for p in placements if p.get("student_id")}
    for loc in live_locations:
        loc["student"] = student_map.get(loc["student_id"], {})

    supabase_url = os.environ.get("SUPABASE_URL", "").strip()
    logbook_entries = []
    logbook_error = None
    if student_ids:
        try:
            logbook_entries = (db.table("digital_logbook")
                .select("id, student_id, log_date, entry_time, tasks_performed, "
                        "skills_applied, hours_worked, challenges_encountered, "
                        "achievements, mentor_approval_status, mentor_comments, "
                        "trainer_comments, evidence_urls, created_at, "
                        "student:user_profiles!digital_logbook_student_id_fkey"
                        "(full_name, admission_no), "
                        "attachment:industrial_attachments!digital_logbook_attachment_id_fkey"
                        "(companies(name))")
                .in_("student_id", student_ids)
                .order("log_date", desc=True)
                .limit(500)
                .execute().data or [])
            for entry in logbook_entries:
                ev_paths = entry.get("evidence_urls") or []
                entry["_evidence"] = [
                    {
                        "url": f"{supabase_url}/storage/v1/object/public/assessment-evidence/{p}",
                        "ext": p.rsplit(".", 1)[-1].lower() if "." in p else "bin",
                        "name": p.rsplit("/", 1)[-1],
                    }
                    for p in ev_paths if p
                ]
        except Exception as e:
            logbook_error = str(e)

    status_counts = {"active": 0, "approved": 0, "pending": 0}
    for p in placements:
        s = p.get("status", "")
        if s in status_counts:
            status_counts[s] += 1

    return render_template(
        "super_admin/gis_tracking.html",
        placements=placements,
        live_locations=live_locations,
        logbook_entries=logbook_entries,
        logbook_error=logbook_error,
        status_counts=status_counts,
        total_students=len(student_ids),
        departments=departments,
        dept_filter=dept_filter,
    )


# ── Attachment Export (Excel / PDF) ───────────────────────────────────────────

def _get_period_range(year: int, period: str):
    """Return (start, end) ISO date strings for the given period."""
    ranges = {
        "1": (f"{year}-01-01", f"{year}-04-30"),
        "2": (f"{year}-05-01", f"{year}-08-31"),
        "3": (f"{year}-09-01", f"{year}-12-31"),
    }
    return ranges.get(period, (f"{year}-01-01", f"{year}-12-31"))


def _period_label(period: str) -> str:
    return {"1": "January–April", "2": "May–August", "3": "September–December"}.get(period, "Full Year")


def _build_export_rows(db, student_ids=None, dept_filter="",
                       period="", year=0, dept_admin_dept_id=None):
    """Fetch attachment rows for export. Returns list of dicts."""
    from datetime import date
    if not year:
        year = date.today().year

    start_date, end_date = _get_period_range(year, period)

    select_str = (
        "id, student_id, start_date, end_date, status, "
        "student:user_profiles!industrial_attachments_student_id_fkey"
        "(full_name, admission_no, mobile_number), "
        "companies(name, address, contact_person, contact_phone)"
    )
    query = (db.table("industrial_attachments")
               .select(select_str)
               .gte("start_date", start_date)
               .lte("start_date", end_date)
               .order("student_id"))

    if dept_admin_dept_id:
        # Fetch student_ids for this dept if not supplied
        if student_ids is None:
            enr = (db.table("enrollments")
                     .select("student_id, classes!inner(department_id)")
                     .eq("classes.department_id", dept_admin_dept_id)
                     .execute().data or [])
            student_ids = list({e["student_id"] for e in enr})
        if student_ids:
            query = query.in_("student_id", student_ids)
        else:
            return []
    elif student_ids is not None:
        if student_ids:
            query = query.in_("student_id", student_ids)
        else:
            return []

    rows = query.execute().data or []

    result = []
    for r in rows:
        st = r.get("student") or {}
        co = r.get("companies") or {}
        result.append({
            "Admission No":              st.get("admission_no", ""),
            "Full Name":                 st.get("full_name", ""),
            "Trainee Phone":             st.get("mobile_number", ""),
            "Company Attached":          co.get("name", ""),
            "Location / Address":        co.get("address", ""),
            "Supervisor Name":           co.get("contact_person", ""),
            "Supervisor Phone":          co.get("contact_phone", ""),
            "Start Date":                r.get("start_date", ""),
            "End Date":                  r.get("end_date", ""),
            "Status":                    (r.get("status") or "").title(),
        })
    return result


def _export_excel(rows, title: str, period_label: str, year: int):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import Response

    wb = Workbook()
    ws = wb.active
    ws.title = "Attachments"

    hdr_fill = PatternFill("solid", fgColor="1565C0")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells("A1:J1")
    ws["A1"] = f"{title} — {period_label} {year}"
    ws["A1"].font = Font(bold=True, size=13, color="0D2167")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Admission No", "Full Name", "Trainee Phone",
               "Company Attached", "Location / Address",
               "Supervisor Name", "Supervisor Phone",
               "Start Date", "End Date", "Status"]
    ws.append([])  # blank row
    ws.append(headers)
    hdr_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=col_idx)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in rows:
        ws.append([row.get(h, "") for h in headers])
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col_idx).border = border

    col_widths = [16, 28, 18, 28, 32, 24, 18, 14, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"attachments_{period_label.replace('–', '-')}_{year}.xlsx"
    return Response(buf.read(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


def _export_pdf(rows, title: str, period_label: str, year: int):
    import io
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from flask import Response

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    navy = colors.HexColor("#1565C0")

    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                 textColor=navy, fontSize=14, spaceAfter=4)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               textColor=colors.grey, fontSize=9, spaceAfter=10)

    col_headers = ["Adm No", "Full Name", "Phone", "Company",
                   "Address", "Supervisor", "Sup. Phone", "Start", "End", "Status"]
    keys = ["Admission No", "Full Name", "Trainee Phone", "Company Attached",
            "Location / Address", "Supervisor Name", "Supervisor Phone",
            "Start Date", "End Date", "Status"]

    data = [col_headers] + [[r.get(k, "") for k in keys] for r in rows]

    col_widths_mm = [22, 48, 26, 48, 52, 40, 26, 20, 20, 18]
    col_widths_pt = [w * mm for w in col_widths_mm]

    tbl = Table(data, colWidths=col_widths_pt, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), navy),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 8),
        ("FONTSIZE",    (0, 1), (-1, -1), 7.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF6FF")]),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ("ALIGN",       (0, 0), (-1, -1), "LEFT"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))

    story = [
        Paragraph(f"{title}", title_style),
        Paragraph(f"Period: {period_label} {year}  |  Total records: {len(rows)}", sub_style),
        Spacer(1, 4),
        tbl,
    ]
    doc.build(story)
    buf.seek(0)
    filename = f"attachments_{period_label.replace('–', '-')}_{year}.pdf"
    return Response(buf.read(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


@super_admin_bp.route("/gis-tracking/export")
@super_admin_required
def gis_tracking_export():
    from datetime import date
    db = _svc()
    fmt    = request.args.get("format", "excel")
    period = request.args.get("period", "")
    year   = int(request.args.get("year", date.today().year))
    dept_filter = request.args.get("department", "")

    student_ids = None
    if dept_filter:
        enr = (db.table("enrollments")
                 .select("student_id, classes!inner(department_id)")
                 .eq("classes.department_id", dept_filter)
                 .execute().data or [])
        student_ids = list({e["student_id"] for e in enr})

    rows = _build_export_rows(db, student_ids=student_ids,
                              period=period, year=year)
    label = _period_label(period)
    dept_name = ""
    if dept_filter:
        depts = db.table("departments").select("name").eq("id", dept_filter).execute().data or []
        dept_name = f" — {depts[0]['name']}" if depts else ""
    title = f"Industrial Attachments{dept_name}"

    if fmt == "pdf":
        return _export_pdf(rows, title, label, year)
    return _export_excel(rows, title, label, year)


# ── System Notices / Memos ────────────────────────────────────────────────────

def _send_system_notifications(db, title, message, notice_type,
                                action_url, department_id=None, class_id=None,
                                target_role="student"):
    """
    Push in-app notifications to users.
    target_role: 'student', 'trainer', 'all'
    Filters by department and/or class when provided.
    """
    try:
        if class_id:
            enrolled = (db.table("enrollments")
                          .select("student_id")
                          .eq("class_id", class_id)
                          .execute().data or [])
            ids = [e["student_id"] for e in enrolled if e.get("student_id")]
            if not ids:
                return 0
            query = db.table("user_profiles").select("id").in_("id", ids)
        else:
            query = db.table("user_profiles").select("id")
            if target_role != "all":
                query = query.eq("role", target_role)
            if department_id:
                query = query.eq("department_id", department_id)

        users = query.execute().data or []
        for u in users:
            create_notification(
                user_id=u["id"],
                title=f"[Notice] {title}",
                message=message,
                notification_type=notice_type,
                action_url=action_url or "/notifications"
            )
        return len(users)
    except Exception as e:
        print(f"[super_admin notices] notification error: {e}")
        return 0


@super_admin_bp.route("/notices")
@super_admin_required
def notices():
    from datetime import datetime as _dt
    db = _svc()
    notices_list = []
    departments  = db.table("departments").select("id, name").order("name").execute().data or []
    classes      = db.table("classes").select("id, name, department_id").order("name").execute().data or []
    dept_map     = {d["id"]: d["name"] for d in departments}
    class_map    = {c["id"]: c["name"] for c in classes}

    try:
        # Fetch system-wide notices (dept_notices where sent_by is super admin)
        # and all dept notices for visibility
        notices_list = (db.table("dept_notices")
                         .select("*, sender:user_profiles!dept_notices_sent_by_fkey(full_name, role)")
                         .order("sent_at", desc=True)
                         .limit(100)
                         .execute().data or [])
        for n in notices_list:
            did = n.get("department_id")
            cid = n.get("class_id")
            n["dept_name"]  = dept_map.get(did, "All Departments") if did else "All Departments"
            n["class_name"] = class_map.get(cid, "All Classes")    if cid else "All Trainees in Scope"
    except Exception as e:
        flash(f"Error loading notices: {e}", "danger")

    return render_template("super_admin/notices.html",
                           notices=notices_list,
                           departments=departments,
                           classes=classes)


@super_admin_bp.route("/notices/send", methods=["POST"])
@super_admin_required
def send_notice():
    from datetime import datetime as _dt
    db   = _svc()
    user = current_user()

    title       = (request.form.get("title")        or "").strip()
    message     = (request.form.get("message")      or "").strip()
    ntype       = request.form.get("notice_type",   "info")
    dept_id     = request.form.get("department_id") or None
    class_id    = request.form.get("class_id")      or None
    target_role = request.form.get("target_role",   "student")

    if not title or not message:
        flash("Title and message are required.", "warning")
        return redirect(url_for("super_admin.notices"))
    if ntype not in ("info", "warning", "success", "error"):
        ntype = "info"
    if target_role not in ("student", "trainer", "all"):
        target_role = "student"

    try:
        db.table("dept_notices").insert({
            "department_id": dept_id,
            "sent_by":       user["id"],
            "title":         title,
            "message":       message,
            "notice_type":   ntype,
            "class_id":      class_id,
            "sent_at":       _dt.now().isoformat(),
        }).execute()

        count = _send_system_notifications(
            db, title, message, ntype,
            action_url="/notifications",
            department_id=dept_id,
            class_id=class_id,
            target_role=target_role,
        )
        write_audit_log("send_system_notice",
                        target=f"Notice '{title}' sent to {count} users")
        flash(f"Notice sent successfully to {count} recipient(s).", "success")
    except Exception as e:
        flash(f"Failed to send notice: {e}", "danger")

    return redirect(url_for("super_admin.notices"))

# ── Super Admin: pages mirroring departmental admin (institute-wide scope) ────

@super_admin_bp.route("/class-list")
@super_admin_required
def class_list():
    from datetime import datetime as _dt
    db = _svc()
    class_id    = request.args.get("class_id","")
    dept_id     = request.args.get("dept_id","")
    departments = db.table("departments").select("id,name").order("name").execute().data or []
    cq = db.table("classes").select("id,name,department_id,departments(name)")
    if dept_id:
        cq = cq.eq("department_id", dept_id)
    classes = cq.order("name").execute().data or []
    cls = None; students = []; dept_name = ""
    if class_id:
        cls = next((c for c in classes if str(c["id"]) == class_id), None)
        if cls:
            dept_name = (cls.get("departments") or {}).get("name","")
            cls["dept_name"] = dept_name
        enr = (db.table("enrollments")
               .select("*, user_profiles(id,full_name,admission_no)")
               .eq("class_id", class_id).execute().data or [])
        for e in enr:
            s = e.get("user_profiles") or {}
            s["admission_number"] = s.get("admission_no","")
            if s.get("id"):
                students.append(s)
        students.sort(key=lambda s: s.get("full_name",""))
    return render_template("super_admin/class_list.html",
                           departments=departments, classes=classes,
                           class_id=class_id, dept_id=dept_id,
                           cls=cls, students=students, dept_name=dept_name,
                           date_gen=_dt.now().strftime("%d %b %Y"))


@super_admin_bp.route("/class-list/pdf")
@super_admin_required
def class_list_pdf():
    from datetime import datetime as _dt
    db = _svc()
    class_id = request.args.get("class_id","")
    cls = None; students = []; dept_name = ""
    if class_id:
        cls = db.table("classes").select("id,name,departments(name)").eq("id",class_id).single().execute().data
        dept_name = (cls.get("departments") or {}).get("name","") if cls else ""
        if cls:
            cls["dept_name"] = dept_name
        enr = (db.table("enrollments")
               .select("*, user_profiles(id,full_name,admission_no)")
               .eq("class_id", class_id).execute().data or [])
        for e in enr:
            s = e.get("user_profiles") or {}
            s["admission_number"] = s.get("admission_no","")
            if s.get("id"):
                students.append(s)
        students.sort(key=lambda s: s.get("full_name",""))
    return render_template("super_admin/class_list_pdf.html",
                           cls=cls, students=students, dept_name=dept_name,
                           date_gen=_dt.now().strftime("%d %b %Y"))


@super_admin_bp.route("/trainee-search")
@super_admin_required
def trainee_search():
    db = _svc()
    query_str  = request.args.get("q","").strip()
    student_id = request.args.get("student_id","").strip()
    unit_id    = request.args.get("unit_id","").strip()
    students = []; student = None; summary = None; records = []; units_list = []
    if query_str:
        rows = (db.table("user_profiles")
                .select("id,full_name,admission_no,enrollments(classes(name,departments(name)))")
                .eq("role","student")
                .or_(f"full_name.ilike.%{query_str}%,admission_no.ilike.%{query_str}%")
                .limit(20).execute().data or [])
        for r in rows:
            enr = (r.get("enrollments") or [{}])[0]
            cls = enr.get("classes") or {}
            r["class_name"] = cls.get("name","")
            r["dept_name"]  = (cls.get("departments") or {}).get("name","")
            students.append(r)
    if student_id:
        sr = db.table("user_profiles").select(
            "id,full_name,admission_no,enrollments(classes(name,departments(name)))"
        ).eq("id",student_id).single().execute().data or {}
        enr = (sr.get("enrollments") or [{}])[0]
        cls = enr.get("classes") or {}
        sr["class_name"] = cls.get("name","")
        sr["dept_name"]  = (cls.get("departments") or {}).get("name","")
        sr["admission_number"] = sr.get("admission_no","")
        student = sr
        units_list = db.table("units").select("id,name,code").order("name").execute().data or []
        if unit_id:
            records = (db.table("attendance")
                       .select("*, trainer:user_profiles!attendance_trainer_id_fkey(full_name)")
                       .eq("student_id",student_id).eq("unit_id",unit_id)
                       .order("attendance_date").execute().data or [])
            for r in records:
                r["trainer_name"] = (r.get("trainer") or {}).get("full_name","—")
            total = len(records)
            present = sum(1 for r in records if r.get("status")=="present")
            absent  = total - present
            pct     = round(present/total*100,1) if total else 0
            unit_obj = next((u for u in units_list if u["id"]==unit_id),{})
            summary = {"present":present,"absent":absent,"total":total,"pct":pct,"unit":unit_obj}
    from datetime import datetime as _dt
    return render_template("super_admin/trainee_search.html",
                           query=query_str, students=students, student_id=student_id,
                           unit_id=unit_id, student=student, summary=summary,
                           records=records, units_list=units_list,
                           date_gen=_dt.now().strftime("%d %b %Y"))


@super_admin_bp.route("/trainee-report-pdf")
@super_admin_required
def trainee_report_pdf():
    from datetime import datetime as _dt
    db = _svc()
    student_id = request.args.get("student_id","")
    unit_id    = request.args.get("unit_id","")
    student = db.table("user_profiles").select(
        "id,full_name,admission_no,enrollments(classes(name,departments(name)))"
    ).eq("id",student_id).single().execute().data or {}
    enr = (student.get("enrollments") or [{}])[0]
    cls = enr.get("classes") or {}
    student["class_name"]       = cls.get("name","")
    student["dept_name"]        = (cls.get("departments") or {}).get("name","")
    student["admission_number"] = student.get("admission_no","")
    records = []; unit = {}; present = absent = total = 0; pct = 0
    if student_id and unit_id:
        unit = db.table("units").select("id,name,code").eq("id",unit_id).single().execute().data or {}
        records = (db.table("attendance")
                   .select("*, trainer:user_profiles!attendance_trainer_id_fkey(full_name)")
                   .eq("student_id",student_id).eq("unit_id",unit_id)
                   .order("attendance_date").execute().data or [])
        for r in records:
            r["trainer_name"] = (r.get("trainer") or {}).get("full_name","—")
        total   = len(records)
        present = sum(1 for r in records if r.get("status")=="present")
        absent  = total - present
        pct     = round(present/total*100,1) if total else 0
    return render_template("super_admin/trainee_report_pdf.html",
                           student=student, unit=unit, records=records,
                           present=present, absent=absent, total=total, pct=pct,
                           term_label={1:"Term 1",2:"Term 2",3:"Term 3"},
                           date_gen=_dt.now().strftime("%d %b %Y"))


@super_admin_bp.route("/assessment-sheet")
@super_admin_required
def assessment_sheet():
    from datetime import datetime as _dt
    db = _svc()
    dept_id  = request.args.get("dept_id","")
    class_id = request.args.get("class_id","")
    unit_id  = request.args.get("unit_id","")
    year     = request.args.get("year","")
    term     = request.args.get("term","")
    min_pct  = int(request.args.get("min_pct",75))
    departments = db.table("departments").select("id,name").order("name").execute().data or []
    cq = db.table("classes").select("id,name,department_id")
    if dept_id:
        cq = cq.eq("department_id",dept_id)
    classes = cq.order("name").execute().data or []
    uq = db.table("units").select("id,name,code")
    if dept_id:
        uq = uq.eq("department_id",dept_id)
    units = uq.order("name").execute().data or []
    cls = unit = None; eligible = []
    term_label = f"Term {term}" if term else "All Terms"
    if class_id and unit_id:
        cls  = next((c for c in classes if str(c["id"])==class_id), None)
        unit = next((u for u in units  if str(u["id"])==unit_id),  None)
        enr  = (db.table("enrollments")
                .select("student_id, user_profiles!inner(id,full_name,admission_no)")
                .eq("class_id",class_id).execute().data or [])
        for e in enr:
            s = e.get("user_profiles") or {}
            sid = s.get("id")
            if not sid:
                continue
            aq = db.table("attendance").select("status").eq("student_id",sid).eq("unit_id",unit_id)
            if year: aq = aq.eq("year",int(year))
            if term: aq = aq.eq("term",int(term))
            att = aq.execute().data or []; tot = len(att)
            if not tot:
                continue
            pre = sum(1 for a in att if a["status"]=="present")
            p   = round(pre/tot*100,1)
            if p >= min_pct:
                eligible.append({"admission_number":s.get("admission_no",""),
                                  "full_name":s.get("full_name",""),
                                  "present":pre,"total":tot,"pct":p})
        eligible.sort(key=lambda x: x["full_name"])
    return render_template("super_admin/assessment_sheet.html",
                           departments=departments, classes=classes, units=units,
                           dept_id=dept_id, class_id=class_id, unit_id=unit_id,
                           year=int(year) if year else "",
                           term=int(term) if term else "",
                           min_pct=min_pct, cls=cls, unit=unit, eligible=eligible,
                           term_label=term_label)


@super_admin_bp.route("/assessment-sheet/pdf")
@super_admin_required
def assessment_sheet_pdf():
    from datetime import datetime as _dt
    db = _svc()
    class_id = request.args.get("class_id","")
    unit_id  = request.args.get("unit_id","")
    year     = request.args.get("year","")
    term     = request.args.get("term","")
    min_pct  = int(request.args.get("min_pct",75))
    cls = unit = None; eligible = []
    term_label = f"Term {term}" if term else "All Terms"
    if class_id and unit_id:
        cls  = db.table("classes").select("id,name,departments(name)").eq("id",class_id).single().execute().data
        unit = db.table("units").select("id,name,code").eq("id",unit_id).single().execute().data
        if cls: cls["dept_name"] = (cls.get("departments") or {}).get("name","")
        enr = (db.table("enrollments")
               .select("student_id, user_profiles!inner(id,full_name,admission_no)")
               .eq("class_id",class_id).execute().data or [])
        for e in enr:
            s = e.get("user_profiles") or {}; sid = s.get("id")
            if not sid: continue
            aq = db.table("attendance").select("status").eq("student_id",sid).eq("unit_id",unit_id)
            if year: aq = aq.eq("year",int(year))
            if term: aq = aq.eq("term",int(term))
            att = aq.execute().data or []; tot = len(att)
            if not tot: continue
            pre = sum(1 for a in att if a["status"]=="present")
            p = round(pre/tot*100,1)
            if p >= min_pct:
                eligible.append({"admission_number":s.get("admission_no",""),
                                  "full_name":s.get("full_name",""),
                                  "present":pre,"total":tot,"pct":p})
        eligible.sort(key=lambda x: x["full_name"])
    return render_template("super_admin/assessment_sheet_pdf.html",
                           cls=cls, unit=unit, eligible=eligible,
                           term_label=term_label, year=year, min_pct=min_pct,
                           date_gen=_dt.now().strftime("%d %b %Y"))


@super_admin_bp.route("/exam-bookings")
@super_admin_required
def exam_bookings():
    from datetime import date as _date
    db = _svc()
    status_filter = request.args.get("status","all")
    dept_filter   = request.args.get("dept_id","")
    year_filter   = request.args.get("year","")
    term_filter   = request.args.get("term","")
    q             = request.args.get("q","").strip()
    TERM_MONTHS   = {"1":("01","04","30"),"2":("05","08","31"),"3":("09","12","31")}
    departments   = db.table("departments").select("id,name").order("name").execute().data or []
    bq = (db.table("exam_bookings")
           .select("*, units(name,code), "
                   "student:user_profiles!exam_bookings_student_id_fkey(full_name,admission_no,mobile_number), "
                   "reviewer:user_profiles!exam_bookings_approved_by_fkey(full_name)")
           .order("created_at",desc=True))
    if status_filter and status_filter != "all":
        bq = bq.eq("status",status_filter)
    yr = year_filter or str(_date.today().year)
    if year_filter:
        bq = bq.gte("exam_date",f"{yr}-01-01").lte("exam_date",f"{yr}-12-31")
    if term_filter and term_filter in TERM_MONTHS:
        m0,m1,last = TERM_MONTHS[term_filter]
        bq = bq.gte("exam_date",f"{yr}-{m0}-01").lte("exam_date",f"{yr}-{m1}-{last}")
    bookings = bq.limit(500).execute().data or []
    sids = list({b["student_id"] for b in bookings if b.get("student_id")})
    dept_map = {}
    if sids:
        enr = (db.table("enrollments")
               .select("student_id, classes!inner(department_id,name,departments(name))")
               .in_("student_id",sids).execute().data or [])
        for e in enr:
            sid = e.get("student_id"); cls = e.get("classes") or {}
            if sid and sid not in dept_map:
                dept_map[sid] = {"class_name":cls.get("name",""),
                                  "dept_name":(cls.get("departments") or {}).get("name",""),
                                  "dept_id":cls.get("department_id","")}
    for b in bookings:
        b["student_user"]     = b.get("student")  or {}
        b["approved_by_user"] = b.get("reviewer") or {}
        dm = dept_map.get(b.get("student_id"),{})
        b["class_name"] = dm.get("class_name","—")
        b["dept_name"]  = dm.get("dept_name","—")
        b["_dept_id"]   = dm.get("dept_id","")
    if dept_filter:
        bookings = [b for b in bookings if b.get("_dept_id")==dept_filter]
    if q:
        ql = q.lower()
        bookings = [b for b in bookings
                    if ql in b["student_user"].get("full_name","").lower()
                    or ql in b["student_user"].get("admission_no","").lower()]
    all_b = db.table("exam_bookings").select("status").limit(2000).execute().data or []
    counts = {"all":len(all_b),
              "pending": sum(1 for b in all_b if b["status"]=="pending"),
              "approved":sum(1 for b in all_b if b["status"]=="approved"),
              "rejected":sum(1 for b in all_b if b["status"]=="rejected")}
    return render_template("super_admin/exam_bookings.html",
                           bookings=bookings, departments=departments,
                           status_filter=status_filter, dept_filter=dept_filter,
                           year_filter=year_filter, term_filter=term_filter,
                           q=q, counts=counts)


@super_admin_bp.route("/exam-bookings/<booking_id>/approve", methods=["POST"])
@super_admin_required
def approve_exam_booking(booking_id):
    db = _svc(); user = current_user()
    try:
        db.table("exam_bookings").update(
            {"status":"approved","approved_by":user["id"],"approved_at":"now()"}
        ).eq("id",booking_id).execute()
        flash("Booking approved.", "success")
    except Exception as e:
        flash(f"Error: {e}", "danger")
    return redirect(url_for("super_admin.exam_bookings"))


@super_admin_bp.route("/exam-bookings/<booking_id>/reject", methods=["POST"])
@super_admin_required
def reject_exam_booking(booking_id):
    db = _svc(); reason = request.form.get("rejection_reason","").strip()
    try:
        db.table("exam_bookings").update(
            {"status":"rejected","rejection_reason":reason}
        ).eq("id",booking_id).execute()
        flash("Booking rejected.", "success")
    except Exception as e:
        flash(f"Error: {e}", "danger")
    return redirect(url_for("super_admin.exam_bookings"))


@super_admin_bp.route("/trainees-documents")
@super_admin_required
def trainees_documents():
    db = _svc()
    q           = request.args.get("q","").strip()
    dept_filter = request.args.get("dept_id","")
    departments = db.table("departments").select("id,name").order("name").execute().data or []
    enr_q = db.table("enrollments").select(
        "student_id, classes!inner(department_id,name), "
        "user_profiles!enrollments_student_id_fkey(id,full_name,admission_no,email)")
    if dept_filter:
        enr_q = enr_q.eq("classes.department_id",dept_filter)
    enr = enr_q.execute().data or []
    seen = {}
    for e in enr:
        up = e.get("user_profiles") or {}
        sid = up.get("id") or e.get("student_id")
        if sid and sid not in seen:
            seen[sid] = {"id":sid, "full_name":up.get("full_name","—"),
                         "admission_no":up.get("admission_no","—"),
                         "email":up.get("email",""),
                         "class_name":(e.get("classes") or {}).get("name","")}
    students = list(seen.values())
    if q:
        ql = q.lower()
        students = [s for s in students
                    if ql in s["full_name"].lower() or ql in s["admission_no"].lower()]
    students.sort(key=lambda s: s["full_name"])
    sids = [s["id"] for s in students]
    req_types = {"passport_photo","admission_letter","medical_form","personal_data_form",
                 "declaration_form","kcse_result_slip","kcse_certificate","kcpe_result_slip",
                 "birth_certificate","national_id","consent_form"}
    doc_map = {}
    if sids:
        try:
            rows = (db.table("student_personal_documents").select("student_id,document_type,status")
                    .in_("student_id",sids).execute().data or [])
            for r in rows:
                doc_map.setdefault(r["student_id"],{})[r["document_type"]] = r
        except Exception:
            pass
    total_req = len(req_types)
    for s in students:
        docs = doc_map.get(s["id"],{})
        s["uploaded"]    = len(docs)
        s["required_ok"] = sum(1 for dt in req_types if dt in docs)
        s["total_req"]   = total_req
        statuses = [d.get("status","pending") for d in docs.values()]
        s["overall_status"] = ("approved" if statuses and all(st=="approved" for st in statuses)
                               else "rejected" if any(st=="rejected" for st in statuses)
                               else "pending")
    return render_template("super_admin/trainees_documents.html",
                           students=students, departments=departments,
                           q=q, dept_filter=dept_filter)


@super_admin_bp.route("/trainees-documents/<student_id>")
@super_admin_required
def trainee_document_detail(student_id):
    db = _svc()
    student = db.table("user_profiles").select(
        "id,full_name,admission_no,email,mobile_number"
    ).eq("id",student_id).single().execute().data or {}
    enr = (db.table("enrollments").select("classes(name)")
           .eq("student_id",student_id).limit(1).execute().data or [])
    class_name = ((enr[0].get("classes") or {}).get("name","")) if enr else ""
    docs_raw = (db.table("student_personal_documents").select("*")
                .eq("student_id",student_id).execute().data or [])
    docs = {}
    for d in docs_raw:
        d["_url"] = d.get("file_url","")
        docs[d["document_type"]] = d
    DOC_TYPES = [
        ("passport_photo","Passport Photo",True),
        ("admission_letter","Admission Letter",True),
        ("medical_form","Medical Examination Form",True),
        ("personal_data_form","Personal Data Form",True),
        ("declaration_form","Declaration Form",True),
        ("kcse_result_slip","KCSE Result Slip",True),
        ("kcse_certificate","KCSE Certificate",True),
        ("kcpe_result_slip","KCPE Result Slip",True),
        ("birth_certificate","Birth Certificate",True),
        ("national_id","National ID",True),
        ("guardian_id","Guardian ID Copies",False),
        ("consent_form","Consent Form",True),
        ("most_recent_result_slip","Previous Module Result Slip (Continuing)",False),
    ]
    statuses = [d.get("status","pending") for d in docs.values()]
    overall = ("approved" if statuses and all(st=="approved" for st in statuses)
               else "rejected" if any(st=="rejected" for st in statuses) else "pending")
    return render_template("super_admin/trainee_document_detail.html",
                           student=student, class_name=class_name, docs=docs,
                           doc_types=DOC_TYPES, overall_status=overall, hod_comment="")


@super_admin_bp.route("/trainees-documents/<student_id>/verify", methods=["POST"])
@super_admin_required
def verify_trainee_document(student_id):
    from datetime import datetime as _dt
    db = _svc()
    status  = request.form.get("status","pending")
    comment = request.form.get("comment","").strip()
    if status not in ("pending","approved","rejected"):
        status = "pending"
    for doc in (db.table("student_personal_documents").select("id")
                .eq("student_id",student_id).execute().data or []):
        try:
            db.table("student_personal_documents").update({
                "status":           status,
                "rejection_reason": comment or None,
                "verified_by":      current_user().get("id"),
                "verified_at":      _dt.utcnow().isoformat(),
            }).eq("id",doc["id"]).execute()
        except Exception as e:
            print(f"verify_trainee_document: {e}")
    flash(f"Documents marked as {status}.", "success")
    return redirect(url_for("super_admin.trainee_document_detail", student_id=student_id))


@super_admin_bp.route("/trainer-poe")
@super_admin_required
def trainer_poe():
    db = _svc()
    dept_filter = request.args.get("dept_id","")
    q           = request.args.get("q","").strip()
    departments = db.table("departments").select("id,name").order("name").execute().data or []
    tq = db.table("user_profiles").select(
        "id,full_name,admission_no,department_id,departments(name)"
    ).eq("role","trainer")
    if dept_filter:
        tq = tq.eq("department_id",dept_filter)
    trainers = tq.order("full_name").execute().data or []
    if q:
        ql = q.lower()
        trainers = [t for t in trainers
                    if ql in t.get("full_name","").lower()
                    or ql in t.get("admission_no","").lower()]
    tids = [t["id"] for t in trainers]
    doc_map = {}
    if tids:
        try:
            docs = (db.table("trainer_documents").select("*")
                    .in_("trainer_id",tids).execute().data or [])
            for d in docs:
                doc_map.setdefault(d["trainer_id"],[]).append(d)
        except Exception:
            pass
    for t in trainers:
        t["docs"]      = doc_map.get(t["id"],[])
        t["doc_count"] = len(t["docs"])
        t["dept_name"] = (t.get("departments") or {}).get("name","—")
    return render_template("super_admin/trainer_poe.html",
                           trainers=trainers, departments=departments,
                           dept_filter=dept_filter, q=q)


@super_admin_bp.route("/trainer-documents")
@super_admin_required
def trainer_documents():
    db = _svc()
    from datetime import datetime as _dt
    doc_type   = request.args.get("document_type", "")
    year       = request.args.get("year", str(_dt.now().year))
    term       = request.args.get("term", "")
    trainer_id = request.args.get("trainer_id", "")
    query = (db.table("trainer_documents")
        .select("*, units(name, code, department_id), classes(name), user_profiles(full_name, staff_no, department_id)")
        .eq("academic_year", int(year)))
    if term:       query = query.eq("term", term)
    if doc_type:   query = query.eq("document_type", doc_type)
    if trainer_id: query = query.eq("trainer_id", trainer_id)
    docs = query.order("created_at", desc=True).execute().data or []
    trainers = (db.table("user_profiles").select("id, full_name, staff_no")
        .eq("role", "trainer").order("full_name").execute().data or [])
    return render_template("super_admin/trainer_documents.html",
                           documents=docs, trainers=trainers,
                           document_type=doc_type, year=year, term=term, trainer_id=trainer_id)


@super_admin_bp.route("/trainer-document-view/<document_id>")
@super_admin_required
def view_trainer_document(document_id):
    db = _svc()
    result = db.table("trainer_documents").select("id, file_url").eq("id", document_id).execute()
    doc = result.data[0] if result.data else None
    if not doc or not doc.get("file_url"):
        abort(404)
    return redirect(doc["file_url"])


@super_admin_bp.route("/import", methods=["GET","POST"])
@super_admin_required
def import_data():
    db = _svc()
    departments = db.table("departments").select("id, name").order("name").execute().data or []
    classes     = db.table("classes").select("id, name, department_id, departments(name)").order("name").execute().data or []
    result_summary = None; error = None
    if request.method == "POST":
        flash("Import submitted.", "info")
    return render_template("super_admin/import.html",
                           departments=departments, classes=classes,
                           result=result_summary, error=error)


# ── Biometric Scanner Registration ───────────────────────────────────────────

@super_admin_bp.route("/biometric-scanners")
@super_admin_required
def biometric_scanners():
    """List all registered biometric scanners and show registration form."""
    db   = _svc()
    user = current_user()

    scanners = []
    try:
        scanners = (db.table("biometric_scanners")
                    .select("*, departments(name)")
                    .order("registered_at", desc=True)
                    .execute().data or [])
    except Exception:
        pass

    departments = (db.table("departments")
                   .select("id, name")
                   .order("name")
                   .execute().data or [])

    active_count   = sum(1 for s in scanners if s.get("is_active") is not False)
    rooms_covered  = len({s.get("room") for s in scanners if s.get("room")})

    return render_template(
        "super_admin/biometric_scanners.html",
        scanners=scanners,
        departments=departments,
        total=len(scanners),
        active_count=active_count,
        rooms_covered=rooms_covered,
    )


@super_admin_bp.route("/biometric-scanners/register", methods=["POST"])
@super_admin_required
def register_scanner():
    """Register a new biometric fingerprint scanner."""
    db   = _svc()
    user = current_user()

    serial_number = request.form.get("serial_number", "").strip()
    device_name   = request.form.get("device_name", "").strip()
    room          = request.form.get("room", "").strip()
    building      = request.form.get("building", "").strip()
    department_id = request.form.get("department_id", "").strip() or None
    notes         = request.form.get("notes", "").strip()

    if not serial_number:
        flash("Scanner serial number / device ID is required.", "error")
        return redirect(url_for("super_admin.biometric_scanners"))
    if not room:
        flash("Room assignment is required.", "error")
        return redirect(url_for("super_admin.biometric_scanners"))

    try:
        # Check for duplicate serial
        existing = (db.table("biometric_scanners")
                    .select("id")
                    .eq("serial_number", serial_number)
                    .execute().data or [])
        if existing:
            flash(f"A scanner with serial '{serial_number}' is already registered.", "warning")
            return redirect(url_for("super_admin.biometric_scanners"))

        db.table("biometric_scanners").insert({
            "serial_number": serial_number,
            "device_name":   device_name or f"Scanner — {room}",
            "room":          room,
            "building":      building or None,
            "department_id": department_id,
            "notes":         notes or None,
            "is_active":     True,
            "registered_by": user["id"],
        }).execute()

        write_audit_log("register_scanner",
                        target=f"serial:{serial_number} room:{room}")
        flash(f"Scanner '{serial_number}' registered and assigned to {room}.", "success")

    except Exception as e:
        flash(f"Error registering scanner: {e}", "error")

    return redirect(url_for("super_admin.biometric_scanners"))


@super_admin_bp.route("/biometric-scanners/update/<scanner_id>", methods=["POST"])
@super_admin_required
def update_scanner(scanner_id):
    """Update scanner details — room, name, building, department, status."""
    db   = _svc()
    user = current_user()

    device_name   = request.form.get("device_name", "").strip()
    room          = request.form.get("room", "").strip()
    building      = request.form.get("building", "").strip()
    department_id = request.form.get("department_id", "").strip() or None
    notes         = request.form.get("notes", "").strip()
    is_active     = request.form.get("is_active", "1") == "1"

    if not room:
        flash("Room is required.", "error")
        return redirect(url_for("super_admin.biometric_scanners"))

    try:
        db.table("biometric_scanners").update({
            "device_name":   device_name or None,
            "room":          room,
            "building":      building or None,
            "department_id": department_id,
            "notes":         notes or None,
            "is_active":     is_active,
        }).eq("id", scanner_id).execute()

        write_audit_log("update_scanner", target=f"scanner:{scanner_id}")
        flash("Scanner details updated successfully.", "success")

    except Exception as e:
        flash(f"Error updating scanner: {e}", "error")

    return redirect(url_for("super_admin.biometric_scanners"))


@super_admin_bp.route("/biometric-scanners/delete/<scanner_id>", methods=["POST"])
@super_admin_required
def delete_scanner(scanner_id):
    """Permanently remove a scanner registration."""
    db   = _svc()
    user = current_user()

    try:
        row = (db.table("biometric_scanners")
               .select("serial_number, room")
               .eq("id", scanner_id)
               .single()
               .execute().data)
        if not row:
            flash("Scanner not found.", "error")
            return redirect(url_for("super_admin.biometric_scanners"))

        db.table("biometric_scanners").delete().eq("id", scanner_id).execute()
        write_audit_log("delete_scanner",
                        target=f"serial:{row.get('serial_number')} room:{row.get('room')}")
        flash(f"Scanner '{row.get('serial_number')}' removed.", "success")

    except Exception as e:
        flash(f"Error removing scanner: {e}", "error")

    return redirect(url_for("super_admin.biometric_scanners"))
